class GameState:
    def __init__(self):
        self.career_high_score = 0
        self.temp_high_score_achievement_unlocked = False
        self.current_match_highscore = 0
        self.player_runs_1stinn = 0
        self.ai_runs_1stinn = 0
        self.player_runs_2ndinn = 0
        self.ai_runs_2ndinn = 0
        self.player_wickets_1stinn = 0
        self.ai_wickets_1stinn = 0
        self.player_wickets_2ndinn = 0
        self.ai_wickets_2ndinn = 0
        self.player_runs_conceded_1stinn = 0
        self.ai_runs_conceded_1stinn = 0
        self.player_runs_conceded_2ndinn = 0
        self.ai_runs_conceded_2ndinn = 0
        self.player_balls_bowled_1stinn = 0
        self.ai_balls_bowled_1stinn = 0
        self.player_balls_bowled_2ndinn = 0
        self.ai_balls_bowled_2ndinn = 0
        self.balls_played_by_player = 0
        self.balls_played_by_ai = 0
        self.balls_played_first_inn = 0
        self.balls_played_sec_inn = 0
        self.current_match_wickets = 0
        self.match_result = ''
        self.lifetime_runs = 0
        self.lifetime_wickets = 0
        self.lifetime_balls_faced = 0
        self.lifetime_balls_bowled = 0
        self.lifetime_runs_conceded = 0
        self.half_centuries = 0
        self.centuries = 0
        self.total_matches_played = 0
        self.total_wins = 0
        self.total_losses = 0
        self.total_draws = 0
        self.average_runs_per_match = 0
        self.player_profile = {}
        self.match_summary = {}
        self.difficulty = 'medium'
        self.commentary_enabled = True
        self.toss_result = ''
        self.over = 0
        self.balls = 0
        self.player_name = ''
        self.total_sixes = 0
    def update(self, data: dict):
        if isinstance(data, GameState):
            data = data.to_dict()
        for key, value in data.items():
            if hasattr(self, key):
                setattr(self, key, value)
    
    def to_dict(self):
        result = {}
        for k, v in self.__dict__.items():
            if isinstance(v, dict):
                result[k] = v.copy()
            else:
                result[k] = v
        return result
    todict = to_dict 
    
    def from_dict(self, data: dict):
        if data is None:
            return
        if isinstance(data, GameState):
            data = data.to_dict()
        for key, value in data.items():
            if hasattr(self, key):
                setattr(self, key, value)
    fromdict = from_dict

    def reset_lifetime(self):
        self.lifetime_runs = 0
        self.lifetime_wickets = 0
        self.total_matches_played = 0
        self.total_wins = 0
        self.total_losses = 0
        self.total_draws = 0
        self.average_runs_per_match = 0
        self.lifetime_balls_faced = 0
        self.lifetime_balls_bowled = 0
        self.lifetime_runs_conceded = 0
        self.centuries = 0
        self.half_centuries = 0
        self.total_sixes = 0
        self.current_match_highscore = 0
        self.temp_high_score_achievement_unlocked = False
class HandCricketDB:
    def __init__(self, config):
        self.DB_CONFIG = config
        self._db = None
        self.current_user = None

    def create_database(self):
        try:
            sql_server = mysql.connector.connect(
                host=self.DB_CONFIG["host"],
                user=self.DB_CONFIG["user"],
                password=self.DB_CONFIG["password"],
            )
            cursor = sql_server.cursor()
            cursor.execute(f"CREATE DATABASE IF NOT EXISTS {self.DB_CONFIG['database']}")                                                                                                                                                                                                           
            cursor.close()
            sql_server.close()
            console.print(f"[bold yellow]Database[/bold yellow] [bold red]'{self.DB_CONFIG['database']}'[/bold red] [bold yellow]has been created/loaded successfully.[/bold yellow]")
        except mysql.connector.Error as e:
            console.print(f"[grey37]Database could not be created due to error:[/grey37] [bold red]{e}[/bold red]")
            self.achievement_manager.save_achievements()
            sys.exit(1)

    def get_db(self):
        try:
            if self._db is None or not self._db.is_connected():
                if self._db:
                    self._db.close()
                self._db = mysql.connector.connect(**self.DB_CONFIG, buffered = True)
            return self._db
        except mysql.connector.Error as error:
            console.print(f"[grey37]Database connection error:[/grey37] [bold red]{error}[/bold red]")
            return None

    def init_db(self):
        ddl_profile = """
            CREATE TABLE IF NOT EXISTS player_profile ( 
                id               INT AUTO_INCREMENT PRIMARY KEY, 
                name             VARCHAR(100) UNIQUE NOT NULL,
                password         VARCHAR(100) NOT NULL, 
                lifetime_runs    INT  DEFAULT 0, 
                lifetime_wickets INT  DEFAULT 0, 
                total_matches    INT  DEFAULT 0, 
                total_wins       INT  DEFAULT 0, 
                total_losses     INT  DEFAULT 0, 
                total_draws      INT  DEFAULT 0,
                lifetime_balls_faced INT  DEFAULT 0,
                lifetime_balls_bowled INT  DEFAULT 0,
                lifetime_runs_conceded INT  DEFAULT 0,
                centuries        INT  DEFAULT 0,
                half_centuries   INT  DEFAULT 0,
                avg_runs         FLOAT  DEFAULT 0.0
            )
            """
        ddl_match_data = """
            CREATE TABLE IF NOT EXISTS match_data (
                id INT AUTO_INCREMENT PRIMARY KEY,
                player_name VARCHAR(100) NOT NULL,
                runs INT  DEFAULT 0,
                wickets INT  DEFAULT 0,
                balls_faced INT  DEFAULT 0,
                player_balls_bowled INT  DEFAULT 0,
                player_runs_conceded INT  DEFAULT 0,
                result VARCHAR(10),
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_player_name (player_name)
            )
            """
        
        ddl_achievements = """
            CREATE TABLE IF NOT EXISTS achievements (
                id INT AUTO_INCREMENT PRIMARY KEY,
                player_name VARCHAR(100) NOT NULL,
                achievement VARCHAR(100) NOT NULL,
                achieved_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY unique_achievement (player_name, achievement),
                INDEX idx_player_name (player_name)
            )
            """
        
        db = self.get_db()
        if db:
            cur = db.cursor()
            try:
                cur.execute(ddl_profile)
                cur.execute(ddl_match_data)
                cur.execute(ddl_achievements)
                db.commit()
            finally:
                cur.close()

    def get_password(self, player_name):
        db = self.get_db()
        cursor = db.cursor()
        cursor.execute("SELECT password FROM player_profile WHERE name = %s", (player_name,))
        row = cursor.fetchone()
        cursor.close()
        return row[0] if row else None

    def update_password(self, player_name, new_password):
        db = self.get_db()
        cursor = db.cursor()
        cursor.execute("UPDATE player_profile SET password = %s WHERE name = %s", (new_password, player_name))
        db.commit()
        cursor.close()

    def save_profile(self, profile: dict, player_name: str = "", password: str = None):
        if not player_name:
            raise ValueError("player name must be provided")

        sql = """
              INSERT INTO player_profile
              (name, password, lifetime_runs, lifetime_wickets, total_matches,
               total_wins, total_losses, total_draws, avg_runs, lifetime_balls_faced, lifetime_balls_bowled, lifetime_runs_conceded,
               centuries, half_centuries)
              VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) 
              ON DUPLICATE KEY UPDATE
                password = VALUES(password),
                lifetime_runs = VALUES(lifetime_runs),
                lifetime_wickets = VALUES(lifetime_wickets),
                total_matches = VALUES(total_matches),
                total_wins = VALUES(total_wins),
                total_losses = VALUES(total_losses),
                total_draws = VALUES(total_draws),
                avg_runs = VALUES(avg_runs),
                lifetime_balls_faced = VALUES(lifetime_balls_faced),
                lifetime_balls_bowled = VALUES(lifetime_balls_bowled),
                lifetime_runs_conceded = VALUES(lifetime_runs_conceded),
                centuries = VALUES(centuries),
                half_centuries = VALUES(half_centuries)
              """
        vals = (
            player_name,
            password,
            profile["Lifetime runs"],
            profile["Lifetime wickets"],
            profile["Total Matches Played"],
            profile["Total Wins"],
            profile["Total Losses"],
            profile["Total Draws"],
            profile["Average Runs per Match"],
            profile.get("Lifetime Balls Faced", 0), 
            profile.get("Lifetime Balls Bowled", 0), 
            profile.get("Lifetime Runs Conceded", 0),
            profile.get("Centuries", 0),            
            profile.get("Half Centuries", 0), 
        )

        achievements = profile.get("Achievements", [])
        if achievements:
            ach_sql = "INSERT IGNORE INTO achievements (player_name, achievement) VALUES (%s, %s)"
            ach_vals = [(player_name, ach) for ach in achievements]
            try:
                db = self.get_db()
                ach_cur = db.cursor()
                ach_cur.executemany(ach_sql, ach_vals)
                db.commit()
            except Exception as e:
                console.print(f"[bold red]Failed to save achievements:[/bold red] {e}")
            finally:
                try:
                    if ach_cur:
                        ach_cur.close()
                except:
                    pass
        cur = None
        try:
            db = self.get_db()
            cur = db.cursor()
            cur.execute(sql, vals)
            db.commit()
        
        except KeyError as e:
            console.print(f"[grey37]Missing required field in profile:[/grey37] [bold red]{e}[/bold red]")
        except Exception as e:
            console.print(f"[bold red]Failed to save profile:[/bold red] {e}")
            try:
                self.get_db().rollback()
            except:
                pass
        finally:
            try:
                if cur:
                    cur.close()
            except:
                pass

    def save_match_data(self, game_state, player_name: str, player_runs_conceded: int, player_balls_bowled: int, player_match_wickets: int):
        try:
            db = self.get_db()
            cursor = db.cursor()
            sql = """
                INSERT INTO match_data (player_name, runs, wickets, balls_faced, player_runs_conceded, player_balls_bowled, result)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            if hasattr(game_state, 'to_dict'):
                state = game_state.to_dict()

            match_runs = state.get('player_runs_1stinn', 0) + state.get('player_runs_2ndinn', 0)
            match_wickets = player_match_wickets
            match_balls_faced = state.get("balls_played_by_player", 0)

            values = (
                player_name,
                match_runs,
                match_wickets,
                match_balls_faced,
                player_runs_conceded,
                player_balls_bowled,
                state.get("match_result", "unknown"),
            )
            cursor.execute(sql, values)
            db.commit()
            cursor.close()
        except Exception as e:
            console.print(f"[bold red]Error saving match data:[/bold red] {e}")
            try:
                if cursor:
                    cursor.close()
            except:
                pass

    def load_profile(self, player_name):
        try:
            db = self.get_db()
            cursor = db.cursor()
            cursor.execute("""
                SELECT lifetime_runs, lifetime_wickets, total_matches, 
                    total_wins, total_losses, total_draws, avg_runs,
                    lifetime_balls_faced, lifetime_balls_bowled, centuries, half_centuries, lifetime_runs_conceded
                FROM player_profile WHERE name = %s
            """, (player_name,))
            row = cursor.fetchone()
            cursor.close()
            if row:
                return {
                    "lifetime_runs": row[0],
                    "lifetime_wickets": row[1],
                    "total_matches_played": row[2],
                    "total_wins": row[3],
                    "total_losses": row[4],
                    "total_draws": row[5],
                    "average_runs_per_match": row[6],
                    "lifetime_balls_faced": row[7],   
                    "lifetime_balls_bowled": row[8],  
                    "centuries": row[9],              
                    "half_centuries": row[10],
                    "lifetime_runs_conceded": row[11]
                }
        except Exception as e:
            console.print(f"[bold red]Error loading profile from database:[/bold red] {e}")
        return None
    
    def player_exists(self, player_name: str):
        try:
            db = self.get_db()
            cursor = db.cursor()
            cursor.execute("SELECT 1 FROM player_profile WHERE name = %s", (player_name,))
            result = cursor.fetchone()
            cursor.close()
            return result is not None
        except Exception as e:
            console.print(f"[bold red]Error checking player existence:[/bold red] {e}")
            return False

    def close_db(self):
        try:
            if self._db:
                self._db.close()
                self._db = None
        except:
            pass
    
    def get_achievements(self, player_name: str):
        try:
            db = self.get_db()
            cursor = db.cursor()
            cursor.execute("SELECT achievement FROM achievements WHERE player_name = %s", (player_name,))
            rows = cursor.fetchall()
            cursor.close()
            return [row[0] for row in rows] if rows else []
        except Exception as e:
            console.print(f"[bold red]Error fetching achievements from DB:[/bold red] {e}")
            return []
    
    def update_achievements(self, player_name: str, achievements: list):
        try:
            if not achievements:
                return
            db = self.get_db()
            cursor = db.cursor()
            sql = "INSERT IGNORE INTO achievements (player_name, achievement) VALUES (%s, %s)"
            vals = [(player_name, ach) for ach in achievements]
            cursor.executemany(sql, vals)
            db.commit()
            cursor.close()
        except Exception as e:
            console.print(f"[bold red]Error updating achievements in DB:[/bold red] {e}")
            try:
                if cursor:
                    cursor.close()
            except:
                pass
    
    def get_leaderboard_leader(self, category: str):
        cursor = self.get_db().cursor()
        if category == "runs":
            sql = "SELECT player_name FROM match_data GROUP BY player_name ORDER BY SUM(runs) DESC LIMIT 1"
        elif category == "wickets":
            sql = "SELECT player_name FROM match_data GROUP BY player_name ORDER BY SUM(wickets) DESC LIMIT 1"
        elif category == "balls":
            sql = "SELECT player_name FROM match_data GROUP BY player_name ORDER BY SUM(balls_faced) DESC LIMIT 1"
        elif category == "win_rate":
            sql = """
            SELECT player_name
            FROM (
              SELECT player_name,
                     SUM(result='win')/COUNT(*) AS rate
              FROM match_data
              GROUP BY player_name
            ) AS t
            ORDER BY rate DESC
            LIMIT 1
            """
        elif category == "high_score":
            sql = "SELECT player_name FROM match_data GROUP BY player_name ORDER BY MAX(runs) DESC LIMIT 1"
        elif category == "lifetime_run_rate":
            sql = """
            SELECT name AS player_name
            FROM player_profile
            WHERE lifetime_balls_faced>0
            ORDER BY (lifetime_runs*6.0)/lifetime_balls_faced DESC
            LIMIT 1
            """
        elif category == "lifetime_strike_rate":
            sql = """
            SELECT name AS player_name
            FROM player_profile
            WHERE lifetime_balls_faced>0
            ORDER BY (lifetime_runs*100.0)/lifetime_balls_faced DESC
            LIMIT 1
            """
        elif category == "lifetime_economy_rate":
            sql = """
            SELECT name AS player_name
            FROM player_profile
            WHERE lifetime_balls_bowled>0
            ORDER BY (lifetime_runs_conceded*6.0)/lifetime_balls_bowled ASC
            LIMIT 1
            """
        elif category == "half_centuries":
            sql = "SELECT name AS player_name FROM player_profile ORDER BY half_centuries DESC LIMIT 1"
        elif category == "centuries":
            sql = "SELECT name AS player_name FROM player_profile ORDER BY centuries DESC LIMIT 1"
        else:
            return None
        cursor.execute(sql)
        row = cursor.fetchone()
        cursor.close()
        return row[0] if row else None
class GameDataManager:
    def __init__(self, json_folder_path: str, excel_folder_path: str, player_name: str):
        self.json_folder_path = json_folder_path
        self.excel_folder_path = excel_folder_path
        self.name = player_name

    def calculate_strike_rate(self, runs, balls_faced):
        if balls_faced == 0:
            return 0.0
        
        return round((runs / balls_faced) * 100, 2)
    
    def calculate_economy_rate(self, runs_conceded, balls_bowled):
        if balls_bowled == 0:
            return 0.0
        
        return round((runs_conceded / balls_bowled) * 6, 2)

    def file_exists(self):
        filename = f"{self.name}_game_save.json"
        full_path = f"{self.json_folder_path}/{filename}"
        try:
            with open(full_path, 'r'):
                return True
        except FileNotFoundError:
            return False

    def save_game_to_file(self, game_state):
        filename = f"{self.name}_game_save.json"
        full_path = f"{self.json_folder_path}/{filename}"
        try:
            if hasattr(game_state, 'to_dict'):
                state = game_state.to_dict()
            with open(full_path, 'w') as file:
                json.dump(state, file, indent=4)
            console.print(f"[green]Game saved successfully to[/green] [blue]{filename}[/blue]")
        except Exception as e:
            console.print(f"[bold red]Error saving game:[/bold red] {e}")


    def load_game_from_file(self):
        filename = f"{self.name}_game_save.json"
        full_path = f"{self.json_folder_path}/{filename}"
        try:
            with open(full_path, 'r') as file:
                loaded_data = json.load(file)

            defaults = {'half_centuries': 0, 'centuries': 0}
            for key, value in defaults.items():
                loaded_data.setdefault(key, value)

            game_state = GameState()
            game_state.from_dict(loaded_data)

            console.print(f"[green]Game loaded successfully from[/green] [blue]{filename}[/blue]")
            return game_state

        except FileNotFoundError:
            console.print(f"[red]No save file found:[/red] [blue]{filename}[/blue]")
            return None
        except Exception as e:
            console.print(f"[bold red]Error loading game:[/bold red] {e}")
            return None

    def apply_formatting(self, wb, sheet_name, headers):
            try:
                ws = wb[sheet_name]
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="4F81BD")
                fill_colour = PatternFill("solid", fgColor="DCE6F1")
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for col_num, header_text in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_num, value=header_text) 
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
        
                for col in ws.columns:
                    max_length = 0
                    column_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[column_letter].width = max_length + 2

                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row,
                                                           min_col=1, max_col=ws.max_column), start=1):
                    for cell in row:
                        cell.border = thin_border
                    if row_idx > 1 and (row_idx % 2 == 0): 
                        for cell in row:
                            cell.fill = fill_colour
                    elif row_idx > 1: 
                        for cell in row:
                            cell.fill = PatternFill(fill_type=None)
                ws.freeze_panes = "A2"
            except Exception as e:
                console.print(f"[bold red]Error applying formatting:[/bold red] {e}")

    def save_profile_and_match_data(self, player_profile, current_match_data):
        try:
            os.makedirs(self.excel_folder_path, exist_ok=True)
            player_name = player_profile.get('Player Name', self.name)
            full_path = os.path.join(self.excel_folder_path, f"{player_name}_profile_excel.xlsx")

            profile_data = {}
            for key, value in player_profile.items():
                if key != 'Player Name':
                    profile_data[key] = [value]
            
            profile_df = pd.DataFrame(profile_data)

            existing_matches = pd.DataFrame()
            if os.path.exists(full_path):
                try:
                    existing_matches = pd.read_excel(full_path, sheet_name="Match_History")
                except Exception:
                    existing_matches = pd.DataFrame()

            if current_match_data:
                new_match_df = pd.DataFrame([current_match_data])
                combined_matches = pd.concat([existing_matches, new_match_df], ignore_index=True)
                if "Match Number" in combined_matches.columns:
                    combined_matches.drop_duplicates(subset=["Match Number"], keep="last", inplace=True)
            else:
                combined_matches = existing_matches

            with pd.ExcelWriter(full_path, engine="openpyxl", mode="w") as writer:
                profile_df.to_excel(writer, sheet_name="Lifetime_Stats", index=False, header=True)
                combined_matches.to_excel(writer, sheet_name="Match_History", index=False, header=True)

            try:
                wb = load_workbook(full_path)
                if not profile_df.empty:
                    self.apply_formatting(wb, "Lifetime_Stats", list(profile_df.columns))
                if not combined_matches.empty:
                    self.apply_formatting(wb, "Match_History", list(combined_matches.columns))
                wb.save(full_path)
            except Exception as e:
                console.print(f"[bold red]Warning:[/bold red] [red]Could not apply formatting[/red] → [blue]{e}[/blue]")

            console.print(f"[green]Data successfully written to Excel.[/green]")

        except Exception as e:
            console.print(f"[bold red]Error writing data to Excel file:[/bold red] {e}")

    def get_current_match_data(self, game_state):
        if hasattr(game_state, 'to_dict'):
            game_state = game_state.to_dict()
        player_1st_rr = 0
        player_2nd_rr = 0
        ai_1st_rr = 0
        ai_2nd_rr = 0
        
        if game_state.get('balls_played_first_inn', 0) > 0:
            if game_state['toss_result'] == "batting":
                player_1st_rr = round((game_state['player_runs_1stinn'] * 6) / game_state['balls_played_first_inn'], 2)
            else:
                ai_1st_rr = round((game_state['ai_runs_1stinn'] * 6) / game_state['balls_played_first_inn'], 2)
                
        if game_state.get('balls_played_sec_inn', 0) > 0:
            if game_state['toss_result'] == "bowling":
                player_2nd_rr = round((game_state['player_runs_2ndinn'] * 6) / game_state['balls_played_sec_inn'], 2)
            else:
                ai_2nd_rr = round((game_state['ai_runs_2ndinn'] * 6) / game_state['balls_played_sec_inn'], 2)
        
        player_match_runs = game_state.get('player_runs_1stinn', 0) + game_state.get('player_runs_2ndinn', 0)
        player_match_balls_faced = game_state.get("balls_played_by_player", 0)
        player_match_wickets_taken = game_state.get('current_match_wickets', 0) 
        player_match_runs_conceded = game_state.get('player_runs_conceded_1stinn', 0) + game_state.get('player_runs_conceded_2ndinn', 0)
        player_match_balls_bowled = game_state.get('player_balls_bowled_1stinn', 0) + game_state.get('player_balls_bowled_2ndinn', 0)

        player_match_strike_rate = self.calculate_strike_rate(player_match_runs, player_match_balls_faced)
        player_match_economy_rate = self.calculate_economy_rate(player_match_runs_conceded, player_match_balls_bowled)

        return {
            "Match Number": game_state.get('total_matches_played', 0),
            "Player Runs 1st Innings": game_state.get('player_runs_1stinn', 0),
            "AI Runs 1st Innings": game_state.get('ai_runs_1stinn', 0),
            "Player Runs 2nd Innings": game_state.get('player_runs_2ndinn', 0),
            "AI Runs 2nd Innings": game_state.get('ai_runs_2ndinn', 0),
            "Player 1st Inn Run Rate": player_1st_rr,
            "Player 2nd Inn Run Rate": player_2nd_rr,
            "AI 1st Inn Run Rate": ai_1st_rr,
            "AI 2nd Inn Run Rate": ai_2nd_rr,
            "Player Match Strike Rate": player_match_strike_rate,
            "Player Match Economy Rate": player_match_economy_rate,
            "Player Match Wickets Taken": player_match_wickets_taken, 
            "Player Match Runs Conceded": player_match_runs_conceded,
            "Player Match Balls Bowled": player_match_balls_bowled,
            "Toss Result": game_state.get('toss_result', ""),
            "Balls Played 1st Innings": game_state.get('balls_played_first_inn', 0),
            "Balls Played 2nd Innings": game_state.get('balls_played_sec_inn', 0),
            "Difficulty": game_state.get('difficulty', 'medium'),
            "Match Result": game_state.get('match_result', 'unknown')
        }
class GamePlay:
    def __init__(self, player_name, min_choice=0, max_choice=10, score_alignment=80, difficulty = None, commentator = None, achievement_manager = None):
        self.name = player_name
        self.min_choice = min_choice
        self.max_choice = max_choice
        self.score_alignment = score_alignment
        self.difficulty = difficulty
        self.commentator = commentator
        self.achievement_manager = achievement_manager

    def input_num(self):
        while True:
            try:
                console.print(f"\n[bold yellow]Enter a number between[/bold yellow] {self.min_choice} [bold yellow]and[/bold yellow] {self.max_choice}[bold yellow]:[/bold yellow] ", end="")
                player_choice = int(input(f""))
                sounds.play_click()
                if player_choice < self.min_choice or player_choice > self.max_choice:
                    console.print(f"[red]Invalid number[/red]")
                    continue
                return player_choice
            except ValueError:
                console.print(f"[red]Invalid number[/red]")
            except KeyboardInterrupt:
                console.print(f"\n[red]Can't quit here[/red]")

    def get_ai_choice(self, player_prev_choice=None, ai_role=None):
        if self.difficulty == "easy":
            return random.randint(self.min_choice, self.max_choice)

        elif self.difficulty == "medium":
            choices = list(range(self.min_choice, self.max_choice + 1))
            if ai_role == "batting":
                if player_prev_choice in choices:
                    choices.remove(player_prev_choice)
            elif ai_role == "bowling":
                if player_prev_choice in choices:
                    choices += [player_prev_choice] * 2
            return random.choice(choices)

        elif self.difficulty == "hard":
            if ai_role == "bowling":
                if player_prev_choice != 0 and random.random() < 0.7:
                    return player_prev_choice
                else:
                    return random.randint(self.min_choice, self.max_choice)
            elif ai_role == "batting":
                if player_prev_choice != 0 and random.random() < 0.7:
                    avoid = [i for i in range(self.min_choice, self.max_choice + 1) if i != player_prev_choice]
                    return random.choice(avoid)
                else:
                    return random.randint(self.min_choice, self.max_choice)
        return random.randint(self.min_choice, self.max_choice)

    def calculate_run_rate(self, runs, balls_played):
        if balls_played == 0:
            return 0.0
        return round((runs * 6) / balls_played, 2)

    def calculate_required_run_rate(self, runs_needed, balls_remaining):
        if balls_remaining == 0:
            return 0.0 if runs_needed <= 0 else float('inf')
        return round((runs_needed * 6) / balls_remaining, 2)

    def display_run_rate_after_over(self, runs, balls_played, player_name="Player"):
        if balls_played > 0 and balls_played % 6 == 0:  
            run_rate = self.calculate_run_rate(runs, balls_played)
            over_number = balls_played // 6
            console.print(f"\n[green]---[/green] [magenta]After Over {over_number}[/magenta] [green]---[/green]")
            console.print(f"[blue]{player_name}[/blue] [green]Run Rate:[/green] {run_rate} [green]runs per over[/green]")
            console.print(f"[green]Total:[/green] {runs} [green]runs in[/green] {balls_played} [green]balls[/green] ({over_number} [green]overs[/green])")
            console.print(f"[green]=[/green]" * 35)

    def calculate_strike_rate(self, runs, balls_faced):
        if balls_faced == 0:
            return 0.0
        return round((runs / balls_faced) * 100, 2)
    
    def calculate_economy_rate(self, runs_conceded, balls_bowled):
        if balls_bowled == 0:
            return 0.0
        return round((runs_conceded / balls_bowled) * 6, 2)
    
    def check_milestones(self, current_runs, prev_runs, game_state):
        if prev_runs < 50 <= current_runs:
            game_state.half_centuries += 1
            if self.commentator and self.commentator.enabled:
                self.commentator.milestone_commentary(self.name, 50)

        if prev_runs < 100 <= current_runs:
            game_state.centuries += 1
            if self.commentator and self.commentator.enabled:
                self.commentator.milestone_commentary(self.name, 100)

        if current_runs > game_state.current_match_highscore:
            game_state.current_match_highscore = current_runs

            if current_runs > game_state.career_high_score:
                game_state.career_high_score = current_runs
                if hasattr(self, "achievement_manager"):
                    self.achievement_manager.achievements["High Score"] = game_state.career_high_score

                temp_ach = Panel.fit(
                    f"[bold green]🏆 New Career High Score![/bold green]\n"
                    f"[bold red]{current_runs} runs[/bold red] in a single innings!",
                    border_style="bright_blue"
                )
                console.print(temp_ach)
                sounds.play_achievement()

        if hasattr(self, 'achievement_manager'):
            self.achievement_manager.check_achievements(game_state)
            self.achievement_manager.save_achievements()

    def intro(self):
        console.print(f"\n[bold magenta]Welcome to AVG Hand Cricket,[/bold magenta] [bold blue]{self.name}[/bold blue] [bold magenta]![/bold magenta]")
        console.print(f"[red]- Prattik | XII-'A' SAP JEE[/red]".rjust(self.score_alignment))

    def match_over(self, game_state):
        while True:
            try:
                console.print(f"\n[bold yellow]Enter number of overs:[/bold yellow] ", end="")
                game_state.over = int(input(""))
                sounds.play_click()
                if game_state.over < 1:
                    console.print(f"[grey37]Enter a number greater than 0[/grey37]")
                    continue
                return game_state.over
            except ValueError:
                console.print(f"[bold red]Please enter a valid integer[/bold red]")

    def toss(self, game_state):
        while True:
            console.print(f"\n[bold yellow]Choose Odd or Even:[/bold yellow] ", end="")
            player_decision = input("").strip().lower()
            sounds.play_click()
            if player_decision in ["even", "odd"]:
                ai_decision = "odd" if player_decision == "even" else "even"
                break
            else:
                console.print(f"[red]Invalid odd or even[/red]")

        console.print(f"[blue]{self.name}[/blue] [green]chose[/green] {player_decision} [green]| AI chose[/green] {ai_decision}")        

        while True:
            try:
                console.print(f"\n[bold yellow]Choose a number between {self.min_choice} and {self.max_choice}:[/bold yellow] ", end="")
                player_hand = int(input(""))
                sounds.play_click()
                if self.min_choice <= player_hand <= self.max_choice:
                    break
                else:
                    console.print(f"[red]Invalid number[/red]")
            except ValueError:
                console.print(f"[red]Invalid number[/red]")

        ai_hand = random.randint(self.min_choice, self.max_choice)

        console.print(f"[blue]{self.name}[/blue] [green]chose[/green] {player_hand} [green]| AI chose[/green] {ai_hand}")

        role_picked = player_hand + ai_hand

        if player_decision == "odd":
            if role_picked % 2 == 0:
                ai_role = random.choice(["Bat", "Bowl"])
                console.print(f"[magenta]AI chose to[/magenta] {ai_role}")

                if ai_role == "Bat":
                    if self.commentator and self.commentator.enabled:
                        self.commentator.toss_commentary("AI", ai_role.lower())
                    console.print(f"\n[blue]{self.name}[/blue] [magenta]is bowling[/magenta] \n[blue]AI[/blue] [magenta]is batting[/magenta]")
                    game_state.toss_result = "bowling"
                else:
                    if self.commentator and self.commentator.enabled:
                        self.commentator.toss_commentary("AI", ai_role.lower())
                    console.print(f"\n[blue]{self.name}[/blue] [magenta]is batting[/magenta] \n[blue]AI[/blue] [magenta]is bowling[/magenta]")
                    game_state.toss_result = "batting"
            else:
                while True:
                    console.print(f"\n[bold yellow]Choose to bat or bowl?[/bold yellow] ", end="")
                    player_role = input("").strip().lower()
                    if player_role == "bat":
                        if self.commentator and self.commentator.enabled:
                            self.commentator.toss_commentary(self.name, player_role)
                        console.print(f"\n[blue]{self.name}[/blue] [magenta]is batting[/magenta] \n[blue]AI[/blue] [magenta]is bowling[/magenta]")
                        game_state.toss_result = "batting"
                        break
                    elif player_role == "bowl":
                        if self.commentator and self.commentator.enabled:
                            self.commentator.toss_commentary(self.name, player_role)
                        console.print(f"\n[blue]{self.name}[/blue] [magenta]is bowling[/magenta] \n[blue]AI[/blue] [magenta]is batting[/magenta]")
                        game_state.toss_result = "bowling"
                        break
                    else:
                        console.print(f"[red]Invalid bat or bowl[/red]")
        else:
            if role_picked % 2 == 0:
                while True:
                    console.print(f"\n[bold yellow]Choose to bat or bowl?[/bold yellow] ", end="")
                    player_role = input("").strip().lower()
                    sounds.play_click()

                    if player_role == "bat":
                        if self.commentator and self.commentator.enabled:
                            self.commentator.toss_commentary(self.name, player_role)
                        console.print(f"\n[blue]{self.name}[/blue] [magenta]is batting[/magenta] \n[blue]AI[/blue] [magenta]is bowling[/magenta]")
                        game_state.toss_result = "batting"
                        break
                    elif player_role == "bowl":
                        if self.commentator and self.commentator.enabled:
                            self.commentator.toss_commentary(self.name, player_role)
                        console.print(f"\n[blue]{self.name}[/blue] [magenta]is bowling[/magenta] \n[blue]AI[/blue] [magenta]is batting[/magenta]")
                        game_state.toss_result = "bowling"
                        break
                    else:
                        console.print(f"[red]Invalid bat or bowl[/red]")
            else:
                ai_role = random.choice(["Bat", "Bowl"])
                if ai_role == "Bat":
                    if self.commentator and self.commentator.enabled:
                        self.commentator.toss_commentary("AI", ai_role.lower())
                    console.print(f"\n[blue]{self.name}[/blue] [magenta]is bowling[/magenta] \n[blue]AI[/blue] [magenta]is batting[/magenta]")
                    game_state.toss_result = "bowling"
                else:
                    if self.commentator and self.commentator.enabled:
                        self.commentator.toss_commentary("AI", ai_role.lower())
                    console.print(f"\n[blue]{self.name}[/blue] [magenta]is batting[/magenta] \n[blue]AI[/blue] [magenta]is bowling[/magenta]")
                    game_state.toss_result = "batting"

        return game_state.toss_result

    def one_or_two(self):
        console.print(f"[green]One or Two[/green]")
        while True:
            try:
                console.print(f"\n[bold yellow]Choose 1 or 2:[/bold yellow] ", end="")
                player_12decision = int(input(""))
                sounds.play_click()
                if player_12decision in [1, 2]:
                    break
                else:
                    console.print(f"[red]Invalid input[/red]")
            except ValueError:
                console.print(f"[red]Invalid input[/red]")

        ai_12decision = random.choice([1, 2])

        console.print(f"[blue]{self.name}[/blue] [green]chose[/green] {player_12decision} [green]|[/green] [blue]AI[/blue] [green]chose[/green] {ai_12decision}")
        return player_12decision, ai_12decision

    def first_in(self, game_state):
        console.print(f"\n[bold yellow]First Innings:[/bold yellow]")

        if game_state.toss_result == "bowling":
            for ball in range(1, game_state.balls + 1):
                player_choice = self.input_num()
                ai_role = "bowling" if game_state.toss_result == "batting" else "batting"
                ai_choice = self.get_ai_choice(player_choice, ai_role)

                console.print(f"[yellow]Ball {ball}:[/yellow] [blue]{self.name}[/blue] [green]chose[/green] {player_choice} [green]|[/green] [blue]AI[/blue] [green]chose[/green] {ai_choice}")

                game_state.balls_played_first_inn += 1
                game_state.player_balls_bowled_1stinn += 1

                self.achievement_manager.check_achievements(game_state)

                if player_choice == ai_choice and player_choice != 0:
                    self.commentator.wicket_commentary("AI", bowler = self.name)
                    console.print(f"[bold yellow]First Innings Over[/bold yellow]")
                    game_state.lifetime_wickets += 1
                    game_state.current_match_wickets += 1
                    break

                if player_choice == 0 and ai_choice == 0:
                    player_12decision, ai_12decision = self.one_or_two()
                    if player_12decision == ai_12decision:
                        self.commentator.wicket_commentary("AI", bowler = self.name)
                        console.print(f"[bold yellow]First Innings Over[/bold yellow] \n[blue]AI[/blue] [green]has hit[/green] {game_state.ai_runs_1stinn} [green]runs[/green]")
                        game_state.lifetime_wickets += 1
                        game_state.current_match_wickets += 1
                    
                        break

                game_state.ai_runs_1stinn += ai_choice
                game_state.player_runs_conceded_1stinn += ai_choice
                if self.commentator and self.commentator.enabled:
                    self.commentator.run_commentary(ai_choice)
                console.print(f"[green]Current AI runs:[/green] {game_state.ai_runs_1stinn}".rjust(self.score_alignment))

                self.display_run_rate_after_over(
                    game_state.ai_runs_1stinn, 
                    game_state.balls_played_first_inn, 
                    "AI"
                )

            console.print(f"\n[blue]AI[/blue] [green]as hit[/green] {game_state.ai_runs_1stinn} [green]runs[/green]")
            final_run_rate = self.calculate_run_rate(game_state.ai_runs_1stinn, game_state.balls_played_first_inn)
            console.print(f"[blue]AI[/blue] [yellow]Final Run Rate:[/yellow] {final_run_rate} [green]runs per over[/green]")
            console.print(f"\n[blue]{self.name}[/blue] [magenta]needs[/magenta] {game_state.ai_runs_1stinn + 1} [magenta]runs to win[/magenta]")

        if game_state.toss_result == "batting":
            for ball in range(1, game_state.balls + 1):
                player_choice = self.input_num()
                ai_role = "bowling" if game_state.toss_result == "batting" else "batting"
                ai_choice = self.get_ai_choice(player_choice, ai_role)

                console.print(f"[yellow]Ball {ball}:[/yellow] [blue]{self.name}[/blue] [green]chose[/green] {player_choice} [green]|[/green] [blue]AI[/blue] [green]chose[/green] {ai_choice}")

                game_state.balls_played_first_inn += 1
                game_state.balls_played_by_player += 1

                self.achievement_manager.check_achievements(game_state)

                if player_choice == ai_choice and player_choice != 0:
                    self.commentator.wicket_commentary(self.name, bowler = "AI")
                    console.print(f"\n[bold yellow]First Innings Over[/bold yellow] \n[blue]{self.name}[/blue] [green]has hit[/green] {game_state.player_runs_1stinn} [green]runs[/green]")
                    break

                if player_choice == 0 and ai_choice == 0:
                    player_12decision, ai_12decision = self.one_or_two()
                    if player_12decision == ai_12decision:
                        self.commentator.wicket_commentary(self.name, bowler = "AI")
                        console.print(f"\n[bold yellow]First Innings Over[/bold yellow] \n[blue]{self.name}[/blue] [green]has hit[/green] {game_state.player_runs_1stinn} [green]runs[/green]")
                        break
                    else:
                        continue

                prev_player_runs = game_state.player_runs_1stinn           
                game_state.player_runs_1stinn += player_choice
                game_state.lifetime_runs += player_choice
                if player_choice == 6:
                    game_state.total_sixes += 1
                self.check_milestones(game_state.player_runs_1stinn, prev_player_runs, game_state)

                if self.commentator and self.commentator.enabled:
                    self.commentator.run_commentary(player_choice)
                console.print(f"[yellow]Current player runs:[/yellow] {game_state.player_runs_1stinn}".rjust(self.score_alignment))

                self.display_run_rate_after_over(
                    game_state.player_runs_1stinn, 
                    game_state.balls_played_first_inn, 
                    self.name
                )

            final_run_rate = self.calculate_run_rate(game_state.player_runs_1stinn, game_state.balls_played_first_inn)
            console.print(f"\n[blue]{self.name}[/blue] [yellow]Final Run Rate:[/yellow] {final_run_rate} [green]runs per over[/green]")
            console.print(f"\n[blue]AI[/blue] [magenta]needs[/magenta] {game_state.player_runs_1stinn + 1} [magenta]runs to win[/magenta]")

    def second_in(self, game_state):
        player_choice = 0
        ai_choice = 0
        player_12decision = 0
        ai_12decision = 0

        if game_state.toss_result == "bowling":
            console.print(f"\n[bold yellow]Second Innings:[/bold yellow]")
            console.print(f"\n[blue]{self.name}[/blue] [magenta]is now batting[/magenta] and [blue]AI[/blue] [magenta]is bowling[/magenta]")

            for i in range(1, game_state.balls + 1):
                if game_state.player_runs_2ndinn > game_state.ai_runs_1stinn:
                    console.print(f"[blue]{self.name}[/blue] [bold yellow]has achieved the target runs[/bold yellow]")
                    game_state.match_result = "win"
                    console.print(f"\n[bold yellow]Second Innings Over[/bold yellow]")
                    break

                player_choice = self.input_num()
                ai_role = "bowling" if game_state.toss_result == "batting" else "batting"
                ai_choice = self.get_ai_choice(player_choice, ai_role)

                console.print(f"[yellow]Ball {i}:[/yellow] [blue]{self.name}[/blue] [green]chose[/green] {player_choice} [green]|[/green] [blue]AI[/blue] [green]chose[/green] {ai_choice}")

                game_state.balls_played_sec_inn += 1
                game_state.balls_played_by_player += 1
                
                self.achievement_manager.check_achievements(game_state)

                if player_choice == ai_choice and player_choice != 0:
                    self.commentator.wicket_commentary(self.name, bowler = "AI")
                    console.print(f"\n[bold yellow]Second Innings Over[/bold yellow] \n[blue]{self.name}[/blue] [green]has hit[/green] {game_state.player_runs_2ndinn} [green]runs[/green]")
                    break

                if player_choice == 0 and ai_choice == 0:
                    player_12decision, ai_12decision = self.one_or_two()
                    if player_12decision == ai_12decision:
                        self.commentator.wicket_commentary(self.name, bowler = "AI")
                        console.print(f"\n[bold yellow]Second Innings Over[/bold yellow] \n[blue]{self.name}[/blue] [green]has hit[/green] {game_state.player_runs_2ndinn} [green]runs[/green]")
                        break
                    else:
                        continue

                prev_player_runs = game_state.player_runs_2ndinn
                game_state.player_runs_2ndinn += player_choice
                game_state.lifetime_runs += player_choice
                if player_choice == 6:
                    game_state.total_sixes += 1
                self.check_milestones(game_state.player_runs_2ndinn, prev_player_runs, game_state)
                if self.commentator and self.commentator.enabled:
                    self.commentator.run_commentary(player_choice)
                console.print(f"[yellow]Current player runs:[/yellow] {game_state.player_runs_2ndinn}".rjust(self.score_alignment))
                if i % 6 == 0: 
                    current_run_rate = self.calculate_run_rate(game_state.player_runs_2ndinn, game_state.balls_played_sec_inn)
                    required_run_rate = self.calculate_required_run_rate(
                        game_state.ai_runs_1stinn + 1 - game_state.player_runs_2ndinn, 
                        game_state.balls - game_state.balls_played_sec_inn
                    )
                    console.print(f"\n[magenta]-- After Over[/magenta] [green]{i//6}[/green] [magenta]---[/magenta]")
                    console.print(f"[magenta]Current Run Rate:[/magenta] [green]{current_run_rate}[/green]")
                    console.print(f"[magenta]Required Run Rate:[/magenta] [green]{required_run_rate}[/green]")
                    console.print(f"[magenta]=[/magenta]" * 35)

            console.print(f"\n[blue]{self.name}[/blue] [green]has hit[/green] {game_state.player_runs_2ndinn} [green]runs[/green]")

        elif game_state.toss_result == "batting":
            console.print(f"\n[bold yellow]Second Innings:[/bold yellow]")
            console.print(f"\n[blue]{self.name}[/blue] [magenta]is now bowling[/magenta] and [blue]AI[/blue] [magenta]is batting[/magenta]")

            for i in range(1, game_state.balls + 1):
                if game_state.ai_runs_2ndinn > game_state.player_runs_1stinn:
                    console.print(f"[blue]AI[/blue] [bold yellow]has achieved the target runs[/bold yellow]")
                    game_state.match_result = "loss" 
                    console.print(f"\n[bold yellow]Second Innings Over[/bold yellow]")
                    break

                player_choice = self.input_num()
                ai_role = "bowling" if game_state.toss_result == "batting" else "batting"
                ai_choice = self.get_ai_choice(player_choice, ai_role)

                console.print(f"[yellow]Ball {i}:[/yellow] [blue]{self.name}[/blue] [green]chose[/green] {player_choice} [green]|[/green] [blue]AI[/blue] [green]chose[/green] {ai_choice}")

                game_state.balls_played_sec_inn += 1
                game_state.player_balls_bowled_2ndinn += 1

                self.achievement_manager.check_achievements(game_state)


                if player_choice == ai_choice and player_choice != 0:
                    self.commentator.wicket_commentary("AI", bowler = self.name)
                    console.print(f"\n[bold yellow]Second Innings Over[/bold yellow]")
                    game_state.lifetime_wickets += 1
                    game_state.current_match_wickets += 1

                    self.achievement_manager.check_achievements(game_state)

                    break

                if player_choice == 0 and ai_choice == 0:
                    player_12decision, ai_12decision = self.one_or_two()
                    if player_12decision == ai_12decision:
                        self.commentator.wicket_commentary("AI", bowler = self.name)
                        console.print(f"\n[bold yellow]Second Innings Over[/bold yellow] \n[blue]AI[/blue] [green]has hit[/green] {game_state.ai_runs_2ndinn} [green]runs[/green]")
                        game_state.lifetime_wickets += 1
                        game_state.current_match_wickets += 1

                        break
                    else:
                        continue

                game_state.ai_runs_2ndinn += ai_choice
                game_state.player_runs_conceded_2ndinn += ai_choice
                if self.commentator and self.commentator.enabled:
                    self.commentator.run_commentary(ai_choice)
                console.print(f"[yellow]Current AI runs:[/yellow] {game_state.ai_runs_2ndinn}".rjust(self.score_alignment))
                if i % 6 == 0:  
                    current_run_rate = self.calculate_run_rate(game_state.ai_runs_2ndinn, game_state.balls_played_sec_inn)
                    required_run_rate = self.calculate_required_run_rate(
                        game_state.player_runs_1stinn + 1 - game_state.ai_runs_2ndinn, 
                        game_state.balls - game_state.balls_played_sec_inn
                    )
                    console.print(f"\n[magenta]--- After Over[/magenta] [green]{i//6}[/green] [magenta]---[/magenta]")
                    console.print(f"[magenta]Current Run Rate:[/magenta] [green]{current_run_rate}[/green]")
                    console.print(f"[magenta]Required Run Rate:[/magenta] [green]{required_run_rate}[/green]")
                    console.print(f"[magenta]=[/magenta]" * 35)

            console.print(f"\n[blue]AI[/blue] [green]has hit[/green] {game_state.ai_runs_2ndinn} [green]runs[/green]")

    def super_over(self, game_state):
        while True:
            console.print(f"\n[magenta]*** SUPER OVER! ***[/magenta]")
            console.print(f"[yellow]It’s a super over! Only[/yellow] [green]1 over[/green][yellow], whoever wins - wins the match![/yellow]\n")

            player_super_runs = 0
            ai_super_runs = 0

            console.print(f"[blue]{self.name}[/blue] [green]is batting first in the[/green] [magenta]Super Over[/magenta][green]![/green]")
            for ball in range(1, game_state.balls + 1):
                player_choice = self.input_num()
                ai_choice = self.get_ai_choice(player_choice, ai_role="bowling")
                console.print(f"[blue]{self.name}[/blue] [green]chose[/green] {player_choice} [green]|[/green] [blue]AI[/blue] [green]chose[/green] {ai_choice}")

                game_state.balls_played_by_player += 1

                self.achievement_manager.check_achievements(game_state)

                if player_choice == ai_choice:
                    console.print(f"[bold red]OUT![/bold red] [red]You have lost your wicket![/red]")
                    self.commentator.wicket_commentary(self.name, bowler = "AI")
                    console.print(f"[blue]{self.name}[/blue] [green]has hit[/green] {player_super_runs} [green]runs[/green]")
                    game_state.lifetime_runs += player_super_runs

                    break
                else:
                    player_super_runs += player_choice
                    if player_choice == 6:
                        game_state.total_sixes += 1
                    console.print(f"[yellow]Current player runs:[/yellow] {player_super_runs}".rjust(self.score_alignment))

            console.print(f"\n[blue]AI[/blue] [magenta]is batting now[/magenta] \n[blue]AI[/blue] [green]has to hit[/green] {player_super_runs + 1} [green]to win[/green]")
            for ball in range(1, game_state.balls + 1):
                if ai_super_runs > player_super_runs:
                    break
                player_choice = self.input_num()
                ai_choice = self.get_ai_choice(player_choice, ai_role="batting")
                console.print(f"[blue]{self.name}[/blue] [green]chose[/green] {player_choice} [green]|[/green] [blue]AI[/blue] [green]chose[/green] {ai_choice}")

                self.achievement_manager.check_achievements(game_state)

                if player_choice == ai_choice:
                    console.print(f"[bold red]OUT![/bold red] [blue]AI[/blue] [red]has lost its wicket![/red]")
                    self.commentator.wicket_commentary("AI", bowler = self.name)
                    console.print(f"[blue]AI[/blue] [green]has hit[/green] {ai_super_runs} [green]runs[/green]")
                    game_state.lifetime_wickets += 1
                    game_state.current_match_wickets += 1

                    break
                else:
                    ai_super_runs += ai_choice
                    console.print(f"[yellow]Current AI runs:[/yellow] {ai_super_runs}".rjust(self.score_alignment))

            console.print(f"\n[bold magenta]Super Over Results:[/bold magenta]")
            console.print(f"[blue]{self.name}'s[/blue] [magenta]runs:[/magenta] [green]{player_super_runs}[/green] [magenta]|[/magenta] [blue]AI[/blue] [magenta]runs:[/magenta] [green]{ai_super_runs}[/green]")

            if player_super_runs > ai_super_runs:
                console.print(f"[bold yellow]{self.name} has won the [/bold yellow][bold magenta]Super Over[/bold magenta][bold yellow]![/bold yellow]")
                game_state.match_result = "win" 
                game_state.total_wins += 1
                game_state.match_result = "win"
                break
            elif ai_super_runs > player_super_runs:
                console.print(f"[bold yellow]AI has won the [/bold yellow][bold magenta]Super Over[/bold magenta][bold yellow]![/bold yellow]")
                game_state.match_result = "loss" 
                game_state.total_losses += 1
                game_state.match_result = "loss"
                break
            else:
                console.print(f"[bold magenta]Super Over[/bold magenta] [green]has resulted in a tie! Another[/green] [magenta]super over[/magenta] [green]will be played![/green]")
class DisplayManager:
    def print_match_summary(self, match_summary):
        lines = []
        for inning, stats in match_summary.items():
            lines.append(Text(f"{inning}:", style="green"))
            for key, value in stats.items():
                if key == "run_rate":
                    lines.append(Text(f"Run rate: {value} runs per over", style="magenta"))
                else:
                    score_name = key.replace("_", " ").capitalize()
                    lines.append(Text(f"{score_name}: {value}", style="magenta"))
            lines.append(Text("")) 
        lines.append(Text("-----------------------------------", style="bold magenta"))

        console.print(
            Panel.fit(
                Group(*lines),
                title="Match Summary",
                border_style="bright_blue"
            )
        )
        time.sleep(3)

    def print_player_profile(self, profile):
        lines = []
        for key, value in profile.items():
            score_name = key.replace("_", " ").capitalize()
            lines.append(Text(f"{score_name}: {value}", style="magenta"))
        lines.append(Text("-----------------------------", style="bold magenta"))

        console.print(
            Panel.fit(
                Group(*lines),
                title="Player Profile",
                border_style="bright_blue"
            )
        )
        time.sleep(2)
class GameManager:
    def __init__(self, gameplay_instance, display_instance, db_handler, manager, player_name, achievement_manager):
        self.gameplay = gameplay_instance
        self.display = display_instance
        self.db_handler = db_handler
        self.manager = manager
        self.name = player_name
        self.achievement_manager = achievement_manager

    def result(self, game_state):
        summary = game_state.match_summary
        if game_state.toss_result == "bowling":
            ai_run_rate = self.gameplay.calculate_run_rate(
                game_state.ai_runs_1stinn, 
                game_state.balls_played_first_inn
            )
            player_run_rate = self.gameplay.calculate_run_rate(
                game_state.player_runs_2ndinn, 
                game_state.balls_played_sec_inn
            )
            summary['First Innings'] = {
                "ai_runs": game_state.ai_runs_1stinn,
                "balls_played": game_state.balls_played_first_inn,
                "run_rate": ai_run_rate
            }
            summary['Second Innings'] = {
                "player_runs": game_state.player_runs_2ndinn,
                "balls_played": game_state.balls_played_sec_inn,
                "run_rate": player_run_rate
            }

            total_runs = game_state.ai_runs_1stinn + game_state.player_runs_2ndinn
            total_balls = game_state.balls_played_first_inn + game_state.balls_played_sec_inn
            match_avg_run_rate = self.gameplay.calculate_run_rate(total_runs, total_balls)
            
            player_match_runs = game_state.player_runs_1stinn+ game_state.player_runs_2ndinn
            player_match_balls_faced = game_state.balls_played_by_player
            player_match_runs_conceded = game_state.player_runs_conceded_1stinn + game_state.player_runs_conceded_2ndinn
            player_match_balls_bowled = game_state.player_balls_bowled_1stinn + game_state.player_balls_bowled_2ndinn
            player_match_strike_rate = self.gameplay.calculate_strike_rate(player_match_runs, player_match_balls_faced)
            player_match_economy_rate = self.gameplay.calculate_economy_rate(player_match_runs_conceded, player_match_balls_bowled)


            summary['Match Statistics'] = {
                "average_run_rate": match_avg_run_rate,
                "total_runs_scored": total_runs,
                "total_balls_played": total_balls,
                "Player Match Strike Rate": player_match_strike_rate,
                "Player Match Economy Rate": player_match_economy_rate,
            }

            self.display.print_match_summary(summary)
            if game_state.ai_runs_1stinn > game_state.player_runs_2ndinn:
                console.print(f"[blue]AI[/blue] [green]has won the game.[/green] [blue]You[/blue] [red]have lost.[/red]")
                game_state.total_losses += 1
                game_state.match_result = "loss"

            elif game_state.ai_runs_1stinn < game_state.player_runs_2ndinn:
                console.print(f"[blue]You[/blue] [green]have won the game.[/green] [blue]AI[/blue] [red]has lost.[/red]")
                game_state.total_wins += 1
                game_state.match_result = "win"

            else:
                console.print(f"[yellow]It's a tie![/yellow]")
                game_state.total_draws += 1
                self.gameplay.super_over(game_state)

            game_state.lifetime_balls_faced += player_match_balls_faced
            game_state.lifetime_balls_bowled += player_match_balls_bowled
            game_state.lifetime_runs_conceded += player_match_runs_conceded

            current_match_data = self.manager.get_current_match_data(game_state)
            self.db_handler.save_match_data(game_state, self.name, player_match_runs_conceded, player_match_balls_bowled, game_state.current_match_wickets)
            self.handle_profile_display_and_save(game_state, current_match_data)
            self.achievement_manager.check_achievements(game_state)
            self.achievement_manager.save_achievements()

        elif game_state.toss_result == "batting":
            ai_run_rate = self.gameplay.calculate_run_rate(
                game_state.ai_runs_2ndinn, 
                game_state.balls_played_sec_inn
            )
            player_run_rate = self.gameplay.calculate_run_rate(
                game_state.player_runs_1stinn, 
                game_state.balls_played_first_inn
            )
            summary['First Innings'] = {
                "player_runs": game_state.player_runs_1stinn,
                "balls_played": game_state.balls_played_first_inn,
                "run_rate": player_run_rate
            }
            summary['Second Innings'] = {
                "ai_runs": game_state.ai_runs_2ndinn,
                "balls_played": game_state.balls_played_sec_inn,
                "run_rate": ai_run_rate
            }

            total_runs = game_state.player_runs_1stinn + game_state.ai_runs_2ndinn
            total_balls = game_state.balls_played_first_inn + game_state.balls_played_sec_inn
            match_avg_run_rate = self.gameplay.calculate_run_rate(total_runs, total_balls)
            
            player_match_runs = game_state.player_runs_1stinn + game_state.player_runs_2ndinn
            player_match_balls_faced = game_state.balls_played_by_player
            player_match_runs_conceded = game_state.player_runs_conceded_1stinn + game_state.player_runs_conceded_2ndinn
            player_match_balls_bowled = game_state.player_balls_bowled_1stinn + game_state.player_balls_bowled_2ndinn
            player_match_strike_rate = self.gameplay.calculate_strike_rate(player_match_runs, player_match_balls_faced)
            player_match_economy_rate = self.gameplay.calculate_economy_rate(player_match_runs_conceded, player_match_balls_bowled)

            summary['Match Statistics'] = {
                "average_run_rate": match_avg_run_rate,
                "total_runs_scored": total_runs,
                "total_balls_played": total_balls,
                "Player Match Strike Rate": player_match_strike_rate,
                "Player Match Economy Rate": player_match_economy_rate,
            }

            self.display.print_match_summary(summary)
            if game_state.ai_runs_2ndinn > game_state.player_runs_1stinn:
                console.print(f"[blue]AI[/blue] [green]has won the game.[/green] [blue]You[/blue] [red]have lost.[/red]")
                game_state.total_losses += 1
                game_state.match_result = "loss"

            elif game_state.ai_runs_2ndinn < game_state.player_runs_1stinn:
                console.print(f"[blue]You[/blue] [green]have won the game.[/green] [blue]AI[/blue] [red]has lost.[/red]")
                game_state.total_wins += 1
                game_state.match_result = "win"
                
            else:
                console.print(f"[yellow]It's a tie![/yellow]")
                game_state.total_draws += 1
                self.gameplay.super_over(game_state)

            game_state.lifetime_balls_faced += player_match_balls_faced
            game_state.lifetime_balls_bowled += player_match_balls_bowled
            game_state.lifetime_runs_conceded += player_match_runs_conceded

            current_match_data = self.manager.get_current_match_data(game_state)
            self.db_handler.save_match_data(game_state, self.name, player_match_runs_conceded, player_match_balls_bowled, game_state.current_match_wickets)
            self.handle_profile_display_and_save(game_state, current_match_data)
            self.achievement_manager.check_achievements(game_state)
            self.achievement_manager.save_achievements()

        else:
            console.print(f"[bold red]Error:[/bold red] toss_result [red]must be either 'bowling' or 'batting'.[/red]")

        self.achievement_manager.check_achievements(game_state)
        self.achievement_manager.save_achievements()


    def reset_game(self, game_state):
        self.career_high_score = 0
        game_state.over = 0
        game_state.player_runs_1stinn = 0
        game_state.ai_runs_1stinn = 0
        game_state.player_runs_2ndinn = 0
        game_state.ai_runs_2ndinn = 0
        game_state.toss_result = ""
        game_state.match_summary = {}
        game_state.balls_played_first_inn = 0
        game_state.balls_played_sec_inn = 0
        game_state.current_match_wickets = 0
        game_state.balls_played_by_player = 0
        game_state.player_balls_bowled_1stinn = 0
        game_state.player_balls_bowled_2ndinn = 0
        game_state.player_runs_conceded_1stinn = 0
        game_state.player_runs_conceded_2ndinn = 0
        game_state.match_result = ""
        game_state.total_sixes = 0
        game_state.current_match_highscore = 0
        game_state.temp_high_score_achievement_unlocked = 0

    def get_choice(self):
        while True:
            try:
                print(f"\n")
                menu = Panel.fit("[bold cyan]\n-------------------AVG HANDCRICKET-------------------[/bold cyan]\n" 
                    "\n[green]Type[/green] [blue]tutorial[/blue][green] to view the tutorial.[/green]" 
                    "\n[green]Type[/green] [blue]play[/blue][green] to play a match.[/green]" 
                    "\n[green]Type[/green] [blue]quit[/blue][green] to quit.[/green]" 
                    "\n[green]Type[/green] [blue]profile[/blue][green] to view the player profile.[/green]" 
                    "\n[green]Type[/green] [blue]save[/blue][green] to save current game state.[/green]" 
                    "\n[green]Type[/green] [blue]load[/blue][green] to load previous game state.[/green]" 
                    "\n[green]Type[/green] [blue]leaderboard[/blue][green] to view global stats.[/green]" 
                    "\n[green]Type[/green] [blue]switch[/blue][green] to switch player account.[/green]" 
                    "\n[green]Type[/green] [blue]settings[/blue][green] to change commentary or difficulty.[/green]",
                    title="Main Menu", border_style="bright_blue"
                )
                console.print(menu)
                choice = input(": ").strip().lower()
                if choice in ("tutorial","play", "quit", "profile", "save", "load", "leaderboard", "switch", "settings"):
                    return choice
                console.print(f"[red]Invalid input[/red]")
            except KeyboardInterrupt:
                console.print(f"\n[red]Game was interrupted[/red] [grey37]type no to exit properly[/grey37]")
                continue

    def handle_profile_display_and_save(self, game_state, current_match_data=None):
        if game_state.total_matches_played > 0:
            game_state.average_runs_per_match = (
                game_state.lifetime_runs / game_state.total_matches_played
            )

        lifetime_strike_rate = self.gameplay.calculate_strike_rate(
            game_state.lifetime_runs,
            game_state.lifetime_balls_faced
        )

        lifetime_economy_rate = self.gameplay.calculate_economy_rate(
            game_state.lifetime_runs_conceded,
            game_state.lifetime_balls_bowled
        )

        lifetime_run_rate = self.gameplay.calculate_run_rate(
            game_state.lifetime_runs,
            game_state.lifetime_balls_faced
        )

        game_state.player_profile = {
            "Lifetime runs": game_state.lifetime_runs,
            "Lifetime wickets": game_state.lifetime_wickets,
            "Total Matches Played": game_state.total_matches_played,
            "Total Wins": game_state.total_wins,
            "Total Losses": game_state.total_losses,
            "Total Draws": game_state.total_draws,
            "Average Runs per Match": game_state.average_runs_per_match,
            "Lifetime Balls Faced": game_state.lifetime_balls_faced,
            "Lifetime Balls Bowled": game_state.lifetime_balls_bowled,
            "Lifetime Runs Conceded": game_state.lifetime_runs_conceded,
            "Lifetime Run Rate": lifetime_run_rate,
            "Lifetime Strike Rate": lifetime_strike_rate,
            "Lifetime Economy Rate": lifetime_economy_rate,
            "Centuries": game_state.centuries,
            "Half Centuries": game_state.half_centuries
        }
        self.display.print_player_profile(game_state.player_profile)
        achievements = self.db_handler.get_achievements(self.name)
        self.achievement_manager.display_achievements(achievements)
        pw = self.db_handler.get_password(self.name)
        self.db_handler.save_profile(game_state.player_profile, self.name, pw)

        if current_match_data:
            self.manager.save_profile_and_match_data(game_state.player_profile, current_match_data)

    def show_settings_menu(self, game_state):
        while True:
            settings = Panel.fit("[bold cyan]\n-------------------Settings Menu-------------------[/bold cyan]\n"
                "\n[green]Type[/green] [blue]1[/blue][green] to toggle commentary on/off.[/green]"
                "\n[green]Type[/green] [blue]2[/blue][green] to change difficulty (easy/medium/hard).[/green]"
                "\n[green]Type[/green] [blue]3[/blue][green] to change player name.[/green]"
                "\n[green]Type[/green] [blue]4[/blue][green] to change password.[/green]"
                "\n[green]Type[/green] [blue]5[/blue][green] to go back to main menu.[/green]",
                title="Settings Menu", border_style="bright_blue"
            )
            console.print(settings)
            console.print(f"[cyan]Enter choice (1-5):[/cyan]")
            choice = input("").strip()

            if choice == "1":
                self.toggle_commentary(game_state)
            elif choice == "2":
                self.change_difficulty(game_state)
            elif choice == "3":
                self.change_player_name()
            elif choice == "4":
                self.change_password()
            elif choice == "5":
                break
            else:
                console.print(f"[red]Invalid choice.[/red] [grey37]Please select[/grey37] 1, 2, 3, 4, [grey37]or[/grey37] 5")

    def toggle_commentary(self, game_state):
        current = self.gameplay.commentator.enabled
        self.gameplay.commentator.enabled = not current
        game_state.commentary_enabled = self.gameplay.commentator.enabled
        status = "enabled" if self.gameplay.commentator.enabled else "disabled"
        console.print(f"[green]Commentary is now[/green] [cyan]{status}[/cyan][green].[/green]")

    def change_difficulty(self, game_state):
        while True:
            console.print(f"[green]Enter new difficulty[/green] [cyan](easy, medium, hard)[/cyan][green]:[/green]")
            difficulty = input("").strip().lower()
            if difficulty in ("easy", "medium", "hard"):
                game_state.difficulty = difficulty
                self.gameplay.difficulty = difficulty
                console.print(f"[green]Difficulty set to [cyan]{difficulty}[/cyan][green].[/green]")
                break
            else:
                console.print(f"[red]Invalid choice[/red] [grey37]Try again.[/grey37]")

    def change_player_name(self):
        global name
        db_handler = self.db_handler
        old_name = self.name

        
        console.print(f"[green]Enter new player name [/green][grey37](max 30 characters)[/grey37][green]:[/green]")
        new_name = input("").strip().title()
        if not new_name.replace(" ", "").isalpha():
            console.print(f"[red]Invalid name[/red] [grey37]Only letters and spaces allowed.[/grey37]")
            return
        if len(new_name) == 0 or len(new_name) > 30:
            console.print(f"[red]Name must be 1-30 characters.[/red]")
            return
        if new_name == old_name:
            console.print(f"[red]New name must not match the current name.[/red]")
            return
        if db_handler.player_exists(new_name):
            console.print(f"[red]This player name already exists. Please choose another.[/red]")
            return

        console.print(f"[green]To confirm changing name from[/green] [cyan]{old_name}[/cyan] [green]to[/green] [cyan]{new_name}[/cyan][green], please enter your current password:[/green]")
        old_pass = pwinput.pwinput("").strip()
        sounds.play_click()
        if old_pass != db_handler.get_password(old_name):
            console.print(f"[red]Incorrect password. Name not changed.[/red]")
            return

        db = db_handler.get_db()
        cursor = db.cursor()
        cursor.execute("UPDATE player_profile SET name = %s WHERE name = %s", (new_name, old_name))
        cursor.execute("UPDATE match_data SET player_name = %s WHERE player_name = %s", (new_name, old_name))
        db.commit()
        cursor.close()

        old_json_file = f"{self.manager.json_folder_path}/{old_name}_game_save.json"
        new_json_file = f"{self.manager.json_folder_path}/{new_name}_game_save.json"
    
        old_excel_file = f"{self.manager.excel_folder_path}/{old_name}_profile_excel.xlsx"
        new_excel_file = f"{self.manager.excel_folder_path}/{new_name}_profile_excel.xlsx"
    
        files_renamed = []

        try:
            if os.path.exists(old_json_file):
                os.rename(old_json_file, new_json_file)
                files_renamed.append("game save")
        except Exception as e:
            console.print(f"[bold red]Warning:[/bold red] [red]Could not rename JSON save file: {e}[/red]")
    
        try:
            if os.path.exists(old_excel_file):
                os.rename(old_excel_file, new_excel_file)
                files_renamed.append("Excel profile")
        except Exception as e:
            console.print(f"[bold red]Warning:[/bold red] [red]Could not rename Excel file: {e}[/red]")

        console.print(f"[green]Name changed from[/green] [blue]'{old_name}'[/blue] [green]to[/green] [blue]'{new_name}'[/blue] [green]successfully.[/green]")
        if files_renamed:
            console.print(f"[yellow]Renamed save files: {', '.join(files_renamed)}[/yellow]")

        self.manager.name = new_name
        self.name = new_name
        name = new_name

    def change_password(self):
        db_handler = self.db_handler
        name = self.name

        old_pass = pwinput.pwinput("Enter your current password: ").strip()
        sounds.play_click()
        if old_pass != db_handler.get_password(name):
            console.print(f"[red]Incorrect password. Password not changed.[/red]")
            return

        while True:
            new_pass = pwinput.pwinput("Enter new password: ").strip()
            sounds.play_click()
            if len(new_pass) < 3:
                console.print(f"[red]Password must be at least 3 characters long.[/red]")
                continue
            confirm = pwinput.pwinput("Confirm new password: ").strip()
            if new_pass != confirm:
                console.print(f"[red]Passwords do not match.[/red] [grey37]Try again.[/grey37]")
            else:
                db_handler.update_password(name, new_pass)
                console.print(f"[bold yellow]Password updated successfully.[/bold yellow]")
                break

    def handle_save_game_and_profile(self, game_state):
        if isinstance(game_state, dict):
            gs = GameState()
            gs.from_dict(game_state)
            game_state = gs

        self.manager.save_game_to_file(game_state)
        if game_state.total_matches_played > 0:
            game_state.average_runs_per_match = game_state.lifetime_runs / game_state.total_matches_played
            
        profile = {
            "Player Name": self.name,
            "Lifetime runs": game_state.lifetime_runs,
            "Lifetime wickets": game_state.lifetime_wickets,
            "Total Matches Played": game_state.total_matches_played,
            "Total Wins": game_state.total_wins,
            "Total Losses": game_state.total_losses,
            "Total Draws": game_state.total_draws,
            "Average Runs per Match": game_state.average_runs_per_match,
            "Lifetime Balls Faced": game_state.lifetime_balls_faced,
            "Lifetime Balls Bowled": game_state.lifetime_balls_bowled,
            "Lifetime Runs Conceded": game_state.lifetime_runs_conceded,
            "Centuries": game_state.centuries,
            "Half Centuries": game_state.half_centuries,
        }

        current_match_data = self.manager.get_current_match_data(game_state)
        self.manager.save_profile_and_match_data(profile, current_match_data)
class Commentator:
    def __init__(self, enabled=True):
        self.enabled = enabled

    def say(self, messages, style=""):
        if self.enabled:
            console.print(f"\n[bright blue]Commentator:[/bright blue]", random.choice(messages), style=style)

    def run_commentary(self, runs):
        if runs == 0:
            sounds.play_run()
            self.say(["That's a dot ball!", "No run scored.", "Good ball, no run."], style = "bold grey37 italic")
        elif runs == 1:
            sounds.play_run()
            self.say(["Just a single.", "Quick single taken.", "Keeps the scoreboard ticking."], style = "bold grey37 italic")
        elif runs == 2:
            sounds.play_run()
            self.say(["Two runs added! Great work they're putting it!", "Great running between the wickets."], style = "bold green italic")
        elif runs == 3:
            sounds.play_run()
            self.say(["Three runs! That's rare.", "Excellent placement and running."], style = "bold green italic")
        elif runs == 4:
            sounds.play_score()
            self.say(["That's a FOUR!", "Cracking shot through the gap!", "That races to the boundary!"], style = "bold cyan italic")
        elif runs == 5:
            sounds.play_score()
            self.say(["Five runs! Unbelievable!", "What a shot! That's five!", "They really wanted that five!"], style = "bold cyan italic")
        elif runs == 6:
            sounds.play_score()
            self.say(["SIX! What a hit!", "That's out of the park!", "Massive hit!"], style = "bold red italic")
        elif runs == 7:
            sounds.play_score()
            self.say(["Seven runs! great thinking here.", "They chose for a seven this time.", "They're looking far ahead"], style = "bold red italic")
        elif runs == 8:
            sounds.play_score()
            self.say(["Eight! Great decision.", "An Eight! He's confident.", "He just smashed it!"], style = "bold red italic")
        elif runs == 9:
            sounds.play_score()
            self.say(["Nine runs! they know what they're doing here.", "Makes me wanna sing a song right now.", "Thats it! They can go home happy now!"], style = "bold red italic")
        elif runs == 10:
            sounds.play_score()
            self.say(["A TEN!, What a shot!", "That's out of the planet!","Holy mother of GOD I'm SCREAMINGGGGG"], style = "bold red italic")

    def wicket_commentary(self, out_batsman, bowler = "AI"):
        sounds.play_wicket()
        self.say([
            f"{out_batsman} is OUT! Bowled by {bowler}!",
            f"{out_batsman} departs! What a delivery from {bowler}!",
            f"What a catch! {out_batsman} is gone, bowled by {bowler}!",
            f"{out_batsman} has been dismissed! {bowler} takes the wicket!",
            f"{out_batsman} is back to the pavilion! {bowler} strikes",
            f"{out_batsman} is out! {bowler} with a brilliant delivery!",
            f"{out_batsman} has been sent packing! {bowler} gets the wicket!",
            f"{out_batsman} is gone! {bowler} with a fantastic ball!",
            f"{out_batsman} is out! {bowler} takes the wicket with a superb delivery!",
            f"What a day he must be having right now! Superb bowling from {bowler}!",
            f"He must be feeling on top of the world right now! {bowler} gets the wicket!",
            f"He must've missed his morning grace! {out_batsman} is out!",
            f"{out_batsman} is out! {bowler} bowls a peach of a delivery!",
            f"{out_batsman} is out! {bowler} with a brilliant delivery!",
            f"Wicket! {bowler} gets rid of {out_batsman}!"
        ], style="bold red italic" )

    def toss_commentary(self, winner, decision):
        self.say([
            f"{winner} wins the toss and chooses to {decision}, Dunno what they're thinking.",
            f"The toss is won by {winner}, they opt to {decision}, Makes sense at this time of the day.",
            f"{winner} decides to {decision} after winning the toss, some quality decision making here."
        ], style = "cyan italic")

    def milestone_commentary(self, player_name, milestone):
        sounds.play_milestone()
        if self.enabled:
            if milestone == 50:
                self.say([
                    f"That's a brilliant fifty for {player_name}!",
                    f"{player_name} reaches his half-century with style!",
                    f"Fifty up for {player_name}! What an innings this is turning out to be.",
                    f"{player_name} brings up his fifty! A well-deserved milestone!",
                    f"Half-century for {player_name}! He is looking in great touch!",
                    f"{player_name} has reached his fifty! A fantastic effort!",
                    f"Fifty! {player_name} has played a superb innings so far!",
                    f"What a partnership this is turning out to be! {player_name} has reached his fifty!",
                    f"He lifts his bat up!!! What a moment for {player_name}!",
                ], style = "magenta italic")
            elif milestone == 100:
                self.say([
                    f"A magnificent century for {player_name}! Take a bow!",
                    f"And he lifts his bat up! Imagine how he must be feeling right now!",
                    f"This is the moment he has been waiting for! God Bless Him!",
                    f"{player_name} has shown the world yet again why he is a class player!",
                    f"{player_name} brings up his hundred! An outstanding display of batting!",
                    f"Hundred! {player_name} has reached the magical three-figure mark!",
                ], style = "bold magenta italic")
class LeaderboardManager:
    def __init__(self, db):
        self.db = db  

    def show_menu(self):
        while True:
            print(f"\n")          
            table = Table(title="[bold cyan]🏆 Leaderboard Menu[/bold cyan]", border_style="bright_blue")

            table.add_column("Option", style="blue", justify="center")
            table.add_column("Description", style="green")

            table.add_row("0", "Show leaderboard info.")
            table.add_row("1", "View most runs scored.")
            table.add_row("2", "View most wickets taken.")
            table.add_row("3", "View most balls played.")
            table.add_row("4", "View most active players.")
            table.add_row("5", "View highest scores scored.")
            table.add_row("6", "View most win rate.")
            table.add_row("7", "View highest all-rounders scores.")
            table.add_row("8", "View most lifetime run rate.")
            table.add_row("9", "View most lifetime strike rate.")
            table.add_row("10", "View most lifetime economy rate.")
            table.add_row("11", "View most half centuries.")
            table.add_row("12", "View most centuries.")
            table.add_row("13", "Return to the Main Menu.")

            console.print(table)

            console.print(f"[cyan]Enter choice (0-13):[/cyan]")
            choice = input("").strip()
            sounds.play_click()
            options = {
                "0": self.leaderboard_info,
                "1": self.show_runs,
                "2": self.show_wickets,
                "3": self.show_balls_played,
                "4": self.show_most_active,
                "5": self.show_high_scores,
                "6": self.show_win_rate,
                "7": self.show_all_rounders,
                "8": self.show_lifetime_run_rate,
                "9": self.show_lifetime_strike_rate, 
                "10": self.show_lifetime_economy_rate, 
                "11": self.show_half_centuries,
                "12": self.show_centuries,
            }

            if choice in options:
                options[choice]()
            elif choice == "14":
                break
            else:
                print(f"Invalid choice! Try again.")

    def leaderboard_info(self):
        info = Panel.fit("[bold cyan]\n-------------------Leaderboard Info-------------------[/bold cyan]\n"
            "\n[green]The leaderboard showcases the top players based on various statistics from all matches played.[/green]"
            "\n[green]It includes categories such as most runs, wickets, balls played, highest scores, win rates, and more.[/green]"
            "\n[green]Players are ranked based on their performance in each category, providing a competitive overview of the best players.[/green]"
            "\n[green]Stats are updated after each match to reflect the latest performances.[/green]",
            title="Leaderboard Information", border_style="bright_blue"
        )
        print(f"\n")
        console.print(info)
        time.sleep(5)

    def show_runs(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name, SUM(runs) as total_runs
            FROM match_data
            GROUP BY player_name
            ORDER BY total_runs DESC
            LIMIT 10
        """)
        runresult = cursor.fetchall()
        cursor.close()

        runstable = Table(title="🏆 Top Run Scorers", show_header=True, header_style="bold yellow", box=box.ROUNDED, border_style="bright_blue")
        runstable.add_column("Rank", style="cyan", justify="center")
        runstable.add_column("Player", style="bold green")
        runstable.add_column("Runs", style="bold magenta", justify="right")

        for i, (player_name, total_runs) in enumerate(runresult, start=1):
            runstable.add_row(str(i), player_name, str(total_runs))
        print(f"\n")
        console.print(runstable)
        time.sleep(5)

    def show_wickets(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name, SUM(wickets) as total_wickets
            FROM match_data
            GROUP BY player_name
            ORDER BY total_wickets DESC
            LIMIT 10
        """)
        wicketresults = cursor.fetchall()
        cursor.close()

        wickettable = Table(
            title="🎯 Top Wicket Takers",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_red",
            box=box.ROUNDED
        )
        wickettable.add_column("Rank", style="cyan", justify="center")
        wickettable.add_column("Player", style="bold green")
        wickettable.add_column("Wickets", style="bold magenta", justify="right")

        for i, (player_name, total_wickets) in enumerate(wicketresults, start=1):
            wickettable.add_row(str(i), player_name, str(total_wickets))

        print(f"\n")
        console.print(wickettable)
        time.sleep(5)

    def show_balls_played(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name, SUM(balls_faced) as total_balls
            FROM match_data
            GROUP BY player_name
            ORDER BY total_balls DESC
            LIMIT 10
        """)
        ballsresults = cursor.fetchall()
        cursor.close()

        ballstable = Table(
            title="⚡ Top Balls Faced",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_blue",
            box=box.ROUNDED
        )
        ballstable.add_column("Rank", style="cyan", justify="center")
        ballstable.add_column("Player", style="bold green")
        ballstable.add_column("Balls Faced", style="bold magenta", justify="right")

        for i, (player_name, total_balls) in enumerate(ballsresults, start=1):
            ballstable.add_row(str(i), player_name, str(total_balls))

        print(f"\n")
        console.print(ballstable)
        time.sleep(5)

    def show_most_active(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name, COUNT(*) as matches_played
            FROM match_data
            GROUP BY player_name
            ORDER BY matches_played DESC
            LIMIT 10
        """)
        activeresults = cursor.fetchall()
        cursor.close()

        activetable = Table(
            title="🔥 Most Active Players",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_cyan",
            box=box.ROUNDED
        )
        activetable.add_column("Rank", style="cyan", justify="center")
        activetable.add_column("Player", style="bold green")
        activetable.add_column("Matches Played", style="bold magenta", justify="right")

        for i, (player_name, matches_played) in enumerate(activeresults, start=1):
            activetable.add_row(str(i), player_name, str(matches_played))

        print(f"\n")
        console.print(activetable)
        time.sleep(5)

    def show_high_scores(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name, MAX(runs) as highest_score
            FROM match_data
            GROUP BY player_name
            ORDER BY highest_score DESC
            LIMIT 10
        """)
        highscoreresults = cursor.fetchall()
        cursor.close()

        highscoretable = Table(
            title="🏏 Highest Individual Scores",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_green",
            box=box.ROUNDED
        )
        highscoretable.add_column("Rank", style="cyan", justify="center")
        highscoretable.add_column("Player", style="bold green")
        highscoretable.add_column("High Score", style="bold magenta", justify="right")

        for i, (player_name, highest_score) in enumerate(highscoreresults, start=1):
            highscoretable.add_row(str(i), player_name, str(highest_score))

        print(f"\n")
        console.print(highscoretable)
        time.sleep(5)

    def show_win_rate(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name,
                   SUM(CASE WHEN result = 'win' THEN 1 ELSE 0 END) AS wins,
                   COUNT(*) AS matches,
                   ROUND(SUM(CASE WHEN result = 'win' THEN 1 ELSE 0 END) / COUNT(*) * 100, 2) AS win_rate
            FROM match_data
            GROUP BY player_name
            HAVING matches > 0
            ORDER BY win_rate DESC
            LIMIT 10
        """)

        winrateresults = cursor.fetchall()
        cursor.close()

        winratetable = Table(
            title="🏆 Win Rate Leaderboard",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_magenta",
            box=box.ROUNDED
        )
        winratetable.add_column("Rank", style="cyan", justify="center")
        winratetable.add_column("Player", style="bold green")
        winratetable.add_column("Win Rate %", style="bold blue", justify="right")
        winratetable.add_column("Record (W/M)", style="bold magenta", justify="center")

        for i, (player_name, wins, matches, win_rate) in enumerate(winrateresults, start=1):
            winratetable.add_row(
                str(i),
                player_name,
                f"{win_rate}%",
                f"{wins}/{matches}"
            )

        print(f"\n")
        console.print(winratetable)
        time.sleep(5)

    def show_all_rounders(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name,
                   SUM(runs) + SUM(wickets * 20) as all_rounder_score
            FROM match_data
            GROUP BY player_name
            ORDER BY all_rounder_score DESC
            LIMIT 10
        """)
        rounderresults = cursor.fetchall()
        cursor.close()

        roundertable = Table(
            title="💫 Top All-Rounders",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_green",
            box=box.ROUNDED
        )
        roundertable.add_column("Rank", style="cyan", justify="center")
        roundertable.add_column("Player", style="bold green")
        roundertable.add_column("All-Rounder Score", style="bold magenta", justify="right")

        for i, (player_name, score) in enumerate(rounderresults, start=1):
            roundertable.add_row(str(i), player_name, str(score))

        console.print(roundertable)

    def show_lifetime_run_rate(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT name,
                    lifetime_runs,
                    lifetime_balls_faced,
                    CASE 
                       WHEN lifetime_balls_faced > 0 
                       THEN ROUND((lifetime_runs * 6.0) / lifetime_balls_faced, 2)
                       ELSE 0.0
                    END as lifetime_run_rate
            FROM player_profile
            WHERE lifetime_balls_faced > 0
            ORDER BY lifetime_run_rate DESC
            LIMIT 10
        """)
        
        liferunresults = cursor.fetchall()
        cursor.close()

        liferuntable = Table(
            title="⚡ Lifetime Run Rate Leaderboard",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_blue",
            box=box.ROUNDED
        )
        liferuntable.add_column("Rank", style="cyan", justify="center")
        liferuntable.add_column("Player", style="bold green")
        liferuntable.add_column("Run Rate", style="bold blue", justify="right")
        liferuntable.add_column("Runs/Balls", style="bold magenta", justify="center")

        for i, (player_name, lifetime_runs, lifetime_balls_faced, lifetime_run_rate) in enumerate(liferunresults, start=1):
            liferuntable.add_row(
                str(i),
                player_name,
                str(lifetime_run_rate),
                f"{lifetime_runs}/{lifetime_balls_faced}"
            )

        print(f"\n")
        console.print(liferuntable)
        time.sleep(5)

    def show_lifetime_strike_rate(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT name,
                    lifetime_runs,
                    lifetime_balls_faced,
                    CASE
                       WHEN lifetime_balls_faced > 0
                       THEN ROUND((lifetime_runs * 100.0) / lifetime_balls_faced, 2)
                       ELSE 0.0
                   END as lifetime_strike_rate
            FROM player_profile
            WHERE lifetime_balls_faced > 0
            ORDER BY lifetime_strike_rate DESC
            LIMIT 10
        """)
        lifestrikeresults = cursor.fetchall()
        cursor.close()

        lifestriketable = Table(
            title="🚀 Lifetime Strike Rate Leaderboard",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_red",
            box=box.ROUNDED
        )
        lifestriketable.add_column("Rank", style="cyan", justify="center")
        lifestriketable.add_column("Player", style="bold green")
        lifestriketable.add_column("Strike Rate", style="bold blue", justify="right")
        lifestriketable.add_column("Runs/Balls", style="bold magenta", justify="center")

        for i, (player_name, lifetime_runs, lifetime_balls_faced, strike_rate) in enumerate(lifestrikeresults, start=1):
            lifestriketable.add_row(
                str(i),
                player_name,
                str(strike_rate),
                f"{lifetime_runs}/{lifetime_balls_faced}"
            )   

        print(f"\n")
        console.print(lifestriketable)
        time.sleep(5)

    def show_lifetime_economy_rate(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
                        SELECT name,
                        lifetime_runs_conceded,
                        lifetime_balls_bowled,
                        CASE
                        WHEN lifetime_balls_bowled > 0
                        THEN ROUND(lifetime_runs_conceded * 6.0 / lifetime_balls_bowled, 2)
                        ELSE 0.0
                    END as lifetime_economy_rate
                FROM player_profile
                WHERE lifetime_balls_bowled > 0
                ORDER BY lifetime_economy_rate ASC 
                LIMIT 10
            """)
        lifeecoresults = cursor.fetchall()
        cursor.close()

        lifeecotable = Table(
            title="🛡️ Lifetime Economy Rate Leaderboard",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_cyan",
            box=box.ROUNDED
        )
        lifeecotable.add_column("Rank", style="cyan", justify="center")
        lifeecotable.add_column("Player", style="bold green")
        lifeecotable.add_column("Economy Rate", style="bold blue", justify="right")
        lifeecotable.add_column("Runs Conceded/Balls Bowled", style="bold magenta", justify="center")

        for i, (player_name, lifetime_runs_conceded, lifetime_balls_bowled, economy_rate) in enumerate(lifeecoresults, start=1):
            lifeecotable.add_row(
                str(i),
                player_name,
                str(economy_rate),
                f"{lifetime_runs_conceded}/{lifetime_balls_bowled}"
            )
        
        print(f"\n")
        console.print(lifeecotable)
        time.sleep(5)

    def show_half_centuries(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT name, SUM(half_centuries) AS total_half_centuries
            FROM player_profile
            GROUP BY name
            ORDER BY total_half_centuries DESC
            LIMIT 10
        """)
        halfcentresults = cursor.fetchall()
        cursor.close()

        halfcenttable = Table(
            title="🏏 Half Centuries Leaderboard",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_cyan",
            box=box.ROUNDED
        )

        halfcenttable.add_column("Rank", style="cyan", justify="center")
        halfcenttable.add_column("Player", style="bold green")
        halfcenttable.add_column("Half Centuries", style="bold magenta", justify="right")

        for i, (player_name, total_half) in enumerate(halfcentresults, start=1):
            halfcenttable.add_row(
                str(i),
                player_name,
                str(total_half)
            )

        print(f"\n")
        console.print(halfcenttable)
        time.sleep(5)

    def show_centuries(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT name, SUM(centuries) AS total_centuries
            FROM player_profile
            GROUP BY name
            ORDER BY total_centuries DESC
            LIMIT 10
        """)
        centresults = cursor.fetchall()
        cursor.close()

        centtable = Table(
            title="🏆 Top 10 Centuries",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_cyan",
            box=box.ROUNDED
        )

        centtable.add_column("Rank", style="cyan", justify="center")
        centtable.add_column("Player", style="bold green")
        centtable.add_column("Centuries", style="bold magenta", justify="right")

        for i, (player_name, total_centuries) in enumerate(centresults, start=1):
            centtable.add_row(
                str(i),
                player_name,
                str(total_centuries)
            )

        print(f"\n")
        console.print(centtable)
        time.sleep(5)
class SilentSound:
    def set_volume(self, *args, **kwargs): pass
    def play(self, *args, **kwargs): pass
class SoundManager:
    def __init__(self):
        self.enabled = True
        try:
            pygame.mixer.init()
            pygame.mixer.music.load(os.path.expanduser("~/Downloads/assets/undertalebg.wav"))
            self.milestone = pygame.mixer.Sound(os.path.expanduser("~/Downloads/assets/cheer.wav"))
            self.wicket = pygame.mixer.Sound(os.path.expanduser("~/Downloads/assets/out.wav"))
            self.score = pygame.mixer.Sound(os.path.expanduser("~/Downloads/assets/score.wav"))
            self.click = pygame.mixer.Sound(os.path.expanduser("~/Downloads/assets/click.wav"))
            self.run = pygame.mixer.Sound(os.path.expanduser("~/Downloads/assets/run.wav"))
        except Exception as e:
            console.print(f"[bold red]Warning:[/bold red] [red]Sound initialization failed: {e}[/red]")
            self.enabled = False
        def _safe(path):
            if not self.enabled:
                return SilentSound()
            snd = load_sound_safe(path)
            return snd if snd is not None else SilentSound()

        if self.enabled and os.path.exists(os.path.expanduser("~/Downloads/assets/undertalebg.wav")):
            try:
                pygame.mixer.music.load(os.path.expanduser("~/Downloads/assets/undertalebg.wav"))
            except Exception:
                self.enabled = False  

        self.milestone = _safe(os.path.expanduser("~/Downloads/assets/cheer.wav"))
        self.wicket   = _safe(os.path.expanduser("~/Downloads/assets/out.wav"))
        self.score    = _safe(os.path.expanduser("~/Downloads/assets/score.wav"))
        self.click    = _safe(os.path.expanduser("~/Downloads/assets/click.wav"))
        self.run      = _safe(os.path.expanduser("~/Downloads/assets/run.wav"))
        self.achievement = _safe(os.path.expanduser("~/Downloads/assets/achievement.wav"))

    def play_bg(self):
        try:
            pygame.mixer.music.set_volume(0.2)
            pygame.mixer.music.play(-1)
        except Exception:
            self.enabled = False
    
    def stop_bg(self):
        try:
            pygame.mixer.music.stop()
        except Exception:
            pass

    def play_achievement(self):
        self.achievement.set_volume(0.7)
        self.achievement.play()
    
    def play_click(self):
        self.click.set_volume(0.5)
        self.click.play()

    def play_milestone(self):
        self.milestone.set_volume(0.5)
        self.milestone.play()
    
    def play_wicket(self):
        self.wicket.set_volume(0.6)
        self.wicket.play()
    
    def play_score(self):
        self.score.set_volume(0.9)
        self.score.play()

    def play_run(self):
        self.run.set_volume(0.9)
        self.run.play()
class AchievementManager:
    def __init__(self, db_handler, player_name):
        self.db_handler = db_handler
        self.player_name = player_name
        self.achievements = {
            "Noob": False,
            "You've got guts": False,
            "ITS WAR": False,
            "High Score": 0,
            "First shot": False,
            "First Step": False,
            "First Blood": False,
            "Century Maker": False,
            "Half-Century Hero": False,
            "Wicket Taker": False,
            "All-Rounder": False,
            "Match Winner": False,
            "Consistent Performer": False,
            "Run Machine": False,
            "Bowling Maestro": False,
            "Fielding Star": False,
            "Captain Fantastic": False,
            "Unstoppable": False,
            "Legend": False,
            "Immortal": False,
            "Perfectionist": False,
            "Marathoner": False,
            "Clutch Player": False,
            "Rising Star": False,
            "Veteran": False,
            "Champion": False,
            "Record Breaker": False,
            "Ultimate Player": False,
            "Cricket Guru": False,
            "Game Changer": False,
            "Master Blaster": False,
            "Silent Assassin": False,
            "The Wall": False,
            "The Finisher": False,
            "The Enforcer": False,
            "The Magician": False,
            "The Strategist": False,
            "The Entertainer": False,
            "The Showman": False,
            "Tiny Duck": False,
            "Golden Arm": False,
            "Hat-Trick Hero": False,
            "Speedster": False,
            "Periya kai": False,
            "Sixer King": False,
            "Namadhaaaaaaaaaa": False,
            "Epovume Namadhaaaaaaa": False,
            "Cain": False,
            "God": False,
            "Karadi": False,
            "Super Saiyan": False,
        }
        self.load_achievements()

    def load_achievements(self):
        stored_achievements = self.db_handler.get_achievements(self.player_name)
        if stored_achievements:
            for name in stored_achievements:
                self.achievements[name] = True

    def check_achievements(self, game_state):
        new_achievements = []

        if not self.achievements["Noob"] and difficulty == "easy":
            self.achievements["Noob"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold blue]Noob![/bold blue]\n\n[green]You're a noob!, nothing to feel bad about that lol[/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Noob")

        if not self.achievements["You've got guts"] and difficulty == "medium":
            self.achievements["You've got guts"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold blue]You've got guts![/bold blue]\n\n[green]You've got guts I must admit, good luck..[/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("You've got guts")

        if not self.achievements["ITS WAR"] and difficulty == "hard":
            self.achievements["ITS WAR"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red italic]ITS WAR![/bold red italic]\n\n[green]GET READY TO DIEEEE![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("ITS WAR")

        if not self.achievements["First shot"] and game_state.lifetime_runs >= 1:
            self.achievements["First shot"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]First Shot![/bold red]\n\n[green]Congratulations on scoring your first run! Keep it up![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("First shot")

        if not self.achievements["First Step"] and game_state.total_matches_played >= 1:
            self.achievements["First Step"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]First Step![/bold red]\n\n[green]Congratulations on playing your first match! Keep going![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("First Step")

        if not self.achievements["First Blood"] and game_state.lifetime_wickets >= 1:
            self.achievements["First Blood"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]First Blood![/bold red]\n\n[green]Congratulations on taking your first wicket! Keep it up![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("First Blood")

        if not self.achievements["Century Maker"] and game_state.lifetime_runs >= 100:
            self.achievements["Century Maker"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Century Maker![/bold red]\n\n[green]Amazing! You've scored a century! You're a batting legend![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Century Maker")

        if not self.achievements["Half-Century Hero"] and game_state.lifetime_runs >= 50:
            self.achievements["Half-Century Hero"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Half-Century Hero![/bold red]\n\n[green]Great job! You've reached a half-century! Keep the runs coming![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Half-Century Hero")

        if not self.achievements["Wicket Taker"] and game_state.lifetime_wickets >= 10:
            self.achievements["Wicket Taker"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Wicket Taker![/bold red]\n\n[green]Fantastic! You've taken 10 wickets! You're a bowling star![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Wicket Taker")

        if not self.achievements["All-Rounder"] and game_state.lifetime_runs >= 50 and game_state.lifetime_wickets >= 5:
            self.achievements["All-Rounder"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]All-Rounder![/bold red]\n\n[green]Incredible! You've excelled in both batting and bowling! True all-rounder![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("All-Rounder")

        if not self.achievements["Match Winner"] and game_state.total_wins >= 1:
            self.achievements["Match Winner"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Match Winner![/bold red]\n\n[green]Congratulations on leading your team to victory! You're a true champion![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Match Winner")

        if not self.achievements["Consistent Performer"] and game_state.total_matches_played >= 10:
            self.achievements["Consistent Performer"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Consistent Performer![/bold red]\n\n[green]Well done! You've shown great consistency over multiple matches! Keep it up![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Consistent Performer")

        if not self.achievements["Run Machine"] and game_state.lifetime_runs >= 500:
            self.achievements["Run Machine"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Run Machine![/bold red]\n\n[green]Outstanding! You've amassed 500 runs! You're a run-scoring machine![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Run Machine")

        if not self.achievements["Bowling Maestro"] and game_state.lifetime_wickets >= 50:
            self.achievements["Bowling Maestro"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Bowling Maestro![/bold red]\n\n[green]Exceptional! You've taken 50 wickets! You're a maestro with the ball![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Bowling Maestro")

        if not self.achievements["Fielding Star"] and game_state.current_match_wickets >= 20:
            self.achievements["Fielding Star"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Fielding Star![/bold red]\n\n[green]Brilliant! You've taken 20 catches! Your fielding is top-notch![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Fielding Star")
        
        if not self.achievements["Captain Fantastic"] and game_state.total_wins >= 5:
            self.achievements["Captain Fantastic"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Captain Fantastic![/bold red]\n\n[green]Amazing leadership! You've led your team to 5 victories! You're a fantastic captain![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Captain Fantastic")

        if not self.achievements["Unstoppable"] and game_state.lifetime_runs >= 200 and game_state.lifetime_wickets >= 20:
            self.achievements["Unstoppable"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Unstoppable![/bold red]\n\n[green]Incredible! You've scored 200 runs and taken 20 wickets! You're unstoppable![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Unstoppable")

        if not self.achievements["Legend"] and game_state.lifetime_runs >= 500 and game_state.lifetime_wickets >= 50:
            self.achievements["Legend"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Legend![/bold red]\n\n[green]Phenomenal! You've scored 500 runs and taken 50 wickets! You're a cricketing legend![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Legend")

        if not self.achievements["Immortal"] and game_state.lifetime_runs >= 1000 and game_state.lifetime_wickets >= 100:
            self.achievements["Immortal"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Immortal![/bold red]\n\n[green]Unbelievable! You've scored 1000 runs and taken 100 wickets! You're immortal in the cricketing world![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Immortal")

        if not self.achievements["Perfectionist"] and game_state.total_matches_played >= 20 and game_state.total_wins == game_state.total_matches_played:
            self.achievements["Perfectionist"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Perfectionist![/bold red]\n\n[green]Flawless! You've won all your matches! You're a perfectionist on the field![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Perfectionist")

        if not self.achievements["Marathoner"] and game_state.lifetime_balls_faced >= 1000:
            self.achievements["Marathoner"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Marathoner![/bold red]\n\n[green]Endurance! You've faced 1000 balls! You're a true marathoner at the crease![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Marathoner")

        if not self.achievements["Clutch Player"] and game_state.total_wins >= 3 and game_state.total_matches_played <= 5:
            self.achievements["Clutch Player"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Clutch Player![/bold red]\n\n[green]Pressure Performer! You've won 3 out of 5 matches! You're a clutch player![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Clutch Player")
        
        if not self.achievements["Rising Star"] and game_state.total_matches_played >= 5 and game_state.total_wins >= 3:
            self.achievements["Rising Star"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Rising Star![/bold red]\n\n[green]Impressive! You've played 5 matches and won 3! You're a rising star in cricket![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Rising Star")
        
        if not self.achievements["Veteran"] and game_state.total_matches_played >= 50:
            self.achievements["Veteran"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Veteran![/bold red]\n\n[green]Experience! You've played 50 matches! You're a seasoned veteran of the game![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Veteran")
        
        if not self.achievements["Champion"] and game_state.total_wins >= 10:
            self.achievements["Champion"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Champion![/bold red]\n\n[green]Glorious! You've won 10 matches! You're a true champion of cricket![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Champion")
        
        if not self.achievements["Ultimate Player"] and game_state.lifetime_runs >= 2000 and game_state.lifetime_wickets >= 200:
            self.achievements["Ultimate Player"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Ultimate Player![/bold red]\n\n[green]Legendary! You've scored 2000 runs and taken 200 wickets! You're the ultimate player![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Ultimate Player")
        
        if not self.achievements["Cricket Guru"] and game_state.total_matches_played >= 100:
            self.achievements["Cricket Guru"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Cricket Guru![/bold red]\n\n[green]Wisdom! You've played 100 matches! You're a true cricket guru![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Cricket Guru")
        
        if not self.achievements["Game Changer"] and game_state.total_wins >= 20:
            self.achievements["Game Changer"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Game Changer![/bold red]\n\n[green]Impactful! You've won 20 matches! You're a true game changer![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Game Changer")
        
        if not self.achievements["Master Blaster"] and game_state.lifetime_runs >= 3000:
            self.achievements["Master Blaster"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Master Blaster![/bold red]\n\n[green]Explosive! You've scored 3000 runs! You're a master blaster with the bat![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Master Blaster")
        
        if not self.achievements["Silent Assassin"] and game_state.lifetime_wickets >= 150:
            self.achievements["Silent Assassin"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Silent Assassin![/bold red]\n\n[green]Lethal! You've taken 150 wickets! You're a silent assassin with the ball![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Silent Assassin")
        
        if not self.achievements["The Wall"] and game_state.current_match_highscore >= 150:
            self.achievements["The Wall"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]The Wall![/bold red]\n\n[green]Impenetrable! You've scored 150 runs in an innings! You're the wall of the team![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("The Wall")
        
        if not self.achievements["The Finisher"] and game_state.lifetime_runs >= 2500:
            self.achievements["The Finisher"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]The Finisher![/bold red]\n\n[green]Clutch! You've scored 2500 runs! You're the finisher every team needs![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("The Finisher")
        
        if not self.achievements["The Enforcer"] and game_state.lifetime_wickets >= 100:
            self.achievements["The Enforcer"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]The Enforcer![/bold red]\n\n[green]Dominant! You've taken 100 wickets! You're the enforcer of the bowling attack![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("The Enforcer")

        if not self.achievements["The Magician"] and game_state.current_match_wickets >= 50:
            self.achievements["The Magician"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]The Magician![/bold red]\n\n[green]Spectacular! You've taken 50 catches! You're a fielding magician![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("The Magician")
        
        if not self.achievements["The Strategist"] and game_state.total_wins >= 15:
            self.achievements["The Strategist"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]The Strategist![/bold red]\n\n[green]Tactical Genius! You've won 15 matches! You're a true strategist on the field![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("The Strategist")
        
        if not self.achievements["The Entertainer"] and game_state.lifetime_runs >= 1500:
            self.achievements["The Entertainer"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]The Entertainer![/bold red]\n\n[green]Crowd Pleaser! You've scored 1500 runs! You're the entertainer of the game![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("The Entertainer")
        
        if not self.achievements["The Showman"] and game_state.current_match_highscore >= 120:
            self.achievements["The Showman"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]The Showman![/bold red]\n\n[green]Dazzling! You've scored 120 runs in an innings! You're the showman of the team![/green]", border_Style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("The Showman")
        
        if not self.achievements["Tiny Duck"] and game_state.lifetime_runs == 0 and game_state.lifetime_balls_faced >= 20:
            self.achievements["Tiny Duck"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Tiny Duck![/bold red]\n\n[green]Oh no! You've faced 20 balls without scoring! Better luck next time![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Tiny Duck")
        
        if not self.achievements["Golden Arm"] and game_state.lifetime_wickets >= 75:
            self.achievements["Golden Arm"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Golden Arm![/bold red]\n\n[green]Exceptional! You've taken 75 wickets! Your bowling arm is golden![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Golden Arm")
        
        if not self.achievements["Hat-Trick Hero"] and game_state.current_match_wickets >= 3:
            self.achievements["Hat-Trick Hero"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Hat-Trick Hero![/bold red]\n\n[green]Incredible! You've taken a hat-trick! You're a true hero with the ball![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Hat-Trick Hero")
        
        if not self.achievements["Speedster"] and game_state.lifetime_wickets >= 30:
            self.achievements["Speedster"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Speedster![/bold red]\n\n[green]Blazing Fast! You've taken 30 wickets! You're a speedster with the ball![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Speedster")
        
        if not self.achievements["Periya kai"] and game_state.lifetime_runs >= 333:
            self.achievements["Periya kai"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Periya kai![/bold red]\n\n[green]Awesome! You've scored 333 runs! You're a true 'Periya kai' in cricket![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Periya kai")
        
        if not self.achievements["Sixer King"] and game_state.total_sixes >= 50:
            self.achievements["Sixer King"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Sixer King![/bold red]\n\n[green]Powerful! You've hit 50 sixes! You're the king of sixes![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Sixer King")
        
        if not self.achievements["Namadhaaaaaaaaaa"] and game_state.lifetime_runs >= 69:
            self.achievements["Namadhaaaaaaaaaa"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Namadhaaaaaaaaaa![/bold red]\n\n[green]Haha! You've scored 69 runs! Namadhaaaaaaaaaa![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Namadhaaaaaaaaaa")
        
        if not self.achievements["Epovume Namadhaaaaaaa"] and game_state.lifetime_runs >= 1000:
            self.achievements["Epovume Namadhaaaaaaa"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Epovume Namadhaaaaaaa![/bold red]\n\n[green]Legendary! You've scored 1000 runs! Epovume Namadhaaaaaaa![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Epovume Namadhaaaaaaa")

        if not self.achievements["Karadi"] and game_state.lifetime_wickets >= 5:
            self.achievements["Karadi"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Karadi![/bold red]\n\n[green]Easter Egg for karadi!![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Karadi")

        if not self.achievements["Super Saiyan"] and game_state.lifetime_runs >= 777:
            self.achievements["Super Saiyan"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Super Saiyan![/bold red]\n\n[green]Legendary! You've scored 777 runs! You've become a Super Saiyan![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Super Saiyan")

        if not self.achievements["Cain"] and game_state.lifetime_wickets >= 666:
            self.achievements["Cain"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]Cain![/bold red]\n\n[green]Sinister! You've taken 666 wickets! You're the Cain of cricket![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("Cain")

        if not self.achievements["God"] and game_state.lifetime_runs >= 10000 and game_state.lifetime_wickets >= 10000:
            self.achievements["God"] = True
            achievement_display = Panel.fit("[bold green]🏆 Achievement Unlocked:[/bold green] [bold red]God![/bold red]\n\n[green]Divine! You've scored 10,000 runs and taken 10,000 wickets! You're the God of cricket![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
            new_achievements.append("God")

        if game_state.career_high_score > self.achievements.get("High Score", 0):
            self.achievements["High Score"] = game_state.career_high_score   
            achievement_display = Panel.fit(f"[bold green]🏆 New High Score!:[/bold green] [bold red]{self.achievements["High Score"]}[/bold red]\n\n[green]Keep working hard mate![/green]", border_style="bright_blue")
            console.print(achievement_display)
            sounds.play_achievement()
        
        if new_achievements:
            self.save_achievements()
    
    def save_achievements(self):
        unlocked = [name for name, unlocked in self.achievements.items() if unlocked]
        self.db_handler.update_achievements(self.player_name, unlocked)

    def has_new_achievements(self):
        return any(self.achievements.values())

    def display_achievements(self, achievements: list):
        if not achievements:
            console.print("\n[bold yellow]No achievements unlocked yet. Keep playing to earn achievements![/bold yellow]")
            return

        table = Table(title="🏆 Achievements Unlocked 🏆",
                    show_header=True,
                    header_style="bold magenta",
                    box=box.ROUNDED,
                    border_style="bright_blue")
        table.add_column("Achievement", style="bold green")

        for achievement in achievements:
            if achievement == "High Score":
                value = self.achievements.get("High Score", 0)
                if value > 0:
                    table.add_row(f"{achievement} – {value}")
                else:
                    table.add_row(f"{achievement} – Not achieved yet")
            else:
                table.add_row(achievement)

        console.print("\n")
        console.print(table)
        time.sleep(5)

    def reset_achievements(self):
        for key in self.achievements.keys():
            self.achievements[key] = False
        self.save_achievements()
        console.print("\n[bold green]All achievements have been reset.[/bold green]") 
def load_sound_safe(path):
    try:
        if not os.path.exists(path):
            return None  
        return pygame.mixer.Sound(path)
    except Exception:
        return None
def ask_commentary_setting():
    while True:
        try:
            choice = Prompt.ask(f"\n[green]Do you want commentary?[/green] [cyan](yes/no)[/cyan][green]:[/green] ").strip().lower()
            sounds.play_click()
            if choice == "yes":
                return True
            elif choice == "no":
                return False
            else:
                console.print(f"[red]Invalid input! Please enter yes or no.[/red]")
        except KeyboardInterrupt:
            console.print(f"[bold red]Game interrupted! try again[/bold red]")
def ask_difficulty_setting():
    while True:
        try:
            difficulty = Prompt.ask("\n[green]Choose difficulty[/green] [cyan](easy / medium / hard)[/cyan][green]:[/green] ").strip().lower()
            sounds.play_click()
            if difficulty in ("easy", "medium", "hard"):
                console.print(f"\n[green]Difficulty set to[/green] [bright blue]{difficulty}[/bright blue][green].[/green]")
                return difficulty
            else:
                console.print(f"\n[red]Invalid choice. Try again.[/red]")
        except KeyboardInterrupt:
            console.print(f"\n[bold red]Game interrupted! try again[/bold red]")
def login_menu(db_handler, name):
    while True:
        sounds.play_click()

        print(f"\n--- Login Menu ---")
        print(f"1. Start Game")
        print(f"2. Change Password")
        print(f"3. Logout")
        choice = input("Choose: ").strip()

        sounds.play_click()
        
        if choice == "1":
            return True
        elif choice == "2":
            old_pass = pwinput.pwinput("Enter current password: ").strip()
            if old_pass == db_handler.get_password(name):
                new_pass = pwinput.pwinput("Enter new password: ").strip()
                confirm = pwinput.pwinput("Confirm new password: ").strip()
                if new_pass == confirm:
                    db_handler.update_password(name, new_pass)
                    print(f"Password updated successfully.")
                else:
                    print(f"Passwords do not match.")
            else:
                print(f"Incorrect current password.")
        elif choice == "3":
            return False
        else:
            print(f"Invalid choice.")
def player_login(db_handler):
    while True:
        try:
            name = Prompt.ask(f"\n[yellow]Enter your name[/yellow] [cyan](max 30 characters)[/cyan][green]:[/green] ").strip().title()
            sounds.play_click()
            if not name.replace(" ", "").isalpha():
                console.print(f"\n[red]Invalid name. Only letters and spaces allowed.[/red]")
                continue
            if len(name) == 0:
                console.print(f"\n[red]Name cannot be empty.[/red]")
                continue
            if len(name) > 30:
                print(f"\n[red]Name too long![/red] [blue]({len(name)}/30 characters)[/blue]")
                continue

            if db_handler.player_exists(name):
                console.print(f"\n[bold yellow]Welcome back,[/bold yellow] [blue]{name}[/blue][bold yellow]![/bold yellow]")
                stored_password = db_handler.get_password(name)
                authenticated = False
                console.print(f"\n[bold green italic]You have 3 attempts to enter your password.[/bold green italic]")
                i=0

                for attempt in range(3):
                    i+=1
                    console.print(f"\n[green]Enter your password:[/green] ", end="")
                    entered_pass = pwinput.pwinput("").strip()
                    if entered_pass == stored_password:
                        authenticated = True
                        break
                    elif i < 3:
                        console.print(f"\n[red]Incorrect password. You have[/red] [blue]{3 - i}[/blue] [red]attempts left.[/red]")
                if not authenticated:
                    console.print(f"[bold red]Too many failed attempts. Exiting login.[/bold red]")
                    return None
                
                console.print(f"[magenta italic]Logging in as[/magenta italic] [blue]{name}[/blue][magenta italic]...[/magenta italic]")
                time.sleep(2)
                reset = Prompt.ask(f"\n[green italic]Do you want to reset your stats?[/green italic] [yellow](yes/no)[/yellow][green italic]:[/green italic] ").strip().lower()
                sounds.play_click()
                if reset == "yes":
                    console.print(f"[cyan italic]Resetting profile and match history...[/cyan italic]")
                    time.sleep(2)
                    game_manager.reset_game(game_state)
                    profile = {
                        "Lifetime runs": 0,
                        "Lifetime wickets": 0,
                        "Total Matches Played": 0,
                        "Total Wins": 0,
                        "Total Losses": 0,
                        "Total Draws": 0,
                        "Average Runs per Match": 0,
                        "Lifetime Balls Faced": 0,
                        "Lifetime Balls Bowled": 0,
                        "Lifetime Runs Conceded": 0,
                        "Centuries": 0,
                        "Half Centuries": 0,
                    }
                    db_handler.save_profile(profile, name, stored_password)
                    
                    db = db_handler.get_db()
                    cursor = db.cursor()
                    cursor.execute("DELETE FROM match_data WHERE player_name = %s", (name,))
                    db.commit()
                    cursor.close()
                    
                    console.print(f"\n[green]Profile and match history reset.[/green]")
                return name
            else:
                console.print(f"[yellow]Profile for[/yellow] [blue]{name}[/blue] [yellow]does not exist. Creating a new profile...[/yellow]")
                time.sleep(2)
                while True:
                    console.print(f"\n[green]Set a password (min 3 characters):[/green] ", end="")
                    password = pwinput.pwinput("").strip()
                    if len(password) < 3:
                        console.print(f"[red]Password too short![/red] [blue]({len(password)}/3 characters)[/blue]")
                        continue
                    console.print(f"[green]Confirm password:[/green] ", end="")
                    confirm = pwinput.pwinput("").strip()
                    if password != confirm:
                        console.print(f"[red]Passwords do not match. Try again.[/red]")
                    else:
                        break
                console.print(f"\n[cyan italic]Creating profile for[/cyan italic] [blue]{name}[/blue][cyan italic]...[/cyan italic]")
                time.sleep(3)
                console.print(f"[bold green]Profile created successfully! Welcome,[/bold green] [blue]{name}[/blue][bold green]![/bold green]")
                profile = {
                    "Lifetime runs": 0,
                    "Lifetime wickets": 0,
                    "Total Matches Played": 0,
                    "Total Wins": 0,
                    "Total Losses": 0,
                    "Total Draws": 0,
                    "Average Runs per Match": 0,
                    "Lifetime Balls Faced": 0,
                    "Lifetime Balls Bowled": 0,
                    "Lifetime Runs Conceded": 0,
                    "Centuries": 0,
                    "Half Centuries": 0,
                }
                db_handler.save_profile(profile, name, password)

            return name
        except KeyboardInterrupt:
            console.print(f"\n[bold red]Game interrupted! try again[/bold red]")
def main(game_manager, game_state):
    global manager, name, gameplay, commentator, display
    while True:
        choice = game_manager.get_choice()

        sounds.play_click()

        if choice == "tutorial":
            print('\n')
            tutorialtable = Table(title="🎉 Welcome to the Cricket Game Tutorial! 🎉", show_header=True, header_style="bold yellow", box=box.ROUNDED, border_style="bright_blue")
            tutorialtable.add_column("Instructions", style="bold green")
            tutorialtable.add_row("1. Choose Odd or Even to start the game.")
            tutorialtable.add_row("2. You will be playing against an AI opponent.")
            tutorialtable.add_row("3. AI will choose the opposite of what you choose.")
            tutorialtable.add_row("4. If the sum of the numbers corresponds to what you choose, you'll get the opportunity to choose to bat or bowl.")
            tutorialtable.add_row("5. There will be two innings: bat to bowl or bowl to bat.")
            tutorialtable.add_row("6. Using the same numbers will drop a wicket.")
            tutorialtable.add_row("7. Each number will add to your total runs.")
            tutorialtable.add_row("8. At the end, the player with the highest amount of runs wins!")
            tutorialtable.add_row("9. Note: If both Player and AI choose 0, it will lead to the 1 or 2 rule (only 1 or 2 is allowed, but they don't count as runs).")
            tutorialtable.add_row("10. In case of a tie, a super over will occur (only 1 over with just 6 balls).")
            console.print(tutorialtable)
            console.print(f"\n[bold yellow]Good luck and have fun![/bold yellow]")

        elif choice == "play":
            game_state.total_matches_played += 1
            game_manager.reset_game(game_state)
            toss_result = gameplay.toss(game_state)
            over = gameplay.match_over(game_state)
            game_state.balls = game_state.over * balls_per_over
            gameplay.first_in(game_state)
            gameplay.second_in(game_state)
            game_manager.result(game_state)

        elif choice == "quit":
            console.print(f"[bold yellow italic]Thank you for playing[/bold yellow italic] [blue]{name}[/blue][bold yellow italic]![/bold yellow italic]")
            pygame.mixer.music.stop()
            achievement_manager.save_achievements()
            db_handler.close_db()
            break

        elif choice == "save":
            console.print(f"\n[yellow italic]Saving game state...[/yellow italic]")
            time.sleep(3)
            game_manager.handle_save_game_and_profile(game_state)

        elif choice == "load":
            console.print(f"[yellow italic]Loading save game state...[/yellow italic]")
            time.sleep(3)
            loaded_game = manager.load_game_from_file()
            if loaded_game:
                game_state.update(loaded_game)
                welcome = Panel.fit(f"[bold yellow]Welcome back {name}!![/bold yellow]\n\n[green]Total matches played:[/green] [blue]{game_state.total_matches_played}[/blue]\n[green]Lifetime runs:[/green] [blue]{game_state.lifetime_runs}[/blue]", border_style="bright_blue")
                console.print(welcome)
                game_state.balls = game_state.over * balls_per_over

        elif choice == "profile":
            console.print(f"\n[magenta italic]Loading profile...[/magenta italic]")
            time.sleep(2)
            game_manager.handle_profile_display_and_save(game_state)

        elif choice == "leaderboard":
            console.print(f"\n [yellow italic]Attempting to open leaderboard...[/yellow italic]")
            time.sleep(2)
            leaderboard_manager = LeaderboardManager(db_handler)
            leaderboard_manager.show_menu()

        elif choice == "switch":
            console.print(f"\n[yellow italic]Attempting to switch...[/yellow italic]")
            time.sleep(2)
            achievement_manager = AchievementManager(db_handler, name)
            new_name = player_login(db_handler)
            if not new_name:
                console.print(f"\n[bold red]Login failed.[/bold red]")
                continue
            
            name = new_name
            
            game_state = GameState()
            game_state.reset_lifetime()

            manager = GameDataManager(folder_path_for_json, folder_path_for_excel, name)

            if manager.file_exists():
                console.print(f"\n[blue italic]A pre-existing save with this name was found.[/blue italic]")
                while True:
                    try:
                        autosave_load_choice = Prompt.ask("\n[green]Would you like to load the save file?[green] [cyan](yes/no)[/cyan][green]:[/green] ").strip().lower()
                        sounds.play_click()
                        if autosave_load_choice == "yes":
                            console.print(f"\n[yellow italic]Loading save game state...[/yellow italic]")
                            time.sleep(3)
                            loaded_game = manager.load_game_from_file()
                            if loaded_game:
                                game_state.update(loaded_game)
                                print('\n')
                                welcome = Panel.fit(f"[bold yellow]Welcome back {name}!![/bold yellow]\n\n[green]Total matches played:[/green] [blue]{game_state.total_matches_played}[/blue]\n[green]Lifetime runs:[/green] [blue]{game_state.lifetime_runs}[/blue]", border_style="bright_blue")
                                console.print(welcome)
                            break
                        elif autosave_load_choice == "no":
                            commentary_enabled = ask_commentary_setting()
                            difficulty = ask_difficulty_setting()
                            game_state.commentary_enabled = commentary_enabled
                            game_state.difficulty = difficulty
                            profile_data = db_handler.load_profile(name)
                            if profile_data:
                                console.print(f"\b[bold yellow]Profile stats loaded from database.[/bold yellow]")
                                game_state.update(profile_data)
                            break
                        else:
                            console.print(f"\n[red]Invalid input! Please enter yes or no.[/red]")
                    except KeyboardInterrupt:
                        console.print(f"\n[bold red]Game interrupted! try again[/bold red]")
            else:
                commentary_enabled = ask_commentary_setting()
                difficulty = ask_difficulty_setting()
                game_state.commentary_enabled = commentary_enabled
                game_state.difficulty = difficulty
                profile_data = db_handler.load_profile(name)
                if profile_data:
                    console.print(f"\n[bold yellow]Profile stats loaded from database.[/bold yellow]")
                    game_state.update(profile_data)

            commentator = Commentator(game_state.commentary_enabled)
            gameplay = GamePlay(name, min_choice, max_choice, score_alignment, game_state.difficulty, commentator, achievement_manager)
            display = DisplayManager()
            game_manager = GameManager(gameplay, display, db_handler, manager, name, achievement_manager)
            gameplay.intro()

        elif choice == "settings":
            console.print(f"\n[yellow italic]Opening settings...[/yellow italic]")
            time.sleep(2)
            game_manager.show_settings_menu(game_state)

        else:
            console.print(f"\n[red]Invalid choice. Try again.[/red]")
def animated_print(*args, **kwargs):
    sep = kwargs.get("sep", " ")
    end = kwargs.get("end", "\n")
    text = sep.join(str(a) for a in args)

    for ch in text:
        sys.stdout.write(ch)
        sys.stdout.flush()
        time.sleep(0.02)

    sys.stdout.write(end)
    sys.stdout.flush()
if __name__ == "__main__":
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table
    from rich.prompt import Prompt
    from rich.console import Group
    from rich.text import Text
    from rich import box
    import random
    import pwinput
    import os
    import builtins
    import time
    import sys
    import mysql.connector
    from mysql.connector import Error
    import json
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    os.environ["PYGAME_HIDE_SUPPORT_PROMPT"] = "1"
    import pygame

    class AnimatedConsole(Console):
        animate = True 

        def print(self, *args, style=None, delay=0.02, **kwargs):
            end = kwargs.pop("end", "\n")

            if self.animate and len(args) == 1 and isinstance(args[0], str):
                full_text: Text = self.render_str(args[0], style=style)

                for index, char in enumerate(full_text.plain):
                    super().print(full_text[index:index+1], end="", **kwargs)
                    time.sleep(delay)

                if end:
                    super().print(end, **kwargs)
            else:
                super().print(*args, style=style, **kwargs)

    console = AnimatedConsole()

    sounds = SoundManager()
    sounds.play_bg()
    
    base_dir = os.path.join(os.path.expanduser("~"), "Documents", "Hand_cricket_saves")
    folder_path_for_json = os.path.join(base_dir, "json_saves")
    folder_path_for_excel = os.path.join(base_dir, "excel_saves")

    os.makedirs(folder_path_for_json, exist_ok=True)
    os.makedirs(folder_path_for_excel, exist_ok=True)

    original_print = print
    builtins.print = animated_print
    
    difficulty = ""

    while True:
        try:
            console.print(Panel.fit("[bold yellow]Enter your MySQL root password to connect to the database:[/bold yellow]", border_style="bright_blue"))
            db_password = pwinput.pwinput("")
            sounds.play_click()
            break
        except KeyboardInterrupt:
            console.print(f"\n[bold red]Game interrupted! try again[/bold red]")
            
    DB_CONFIG = {
        "host": "localhost",
        "user": "root",
        "password": db_password,
        "database": "handcricket"
    }

    _DB = None

    db_handler = HandCricketDB(DB_CONFIG)

    db_handler.create_database()

    db_handler.init_db()

    game_state = GameState()

    name = player_login(db_handler)
    if not name:
        console.print(f"\n[bold red]Login failed. Exiting game.[/bold red]")
        sys.exit()

    achievement_manager = AchievementManager(db_handler, name)

    min_choice = 0
    max_choice = 10
    balls_per_over = 6
    score_alignment = 80

    manager = GameDataManager(folder_path_for_json, folder_path_for_excel, name)
    
    if manager.file_exists():
        console.print(Panel.fit("[bold yellow]A pre-existing save with this name was found.[/bold yellow]", border_style="bright_blue"))
        while True:
            try:
                autosave_load_choice = Prompt.ask("[green]Would you like to load the save file?[/green] (yes/no): ").strip().lower()
                sounds.play_click()
                if autosave_load_choice == "yes":
                    console.print(f"[grey37]Loading save game state...[/grey37]")
                    time.sleep(3)
                    loaded_game = manager.load_game_from_file()
                    if loaded_game:
                        game_state.update(loaded_game)
                        console.print(f"\n[bold green]Welcome back {name}!![/bold green]")
                        console.print(f"[bold green]Total matches played:[/bold green] {game_state.total_matches_played}")
                        console.print(f"[bold green]Lifetime runs:[/bold green] {game_state.lifetime_runs}")
                        break
                elif autosave_load_choice == "no":
                    commentary_enabled = ask_commentary_setting()
                    difficulty = ask_difficulty_setting()
                    game_state.commentary_enabled = commentary_enabled
                    game_state.difficulty = difficulty
                    profile_data = db_handler.load_profile(name)
                    if profile_data:
                        console.print(f"\n[bold yellow]Profile stats loaded from database.[/bold yellow]")
                        game_state.update(profile_data)
                    break
                else:
                    console.print(f"\n[red]Invalid input! Please enter yes or no.[/red]")
            except KeyboardInterrupt:
                console.print(f"\n[bold red]Game interrupted! try again[/bold red]")
    else:
        commentary_enabled = ask_commentary_setting()
        difficulty = ask_difficulty_setting()
        game_state.commentary_enabled = commentary_enabled
        game_state.difficulty = difficulty
        profile_data = db_handler.load_profile(name)
        if profile_data:
            console.print(f"\n[bold yellow]Profile stats loaded from database.[/bold yellow]")
            game_state.update(profile_data)

    def show_personal_leaderboard_panel(db_handler, player_name):
        categories = {
            "Total Runs Scored": "runs",
            "Total Wickets Taken": "wickets",
            "Total Balls Faced": "balls",
            "Win Rate": "win_rate",
            "Highest Individual Score": "high_score",
            "Lifetime Run Rate": "lifetime_run_rate",
            "Lifetime Strike Rate": "lifetime_strike_rate",
            "Lifetime Economy Rate": "lifetime_economy_rate",
            "Half Centuries": "half_centuries",
            "Centuries": "centuries",
        }
        leading = []
        for label, key in categories.items():
            if db_handler.get_leaderboard_leader(key) == player_name:
                leading.append(label)
        if leading:
            lines = [Text(f"🏆 You are ranked #1 in:")] + [Text(f"• {lbl}", style="bold green") for lbl in leading]
            console.print(Panel.fit(
                Text.assemble(*lines),
                title=f"Welcome back, {player_name}!",
                border_style="bright_blue"
            ))


    show_personal_leaderboard_panel(db_handler, name)
    
    commentator = Commentator(game_state.commentary_enabled)
    gameplay = GamePlay(name, min_choice, max_choice, score_alignment, game_state.difficulty, commentator, achievement_manager)
    display = DisplayManager()
    game_manager = GameManager(gameplay, display, db_handler, manager, name, achievement_manager)
    gameplay.intro()
    time.sleep(1)
    main(game_manager, game_state)