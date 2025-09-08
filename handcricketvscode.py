class HandCricketDB:
    def __init__(self, config):
        self.DB_CONFIG = config
        self._db = None
        self.current_user = None

    def create_database(self):
        try:
            sql_server = mysql.connector.connect(
                host=self.DB_CONFIG["host"],
                user=self.DB_CONFIG["user"],
                password=self.DB_CONFIG["password"],
            )
            cursor = sql_server.cursor()
            cursor.execute(f"CREATE DATABASE IF NOT EXISTS {self.DB_CONFIG['database']}")                                                                                                                                                                                                           
            cursor.close()
            sql_server.close()
            console.print(f"[bold yellow]Database[/bold yellow] [bold red]'{self.DB_CONFIG['database']}'[/bold red] [bold yellow]has been created/loaded successfully.[/bold yellow]")
        except Error as e:
            console.print(f"[dark_grey]Database could not be created due to error:[/dark_grey] [bold red{e}[/bold red]")
            exit(1)

    def get_db(self):
        try:
            if self._db is None or not self._db.is_connected():
                if self._db:
                    self._db.close()
                self._db = mysql.connector.connect(**self.DB_CONFIG, buffered = True)
            return self._db
        except mysql.connector.Error as error:
            console.print(f"[dark_grey]Database connection error:[/dark_grey] [bold red]{error}[/bold red]")
            return None

    def init_db(self):
        ddl_profile = """
            CREATE TABLE IF NOT EXISTS player_profile ( 
                id               INT AUTO_INCREMENT PRIMARY KEY, 
                name             VARCHAR(100) UNIQUE NOT NULL,
                password         VARCHAR(100) NOT NULL, 
                lifetime_runs    INT  DEFAULT 0, 
                lifetime_wickets INT  DEFAULT 0, 
                total_matches    INT  DEFAULT 0, 
                total_wins       INT  DEFAULT 0, 
                total_losses     INT  DEFAULT 0, 
                total_draws      INT  DEFAULT 0,
                lifetime_balls_faced INT  DEFAULT 0,
                lifetime_balls_bowled INT  DEFAULT 0,
                lifetime_runs_conceded INT  DEFAULT 0,
                centuries        INT  DEFAULT 0,
                half_centuries   INT  DEFAULT 0,
                avg_runs         FLOAT  DEFAULT 0.0
            )
            """
        ddl_match_data = """
            CREATE TABLE IF NOT EXISTS match_data (
                id INT AUTO_INCREMENT PRIMARY KEY,
                player_name VARCHAR(100) NOT NULL,
                runs INT  DEFAULT 0,
                wickets INT  DEFAULT 0,
                balls_faced INT  DEFAULT 0,
                player_balls_bowled INT  DEFAULT 0,
                player_runs_conceded INT  DEFAULT 0,
                result VARCHAR(10),
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_player_name (player_name)
            )
            """
            
        db = self.get_db()
        if db:
            cur = db.cursor()
            try:
                cur.execute(ddl_profile)
                cur.execute(ddl_match_data)
                db.commit()
            finally:
                cur.close()

    def get_password(self, player_name):
        db = self.get_db()
        cursor = db.cursor()
        cursor.execute("SELECT password FROM player_profile WHERE name = %s", (player_name,))
        row = cursor.fetchone()
        cursor.close()
        return row[0] if row else None

    def update_password(self, player_name, new_password):
        db = self.get_db()
        cursor = db.cursor()
        cursor.execute("UPDATE player_profile SET password = %s WHERE name = %s", (new_password, player_name))
        db.commit()
        cursor.close()

    def save_profile(self, profile: dict, player_name: str = "", password: str = None):
        if not player_name:
            raise ValueError("player name must be provided")

        sql = """
              INSERT INTO player_profile
              (name, password, lifetime_runs, lifetime_wickets, total_matches,
               total_wins, total_losses, total_draws, avg_runs, lifetime_balls_faced, lifetime_balls_bowled, lifetime_runs_conceded,
               centuries, half_centuries)
              VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) 
              ON DUPLICATE KEY UPDATE
                password = VALUES(password),
                lifetime_runs = VALUES(lifetime_runs),
                lifetime_wickets = VALUES(lifetime_wickets),
                total_matches = VALUES(total_matches),
                total_wins = VALUES(total_wins),
                total_losses = VALUES(total_losses),
                total_draws = VALUES(total_draws),
                avg_runs = VALUES(avg_runs),
                lifetime_balls_faced = VALUES(lifetime_balls_faced),
                lifetime_balls_bowled = VALUES(lifetime_balls_bowled),
                lifetime_runs_conceded = VALUES(lifetime_runs_conceded),
                centuries = VALUES(centuries),
                half_centuries = VALUES(half_centuries)
              """
        vals = (
            player_name,
            password,
            profile["Lifetime runs"],
            profile["Lifetime wickets"],
            profile["Total Matches Played"],
            profile["Total Wins"],
            profile["Total Losses"],
            profile["Total Draws"],
            profile["Average Runs per Match"],
            profile.get("Lifetime Balls Faced", 0), 
            profile.get("Lifetime Balls Bowled", 0), 
            profile.get("Lifetime Runs Conceded", 0),
            profile.get("Centuries", 0),            
            profile.get("Half Centuries", 0), 
        )
        cur = None
        try:
            db = self.get_db()
            cur = db.cursor()
            cur.execute(sql, vals)
            db.commit()
        
        except KeyError as e:
            console.print(f"[dark_grey]Missing required field in profile:[/dark_grey] [bold red]{e}[/bold red]")
        except Exception as e:
            console.print(f"[bold red]Failed to save profile:[/bold red] {e}")
            try:
                self.get_db().rollback()
            except:
                pass
        finally:
            try:
                if cur:
                    cur.close()
            except:
                pass

    def save_match_data(self, game_var: dict, player_name: str, player_runs_conceded: int, player_balls_bowled: int, player_match_wickets: int):
        try:
            db = self.get_db()
            cursor = db.cursor()
            sql = """
                INSERT INTO match_data (player_name, runs, wickets, balls_faced, player_runs_conceded, player_balls_bowled, result)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """

            match_runs = game_var.get('player_runs_1stinn', 0) + game_var.get('player_runs_2ndinn', 0)
            match_wickets = player_match_wickets 
            match_balls_faced = game_var.get("balls_played_by_player", 0)

            values = (
                player_name,
                match_runs,
                match_wickets,
                match_balls_faced,
                player_runs_conceded,
                player_balls_bowled,
                game_var.get("match_result", "unknown"),     
            )
            cursor.execute(sql, values)
            db.commit()
            cursor.close()
        except Exception as e:
            console.print(f"[bold red]Error saving match data:[/bold red] {e}")

    def load_profile(self, player_name):
        try:
            db = self.get_db()
            cursor = db.cursor()
            cursor.execute("""
                SELECT lifetime_runs, lifetime_wickets, total_matches, 
                    total_wins, total_losses, total_draws, avg_runs,
                    lifetime_balls_faced, lifetime_balls_bowled, centuries, half_centuries, lifetime_runs_conceded
                FROM player_profile WHERE name = %s
            """, (player_name,))
            row = cursor.fetchone()
            cursor.close()
            if row:
                return {
                    "lifetime_runs": row[0],
                    "lifetime_wickets": row[1],
                    "total_matches_played": row[2],
                    "total_wins": row[3],
                    "total_losses": row[4],
                    "total_draws": row[5],
                    "average_runs_per_match": row[6],
                    "lifetime_balls_faced": row[7],   
                    "lifetime_balls_bowled": row[8],  
                    "centuries": row[9],              
                    "half_centuries": row[10],
                    "lifetime_runs_conceded": row[11]
                }
        except Exception as e:
            console.print(f"[bold red]Error loading profile from database:[/bold red] {e}")
        return None

    def player_exists(self, player_name: str):
        try:
            db = self.get_db()
            cursor = db.cursor()
            cursor.execute("SELECT 1 FROM player_profile WHERE name = %s", (player_name,))
            result = cursor.fetchone()
            cursor.close()
            return result is not None
        except Exception as e:
            console.print(f"[bold red]Error checking player existence:[/bold red] {e}")
            return False

    def close_db(self):
        try:
            if self._db:
                self._db.close()
                self._db = None
        except:
            pass
class GameDataManager:
    def __init__(self, json_folder_path: str, excel_folder_path: str, player_name: str):
        self.json_folder_path = json_folder_path
        self.excel_folder_path = excel_folder_path
        self.name = player_name

    def calculate_strike_rate(self, runs, balls_faced):
        if balls_faced == 0:
            return 0.0
        
        return round((runs / balls_faced) * 100, 2)
    
    def calculate_economy_rate(self, runs_conceded, balls_bowled):
        if balls_bowled == 0:
            return 0.0
        
        return round((runs_conceded / balls_bowled) * 6, 2)

    def file_exists(self):
        filename = f"{self.name}_game_save.json"
        full_path = f"{self.json_folder_path}/{filename}"
        try:
            with open(full_path, 'r'):
                return True
        except FileNotFoundError:
            return False

    def save_game_to_file(self, game_var):
        filename = f"{self.name}_game_save.json"
        full_path = f"{self.json_folder_path}/{filename}"
        try:
            with open(full_path, 'w') as file:
                json.dump(game_var, file, indent=4)
            console.print(f"[green]Game saved successfully to[/green] [blue]{filename}[/blue]")
        except Exception as e:
            console.print(f"[bold red]Error saving game:[/bold red] {e}")

    def load_game_from_file(self):
        filename = f"{self.name}_game_save.json"
        full_path = f"{self.json_folder_path}/{filename}"
        try:
            with open(full_path, 'r') as file:
                loaded_data = json.load(file)
            defaults = {
                'half_centuries': 0,
                'centuries': 0
            }
            for key, value in defaults.items():
                loaded_data.setdefault(key, value)
            console.print(f"[green]Game loaded successfully from[/green] [blue]{filename}[/blue]")
            return loaded_data
        except FileNotFoundError:
            console.print(f"[red]No save file found:[/red] [blue]{filename}[/blue]")
            return None
        except Exception as e:
            console.print(f"[bold red]Error loading game:[/bold red] {e}")
            return None

    def apply_formatting(self, wb, sheet_name, headers):
            try:
                ws = wb[sheet_name]
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="4F81BD")
                fill_colour = PatternFill("solid", fgColor="DCE6F1")
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for col_num, header_text in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_num, value=header_text) 
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
        
                for col in ws.columns:
                    max_length = 0
                    column_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[column_letter].width = max_length + 2

                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row,
                                                           min_col=1, max_col=ws.max_column), start=1):
                    for cell in row:
                        cell.border = thin_border
                    if row_idx > 1 and (row_idx % 2 == 0): 
                        for cell in row:
                            cell.fill = fill_colour
                    elif row_idx > 1: 
                        for cell in row:
                            cell.fill = PatternFill(fill_type=None)
                ws.freeze_panes = "A2"
            except Exception as e:
                console.print(f"[bold red]Error applying formatting:[/bold red] {e}")

    def save_profile_and_match_data(self, player_profile, current_match_data):
        try:
            os.makedirs(self.excel_folder_path, exist_ok=True)
            player_name = player_profile.get('Player Name', self.name)
            full_path = os.path.join(self.excel_folder_path, f"{player_name}_profile_excel.xlsx")

            profile_data = {}
            for key, value in player_profile.items():
                if key != 'Player Name':
                    profile_data[key] = [value]
            
            profile_df = pd.DataFrame(profile_data)

            existing_matches = pd.DataFrame()
            if os.path.exists(full_path):
                try:
                    existing_matches = pd.read_excel(full_path, sheet_name="Match_History")
                except Exception:
                    existing_matches = pd.DataFrame()

            if current_match_data:
                new_match_df = pd.DataFrame([current_match_data])
                combined_matches = pd.concat([existing_matches, new_match_df], ignore_index=True)
                if "Match Number" in combined_matches.columns:
                    combined_matches.drop_duplicates(subset=["Match Number"], keep="last", inplace=True)
            else:
                combined_matches = existing_matches

            with pd.ExcelWriter(full_path, engine="openpyxl", mode="w") as writer:
                profile_df.to_excel(writer, sheet_name="Lifetime_Stats", index=False, header=True)
                combined_matches.to_excel(writer, sheet_name="Match_History", index=False, header=True)

            try:
                wb = load_workbook(full_path)
                if not profile_df.empty:
                    self.apply_formatting(wb, "Lifetime_Stats", list(profile_df.columns))
                if not combined_matches.empty:
                    self.apply_formatting(wb, "Match_History", list(combined_matches.columns))
                wb.save(full_path)
            except Exception as e:
                console.print(f"[bold red]Warning:[/bold red] [red]Could not apply formatting[/red] → [blue]{e}[/blue]")

            console.print("[green]Data successfully written to Excel.[/green]")

        except Exception as e:
            console.print(f"[bold red]Error writing data to Excel file:[/bold red] {e}")

    def get_current_match_data(self, game_var):
        player_1st_rr = 0
        player_2nd_rr = 0
        ai_1st_rr = 0
        ai_2nd_rr = 0
        
        if game_var.get('balls_played_first_inn', 0) > 0:
            if game_var['toss_result'] == "batting":
                player_1st_rr = round((game_var['player_runs_1stinn'] * 6) / game_var['balls_played_first_inn'], 2)
            else:
                ai_1st_rr = round((game_var['ai_runs_1stinn'] * 6) / game_var['balls_played_first_inn'], 2)
                
        if game_var.get('balls_played_sec_inn', 0) > 0:
            if game_var['toss_result'] == "bowling":
                player_2nd_rr = round((game_var['player_runs_2ndinn'] * 6) / game_var['balls_played_sec_inn'], 2)
            else:
                ai_2nd_rr = round((game_var['ai_runs_2ndinn'] * 6) / game_var['balls_played_sec_inn'], 2)
        
        player_match_runs = game_var.get('player_runs_1stinn', 0) + game_var.get('player_runs_2ndinn', 0)
        player_match_balls_faced = game_var.get("balls_played_by_player", 0)
        player_match_wickets_taken = game_var.get('current_match_wickets', 0) 
        player_match_runs_conceded = game_var.get('player_runs_conceded_1stinn', 0) + game_var.get('player_runs_conceded_2ndinn', 0)
        player_match_balls_bowled = game_var.get('player_balls_bowled_1stinn', 0) + game_var.get('player_balls_bowled_2ndinn', 0)

        player_match_strike_rate = self.calculate_strike_rate(player_match_runs, player_match_balls_faced)
        player_match_economy_rate = self.calculate_economy_rate(player_match_runs_conceded, player_match_balls_bowled)

        return {
            "Match Number": game_var.get('total_matches_played', 0),
            "Player Runs 1st Innings": game_var.get('player_runs_1stinn', 0),
            "AI Runs 1st Innings": game_var.get('ai_runs_1stinn', 0),
            "Player Runs 2nd Innings": game_var.get('player_runs_2ndinn', 0),
            "AI Runs 2nd Innings": game_var.get('ai_runs_2ndinn', 0),
            "Player 1st Inn Run Rate": player_1st_rr,
            "Player 2nd Inn Run Rate": player_2nd_rr,
            "AI 1st Inn Run Rate": ai_1st_rr,
            "AI 2nd Inn Run Rate": ai_2nd_rr,
            "Player Match Strike Rate": player_match_strike_rate,
            "Player Match Economy Rate": player_match_economy_rate,
            "Player Match Wickets Taken": player_match_wickets_taken, 
            "Player Match Runs Conceded": player_match_runs_conceded,
            "Player Match Balls Bowled": player_match_balls_bowled,
            "Toss Result": game_var.get('toss_result', ""),
            "Balls Played 1st Innings": game_var.get('balls_played_first_inn', 0),
            "Balls Played 2nd Innings": game_var.get('balls_played_sec_inn', 0),
            "Difficulty": game_var.get('difficulty', 'medium'),
            "Match Result": game_var.get('match_result', 'unknown')
        }
class GamePlay:
    def __init__(self, player_name, min_choice=0, max_choice=10, score_alignment=80, difficulty = None, commentator = None):
        self.name = player_name
        self.min_choice = min_choice
        self.max_choice = max_choice
        self.score_alignment = score_alignment
        self.difficulty = difficulty
        self.commentator = commentator

    def input_num(self):
        while True:
            try:
                console.print(f"\n[bold yellow]Enter a number between[/bold yellow] {self.min_choice} [bold yellow]and[/bold yellow] {self.max_choice}[bold yellow]:[/bold yellow] ", end="")
                player_choice = int(input(f""))
                sounds.play_click()
                if player_choice < self.min_choice or player_choice > self.max_choice:
                    console.print("[red]Invalid number[/red]")
                    continue
                return player_choice
            except ValueError:
                console.print("[red]Invalid number[/red]")
            except KeyboardInterrupt:
                console.print("\n[red]Can't quit here[/red]")

    def get_ai_choice(self, player_prev_choice=None, ai_role=None):
        if self.difficulty == "easy":
            return random.randint(self.min_choice, self.max_choice)

        elif self.difficulty == "medium":
            choices = list(range(self.min_choice, self.max_choice + 1))
            if ai_role == "batting":
                if player_prev_choice in choices:
                    choices.remove(player_prev_choice)
            elif ai_role == "bowling":
                if player_prev_choice in choices:
                    choices += [player_prev_choice] * 2
            return random.choice(choices)

        elif self.difficulty == "hard":
            if ai_role == "bowling":
                if player_prev_choice != 0 and random.random() < 0.7:
                    return player_prev_choice
                else:
                    return random.randint(self.min_choice, self.max_choice)
            elif ai_role == "batting":
                if player_prev_choice != 0 and random.random() < 0.7:
                    avoid = [i for i in range(self.min_choice, self.max_choice + 1) if i != player_prev_choice]
                    return random.choice(avoid)
                else:
                    return random.randint(self.min_choice, self.max_choice)
        return random.randint(self.min_choice, self.max_choice)

    def calculate_run_rate(self, runs, balls_played):
        if balls_played == 0:
            return 0.0
        return round((runs * 6) / balls_played, 2)

    def calculate_required_run_rate(self, runs_needed, balls_remaining):
        if balls_remaining == 0:
            return 0.0 if runs_needed <= 0 else float('inf')
        return round((runs_needed * 6) / balls_remaining, 2)

    def display_run_rate_after_over(self, runs, balls_played, player_name="Player"):
        if balls_played > 0 and balls_played % 6 == 0:  
            run_rate = self.calculate_run_rate(runs, balls_played)
            over_number = balls_played // 6
            console.print(f"\n[green]---[/green] [magenta]After Over {over_number}[/magenta] [green]---[/green]")
            console.print(f"[blue]{player_name}[/blue] [green]Run Rate:[/green] [{run_rate} [green]runs per over[/green]")
            console.print(f"[green]Total:[/green] {runs} [green]runs in[/green] {balls_played} [green]balls[/green] ({over_number} [green]overs[/green])")
            console.print("[green]=[/green]" * 35)

    def calculate_strike_rate(self, runs, balls_faced):
        if balls_faced == 0:
            return 0.0
        return round((runs / balls_faced) * 100, 2)
    
    def calculate_economy_rate(self, runs_conceded, balls_bowled):
        if balls_bowled == 0:
            return 0.0
        return round((runs_conceded / balls_bowled) * 6, 2)
    
    def check_milestones(self, current_runs, prev_runs, game_var):
        if prev_runs < 50 <= current_runs:
            game_var['half_centuries'] += 1
            if self.commentator and self.commentator.enabled:
                self.commentator.milestone_commentary(self.name, 50)
        if prev_runs < 100 <= current_runs:
            game_var['centuries'] += 1
            if self.commentator and self.commentator.enabled:
                self.commentator.milestone_commentary(self.name, 100)

    def intro(self):
        console.print(f"\n[bold magenta]Welcome to AVG Hand Cricket,[/bold magenta] [bold blue]{self.name}[/bold blue] [bold magenta]![/bold magenta]")
        console.print("[red]- Prattik | XII-'A' SAP JEE[/red]".rjust(self.score_alignment))

    def match_over(self, game_var):
        while True:
            try:
                console.print(f"\n[bold yellow]Enter number of overs:[/bold yellow] ", end="")
                game_var['over'] = int(input(""))
                sounds.play_click()
                if game_var['over'] < 1:
                    console.print("[dark_grey]Enter a number greater than 0[/dark_grey]")
                    continue
                return game_var['over']
            except ValueError:
                console.print("[bold red]Please enter a valid integer[/bold red]")

    def toss(self, game_var):
        while True:
            console.print(f"\n[bold yellow]Choose Odd or Even:[/bold yellow] ", end="")
            player_decision = input("").strip().lower()
            sounds.play_click()
            if player_decision in ["even", "odd"]:
                ai_decision = "odd" if player_decision == "even" else "even"
                break
            else:
                console.print("[red]Invalid odd or even[/red]")

        console.print(f"[blue]{self.name}[/blue] [green]chose[/green] {player_decision} [green]| AI chose[/green] {ai_decision}")        

        while True:
            try:
                console.print(f"\n[bold yellow]Choose a number between {self.min_choice} and {self.max_choice}:[/bold yellow] ", end="")
                player_hand = int(input(""))
                sounds.play_click()
                if self.min_choice <= player_hand <= self.max_choice:
                    break
                else:
                    console.print("[red]Invalid number[/red]")
            except ValueError:
                console.print("[red]Invalid number[/red]")

        ai_hand = random.randint(self.min_choice, self.max_choice)

        console.print(f"[blue]{self.name}[/blue] [green]chose[/green] {player_hand} [green]| AI chose[/green] {ai_hand}")

        role_picked = player_hand + ai_hand

        if player_decision == "odd":
            if role_picked % 2 == 0:
                ai_role = random.choice(["Bat", "Bowl"])
                console.print(f"[magenta]AI chose to[/magenta] {ai_role}")

                if ai_role == "Bat":
                    if self.commentator and self.commentator.enabled:
                        self.commentator.toss_commentary("AI", ai_role.lower())
                    console.print(f"\n[blue]{self.name}[/blue] [magenta]is bowling[/magenta] \n[blue]AI[/blue] [magenta]is batting[/magenta]")
                    game_var['toss_result'] = "bowling"
                else:
                    if self.commentator and self.commentator.enabled:
                        self.commentator.toss_commentary("AI", ai_role.lower())
                    console.print(f"\n[blue]{self.name}[/blue] [magenta]is batting[/magenta] \n[blue]AI[/blue] [magenta]is bowling[/magenta]")
                    game_var['toss_result'] = "batting"
            else:
                while True:
                    console.print(f"\n[bold yellow]Choose to bat or bowl?[/bold yellow] ", end="")
                    player_role = input("").strip().lower()
                    if player_role == "bat":
                        if self.commentator and self.commentator.enabled:
                            self.commentator.toss_commentary(self.name, player_role)
                        console.print(f"\n[blue]{self.name}[/blue] [magenta]is batting[/magenta] \n[blue]AI[/blue] [magenta]is bowling[/magenta]")
                        game_var['toss_result'] = "batting"
                        break
                    elif player_role == "bowl":
                        if self.commentator and self.commentator.enabled:
                            self.commentator.toss_commentary(self.name, player_role)
                        console.print(f"\n[blue]{self.name}[/blue] [magenta]is bowling[/magenta] \n[blue]AI[/blue] [magenta]is batting[/magenta]")
                        game_var['toss_result'] = "bowling"
                        break
                    else:
                        console.print("[red]Invalid bat or bowl[/red]")
        else:
            if role_picked % 2 == 0:
                while True:
                    console.print(f"\n[bold yellow]Choose to bat or bowl?[/bold yellow] ", end="")
                    player_role = input("").strip().lower()
                    sounds.play_click()

                    if player_role == "bat":
                        if self.commentator and self.commentator.enabled:
                            self.commentator.toss_commentary(self.name, player_role)
                        console.print(f"\n[blue]{self.name}[/blue] [magenta]is batting[/magenta] \n[blue]AI[/blue] [magenta]is bowling[/magenta]")
                        game_var['toss_result'] = "batting"
                        break
                    elif player_role == "bowl":
                        if self.commentator and self.commentator.enabled:
                            self.commentator.toss_commentary(self.name, player_role)
                        console.print(f"\n[blue]{self.name}[/blue] [magenta]is bowling[/magenta] \n[blue]AI[/blue] [magenta]is batting[/magenta]")
                        game_var['toss_result'] = "bowling"
                        break
                    else:
                        console.print("[red]Invalid bat or bowl[/red]")
            else:
                ai_role = random.choice(["Bat", "Bowl"])
                if ai_role == "Bat":
                    if self.commentator and self.commentator.enabled:
                        self.commentator.toss_commentary("AI", ai_role.lower())
                    console.print(f"\n[blue]{self.name}[/blue] [magenta]is bowling[/magenta] \n[blue]AI[/blue] [magenta]is batting[/magenta]")
                    game_var['toss_result'] = "bowling"
                else:
                    if self.commentator and self.commentator.enabled:
                        self.commentator.toss_commentary("AI", ai_role.lower())
                    console.print(f"\n[blue]{self.name}[/blue] [magneta]is batting[/magenta] \n[blue]AI[/blue] [magenta]is bowling[/magenta]")
                    game_var['toss_result'] = "batting"

        return game_var['toss_result']

    def one_or_two(self):
        console.print("[green]One or Two[/green]")
        while True:
            try:
                console.print(f"\n[bold yellow]Choose 1 or 2:[/bold yellow] ", end="")
                player_12decision = int(input(""))
                sounds.play_click()
                if player_12decision in [1, 2]:
                    break
                else:
                    console.print("[red]Invalid input[/red]")
            except ValueError:
                console.print("[red]Invalid input[/red]")

        ai_12decision = random.choice([1, 2])

        console.print(f"[blue]{self.name}[/blue] [green]chose[/green] {player_12decision} [green]|[/green] [blue]AI[/blue] [green]chose[/green] {ai_12decision}")
        return player_12decision, ai_12decision

    def first_in(self, game_var):
        console.print("\n[bold yellow]First Innings:[/bold yellow]")

        if game_var['toss_result'] == "bowling":
            for ball in range(1, game_var['balls'] + 1):
                player_choice = self.input_num()
                ai_role = "bowling" if game_var['toss_result'] == "batting" else "batting"
                ai_choice = self.get_ai_choice(player_choice, ai_role)

                console.print(f"[yellow]Ball {ball}:[/yellow] [blue]{self.name}[/blue] [green]chose[/green]{player_choice} [green]|[/green] [blue]AI[/blue] [green]chose[/green] {ai_choice}")

                game_var["balls_played_first_inn"] += 1
                game_var["player_balls_bowled_1stinn"] += 1

                if player_choice == ai_choice and player_choice != 0:
                    self.commentator.wicket_commentary("AI", bowler = self.name)
                    console.print("[bold yellow]First Innings Over[/bold yellow]")
                    game_var['lifetime_wickets'] += 1
                    game_var['current_match_wickets'] += 1
                    break

                if player_choice == 0 and ai_choice == 0:
                    player_12decision, ai_12decision = self.one_or_two()
                    if player_12decision == ai_12decision:
                        self.commentator.wicket_commentary("AI", bowler = self.name)
                        console.print(f"[bold yellow]First Innings Over[/bold yellow] \n[blue]AI[/blue] [green]has hit[/green] {game_var['ai_runs_1stinn']} [green]runs[/green]")
                        game_var['lifetime_wickets'] += 1
                        game_var['current_match_wickets'] += 1
                        break

                game_var['ai_runs_1stinn'] += ai_choice
                game_var['player_runs_conceded_1stinn'] += ai_choice
                if self.commentator and self.commentator.enabled:
                    self.commentator.run_commentary(ai_choice)
                console.print(f"[green]Current AI runs:[/green] {game_var['ai_runs_1stinn']}".rjust(self.score_alignment))

                self.display_run_rate_after_over(
                    game_var['ai_runs_1stinn'], 
                    game_var["balls_played_first_inn"], 
                    "AI"
                )

            console.print(f"\n[blue]AI[/blue] [green]as hit[/green] {game_var['ai_runs_1stinn']} [green]runs[/green]")
            final_run_rate = self.calculate_run_rate(game_var['ai_runs_1stinn'], game_var["balls_played_first_inn"])
            console.print(f"[blue]AI[/blue] [yellow]Final Run Rate:[/yellow] {final_run_rate} [green]runs per over[/green]")
            console.print(f"\n[blue]{self.name}[/blue] [magenta]needs[/magenta] {game_var['ai_runs_1stinn'] + 1} [magenta]runs to win[/magenta]")

        if game_var['toss_result'] == "batting":
            for ball in range(1, game_var['balls'] + 1):
                player_choice = self.input_num()
                ai_role = "bowling" if game_var['toss_result'] == "batting" else "batting"
                ai_choice = self.get_ai_choice(player_choice, ai_role)

                console.print(f"[yellow]Ball {ball}:[/yellow] [blue]{self.name}[/blue] [green]chose[/green] {player_choice} [green]|[/green] [blue]AI[blue] [green]chose[/green] {ai_choice}")

                game_var["balls_played_first_inn"] += 1
                game_var["balls_played_by_player"] += 1

                if player_choice == ai_choice and player_choice != 0:
                    self.commentator.wicket_commentary(self.name, bowler = "AI")
                    console.print(f"[bold yellow]First Innings Over[/bold yellow] \n[blue]{self.name}[/blue] [green]has hit[/green] {game_var['player_runs_1stinn']} [green]runs[/green]")
                    break

                if player_choice == 0 and ai_choice == 0:
                    player_12decision, ai_12decision = self.one_or_two()
                    if player_12decision == ai_12decision:
                        self.commentator.wicket_commentary(self.name, bowler = "AI")
                        console.print(f"\n[bold yellow]First Innings Over[/bold yellow] \n[blue]{self.name}[/blue] [green]has hit[/green] {game_var['player_runs_1stinn']} [green]runs[/green]")
                        break
                    else:
                        continue

                prev_player_runs = game_var['player_runs_1stinn']            
                game_var['player_runs_1stinn'] += player_choice
                game_var['lifetime_runs'] += player_choice
                self.check_milestones(game_var['player_runs_1stinn'], prev_player_runs, game_var)

                if self.commentator and self.commentator.enabled:
                    self.commentator.run_commentary(player_choice)
                console.print(f"[yellow]Current player runs:[/yellow] {game_var['player_runs_1stinn']}".rjust(self.score_alignment))

                self.display_run_rate_after_over(
                    game_var['player_runs_1stinn'], 
                    game_var["balls_played_first_inn"], 
                    self.name
                )

            final_run_rate = self.calculate_run_rate(game_var['player_runs_1stinn'], game_var["balls_played_first_inn"])
            console.print(f"\n[blue]{self.name}[/blue] [yellow]Final Run Rate:[/yellow] {final_run_rate} [green]runs per over[/green]")
            console.print(f"\n[blue]AI[/blue] [magenta]needs[/magenta] {game_var['player_runs_1stinn'] + 1} [magenta]runs to win[/magenta]")

    def second_in(self, game_var):
        player_choice = 0
        ai_choice = 0
        player_12decision = 0
        ai_12decision = 0

        if game_var['toss_result'] == "bowling":
            console.print("\n[bold yellow]Second Innings:[/bold yellow]")
            console.print(f"\n[blue]{self.name}[/blue] [magenta]is now batting[/magenta] and [blue]AI[/blue] [magenta]is bowling[/magenta]")

            for i in range(1, game_var['balls'] + 1):
                if game_var['player_runs_2ndinn'] > game_var['ai_runs_1stinn']:
                    console.print(f"[blue]{self.name}[/blue] [bold yellow]has achieved the target runs[/bold yellow]")
                    game_var['match_result'] = "win"
                    console.print("\n[bold yellow]Second Innings Over[/bold yellow]")
                    break

                player_choice = self.input_num()
                ai_role = "bowling" if game_var['toss_result'] == "batting" else "batting"
                ai_choice = self.get_ai_choice(player_choice, ai_role)

                console.print(f"[yellow]Ball {i}:[/yellow] [blue]{self.name}[/blue] [green]chose[/green] {player_choice} [green]|[/green] [blue]AI[/blue] [green]chose[/green] {ai_choice}")

                game_var['balls_played_sec_inn'] += 1
                game_var["balls_played_by_player"] += 1


                if player_choice == ai_choice and player_choice != 0:
                    self.commentator.wicket_commentary(self.name, bowler = "AI")
                    console.print(f"\n[bold yellow]Second Innings Over[/bold yellow] \n[blue]{self.name}[/blue] [green]has hit[/green] {game_var['player_runs_2ndinn']} [green]runs[/green]")
                    break

                if player_choice == 0 and ai_choice == 0:
                    player_12decision, ai_12decision = self.one_or_two()
                    if player_12decision == ai_12decision:
                        self.commentator.wicket_commentary(self.name, bowler = "AI")
                        console.print(f"\n[bold yellow]Second Innings Over[/bold yellow] \n[blue]{self.name}[/blue] [green]has hit[/green] {game_var['player_runs_2ndinn']} [green]runs[/green]")
                        break
                    else:
                        continue

                prev_player_runs = game_var['player_runs_2ndinn']
                game_var['player_runs_2ndinn'] += player_choice
                game_var['lifetime_runs'] += player_choice
                self.check_milestones(game_var['player_runs_2ndinn'], prev_player_runs, game_var)
                if self.commentator and self.commentator.enabled:
                    self.commentator.run_commentary(player_choice)
                console.print(f"[yellow]Current player runs:[/yellow] {game_var['player_runs_2ndinn']}".rjust(self.score_alignment))
                if i % 6 == 0: 
                    current_run_rate = self.calculate_run_rate(game_var['player_runs_2ndinn'], game_var['balls_played_sec_inn'])
                    required_run_rate = self.calculate_required_run_rate(
                        game_var['ai_runs_1stinn'] + 1 - game_var['player_runs_2ndinn'], 
                        game_var['balls'] - game_var['balls_played_sec_inn']
                    )
                    console.print(f"\n[magenta]-- After Over[/magenta] [green]{i//6}[/green] [magenta]---[/magenta]")
                    console.print(f"[magenta]Current Run Rate:[/magenta] [green]{current_run_rate}[/green]")
                    console.print(f"[magenta]Required Run Rate:[/magenta] [green]{required_run_rate}[/green]")
                    console.print("[magenta]=[/magenta]" * 35)

            console.print(f"\n[blue]{self.name}[/blue] [green]has hit[/green] {game_var['player_runs_2ndinn']} [green]runs[/green]")

        elif game_var['toss_result'] == "batting":
            console.print("\n[bold yellow]Second Innings:[/bold yellow]")
            console.print(f"\n[blue]{self.name}[/blue] [magenta]is now bowling[/magenta] and [blue]AI[/blue] [magenta]is batting[/magenta]")

            for i in range(1, game_var['balls'] + 1):
                if game_var['ai_runs_2ndinn'] > game_var['player_runs_1stinn']:
                    console.print("[blue]AI[/blue] [bold yellow]has achieved the target runs[/bold yellow]")
                    game_var['match_result'] = "loss" 
                    console.print("\n[bold yellow]Second Innings Over[/bold yellow]")
                    break

                player_choice = self.input_num()
                ai_role = "bowling" if game_var['toss_result'] == "batting" else "batting"
                ai_choice = self.get_ai_choice(player_choice, ai_role)

                console.print(f"[yellow]Ball {i}:[/yellow] [blue]{self.name}[/blue] [green]chose[/green] {player_choice} [green]|[/green] [blue]AI[/blue] [green]chose[/green] {ai_choice}")

                game_var["balls_played_sec_inn"] += 1
                game_var["player_balls_bowled_2ndinn"] += 1

                if player_choice == ai_choice and player_choice != 0:
                    self.commentator.wicket_commentary("AI", bowler = self.name)
                    console.print(f"\n[bold yellow]Second Innings Over[/bold yellow]")
                    game_var['lifetime_wickets'] += 1
                    game_var['current_match_wickets'] += 1
                    break

                if player_choice == 0 and ai_choice == 0:
                    player_12decision, ai_12decision = self.one_or_two()
                    if player_12decision == ai_12decision:
                        self.commentator.wicket_commentary("AI", bowler = self.name)
                        console.print(f"\n[bold yellow]Second Innings Over[/bold yellow] \n[blue]AI[/blue] [green]has hit[/green] {game_var['ai_runs_2ndinn']} [green]runs[/green]")
                        game_var['lifetime_wickets'] += 1
                        game_var['current_match_wickets'] += 1
                        break
                    else:
                        continue

                game_var['ai_runs_2ndinn'] += ai_choice
                game_var['player_runs_conceded_2ndinn'] += ai_choice
                if self.commentator and self.commentator.enabled:
                    self.commentator.run_commentary(ai_choice)
                console.print(f"[yellow]Current AI runs:[/yellow] {game_var['ai_runs_2ndinn']}".rjust(self.score_alignment))
                if i % 6 == 0:  
                    current_run_rate = self.calculate_run_rate(game_var['player_runs_2ndinn'], game_var['balls_played_sec_inn'])
                    required_run_rate = self.calculate_required_run_rate(
                        game_var['player_runs_1stinn'] + 1 - game_var['ai_runs_2ndinn'], 
                        game_var['balls'] - game_var['balls_played_sec_inn']
                    )
                    console.print(f"\n[magenta]--- After Over[/magenta] [green]{i//6}[/green] [magenta]---[/magenta]")
                    console.print(f"[magenta]Current Run Rate:[/magenta] [green]{current_run_rate}[/green]")
                    console.print(f"[magenta]Required Run Rate:[/magenta] [green]{required_run_rate}[/green]")
                    console.print("[magenta]=[/magenta]" * 35)

            console.print(f"\n[blue]AI[/blue] [green]has hit[/green] {game_var['ai_runs_2ndinn']} [green]runs[/green]")

    def super_over(self, game_var):
        while True:
            console.print("\n[magenta]*** SUPER OVER! ***[/magenta]")
            console.print("[yellow]It’s a super over! Only[/yellow] [green]1 over[/green][yellow], whoever wins - wins the match![/yellow]\n")

            player_super_runs = 0
            ai_super_runs = 0

            console.print(f"[blue]{self.name}[/blue] [green]is batting first in the[/green] [magenta]Super Over[/magenta][green]![/green]")
            for ball in range(1, game_var['balls'] + 1):
                player_choice = self.input_num()
                ai_choice = self.get_ai_choice(player_choice, ai_role="bowling")
                console.print(f"[blue]{self.name}[/blue] [green]chose[/green] {player_choice} [green]|[/green] [blue]AI[/blue] [green]chose[/green] {ai_choice}")

                game_var["balls_played_by_player"] += 1

                if player_choice == ai_choice:
                    console.print("[bold red]OUT![/bold red] [red]You have lost your wicket![/red]")
                    self.commentator.wicket_commentary(self.name, bowler = "AI")
                    console.print(f"[blue]{self.name}[/blue] [green]has hit[/green] {player_super_runs} [green]runs[/green]")
                    game_var['lifetime_runs'] += player_super_runs
                    break
                else:
                    player_super_runs += player_choice
                    console.print(f"[yellow]Current player runs:[/yellow] {player_super_runs}".rjust(self.score_alignment))

            console.print(f"\n[blue]AI[/blue] [magenta]is batting now[/magenta] \n[blue]AI[/blue] [green]has to hit[/green] {player_super_runs + 1} [green]to win[/green]")
            for ball in range(1, game_var['balls'] + 1):
                if ai_super_runs > player_super_runs:
                    break
                player_choice = self.input_num()
                ai_choice = self.get_ai_choice(player_choice, ai_role="batting")
                console.print(f"[blue]{self.name}[/blue] [green]chose[/green] {player_choice} [green]|[/green] [blue]AI[/blue] [green]chose[/green] {ai_choice}")

                if player_choice == ai_choice:
                    console.print("[bold red]OUT![/bold red] [blue]AI[/blue] [red]has lost its wicket![/red]")
                    self.commentator.wicket_commentary("AI", bowler = self.name)
                    console.print(f"[blue]AI[/blue] [green]has hit[/green] {ai_super_runs} [green]runs[/green]")
                    game_var['lifetime_wickets'] += 1
                    game_var['current_match_wickets'] += 1
                else:
                    ai_super_runs += ai_choice
                    console.print(f"[yellow]Current AI runs:[/yellow] {ai_super_runs}".rjust(self.score_alignment))

            console.print("\n[bold magenta]Super Over Results:[/bold magenta]")
            console.print(f"[blue]{self.name}'s[/blue] [magenta]runs:[/magenta] [green]{player_super_runs}[/green] [magenta]|[/magenta] [blue]AI[/blue] [magenta]runs:[/magenta] [green]{ai_super_runs}[/green]")

            if player_super_runs > ai_super_runs:
                console.print(f"[bold yellow]{self.name} has won the [/bold yellow][bold magenta]Super Over[/bold magenta][bold yellow]![/bold yellow]")
                game_var['match_result'] = "win" 
                game_var['total_wins'] += 1
                game_var['match_result'] = "win"
                break
            elif ai_super_runs > player_super_runs:
                console.print("[bold yellow]AI has won the [/bold yellow][bold magenta]Super Over[/bold magenta][bold yellow]![/bold yellow]")
                game_var['match_result'] = "loss" 
                game_var['total_losses'] += 1
                game_var['match_result'] = "loss"
                break
            else:
                console.print("[bold magenta]Super Over[/bold magenta] [green]has resulted in a tie! Another[/green] [magenta]super over[/magenta] [green]will be played![/green]")
class DisplayManager:
    def print_match_summary(self, match_summary):
        console.print(f"[bold magenta]---------- Match Summary ----------[/bold magenta]")
        for inning, stats in match_summary.items():
            console.print(f"[green]{inning}:[/green]")
            for key, value in stats.items():
                if key == "run_rate":
                    console.print(f"[magenta]Run rate:[/magenta] {value} [green]runs per over[/green]")
                else:
                    score_name = key.replace("_", " ").capitalize()
                    console.print(f"[magenta]{score_name}:[/magenta] [green]{value}[/green]")
            print()
        console.print(f"[bold magenta]-----------------------------------[/bold magenta]")

    def print_player_profile(self, profile):
        console.print(f"\n[bold magenta]------- Player Profile -------[/bold magenta]")
        for key, value in profile.items():
            score_name = key.replace("_", " ").capitalize()
            console.print(f"[magenta]{score_name}:[/magenta] [green]{value}[/green]")
        print("[bold magenta]-----------------------------[/bold magenta]")
class GameManager:
    global game_var

    def __init__(self, gameplay_instance, display_instance, db_handler, manager, player_name):
        self.gameplay = gameplay_instance
        self.display = display_instance
        self.db_handler = db_handler
        self.manager = manager
        self.name = player_name

    def result(self, game_var):
        summary = game_var['match_summary']
        if game_var['toss_result'] == "bowling":
            ai_run_rate = self.gameplay.calculate_run_rate(
                game_var['ai_runs_1stinn'], 
                game_var['balls_played_first_inn']
            )
            player_run_rate = self.gameplay.calculate_run_rate(
                game_var['player_runs_2ndinn'], 
                game_var['balls_played_sec_inn']
            )
            summary['First Innings'] = {
                "ai_runs": game_var['ai_runs_1stinn'],
                "balls_played": game_var['balls_played_first_inn'],
                "run_rate": ai_run_rate
            }
            summary['Second Innings'] = {
                "player_runs": game_var['player_runs_2ndinn'],
                "balls_played": game_var['balls_played_sec_inn'],
                "run_rate": player_run_rate
            }

            total_runs = game_var['ai_runs_1stinn'] + game_var['player_runs_2ndinn']
            total_balls = game_var['balls_played_first_inn'] + game_var['balls_played_sec_inn']
            match_avg_run_rate = self.gameplay.calculate_run_rate(total_runs, total_balls)
            
            player_match_runs = game_var.get('player_runs_1stinn', 0) + game_var.get('player_runs_2ndinn', 0)
            player_match_balls_faced = game_var.get("balls_played_by_player", 0)
            player_match_runs_conceded = game_var.get('player_runs_conceded_1stinn', 0) + game_var.get('player_runs_conceded_2ndinn', 0)
            player_match_balls_bowled = game_var.get('player_balls_bowled_1stinn', 0) + game_var.get('player_balls_bowled_2ndinn', 0)
            player_match_strike_rate = self.gameplay.calculate_strike_rate(player_match_runs, player_match_balls_faced)
            player_match_economy_rate = self.gameplay.calculate_economy_rate(player_match_runs_conceded, player_match_balls_bowled)


            summary['Match Statistics'] = {
                "average_run_rate": match_avg_run_rate,
                "total_runs_scored": total_runs,
                "total_balls_played": total_balls,
                "Player Match Strike Rate": player_match_strike_rate,
                "Player Match Economy Rate": player_match_economy_rate,
            }

            self.display.print_match_summary(summary)
            if game_var['ai_runs_1stinn'] > game_var['player_runs_2ndinn']:
                console.print("[blue]AI[/blue] [green]has won the game.[/green] [blue]You[/blue] [red]have lost.[/red]")
                game_var['total_losses'] += 1
                game_var['match_result'] = "loss"

            elif game_var['ai_runs_1stinn'] < game_var['player_runs_2ndinn']:
                print("[blue]You[/blue] [green]have won the game.[/green] [blue]AI[/blue] [red]has lost.[/red]")
                game_var['total_wins'] += 1
                game_var['match_result'] = "win"

            else:
                console.print("[yellow]It's a tie![/yellow]")
                game_var['total_draws'] += 1
                self.gameplay.super_over(game_var)

            game_var['lifetime_balls_faced'] += player_match_balls_faced
            game_var['lifetime_balls_bowled'] += player_match_balls_bowled
            game_var['lifetime_runs_conceded'] += player_match_runs_conceded

            current_match_data = self.manager.get_current_match_data(game_var)
            self.db_handler.save_match_data(game_var, self.name, player_match_runs_conceded, player_match_balls_bowled, game_var['current_match_wickets'])
            self.handle_profile_display_and_save(game_var, current_match_data)

        elif game_var['toss_result'] == "batting":
            ai_run_rate = self.gameplay.calculate_run_rate(
                game_var['ai_runs_2ndinn'], 
                game_var['balls_played_sec_inn']
            )
            player_run_rate = self.gameplay.calculate_run_rate(
                game_var['player_runs_1stinn'], 
                game_var['balls_played_first_inn']
            )
            summary['First Innings'] = {
                "player_runs": game_var['player_runs_1stinn'],
                "balls_played": game_var['balls_played_first_inn'],
                "run_rate": player_run_rate
            }
            summary['Second Innings'] = {
                "ai_runs": game_var['ai_runs_2ndinn'],
                "balls_played": game_var['balls_played_sec_inn'],
                "run_rate": ai_run_rate
            }

            total_runs = game_var['player_runs_1stinn'] + game_var['ai_runs_2ndinn']
            total_balls = game_var['balls_played_first_inn'] + game_var['balls_played_sec_inn']
            match_avg_run_rate = self.gameplay.calculate_run_rate(total_runs, total_balls)
            
            player_match_runs = game_var.get('player_runs_1stinn', 0) + game_var.get('player_runs_2ndinn', 0)
            player_match_balls_faced = game_var.get("balls_played_by_player", 0)
            player_match_runs_conceded = game_var.get('player_runs_conceded_1stinn', 0) + game_var.get('player_runs_conceded_2ndinn', 0)
            player_match_balls_bowled = game_var.get('player_balls_bowled_1stinn', 0) + game_var.get('player_balls_bowled_2ndinn', 0)
            player_match_strike_rate = self.gameplay.calculate_strike_rate(player_match_runs, player_match_balls_faced)
            player_match_economy_rate = self.gameplay.calculate_economy_rate(player_match_runs_conceded, player_match_balls_bowled)

            summary['Match Statistics'] = {
                "average_run_rate": match_avg_run_rate,
                "total_runs_scored": total_runs,
                "total_balls_played": total_balls,
                "Player Match Strike Rate": player_match_strike_rate,
                "Player Match Economy Rate": player_match_economy_rate,
            }

            self.display.print_match_summary(summary)
            if game_var['ai_runs_2ndinn'] > game_var['player_runs_1stinn']:
                console.print("[blue]AI[/blue] [green]has won the game.[/green] [blue]You[/blue] [red]have lost.[/red]")
                game_var['total_losses'] += 1
                game_var['match_result'] = "loss"

            elif game_var['ai_runs_2ndinn'] < game_var['player_runs_1stinn']:
                console.print("[blue]You[/blue] [green]have won the game.[/green] [blue]AI[/blue] [red]has lost.[/red]")
                game_var['total_wins'] += 1
                game_var['match_result'] = "win"
                
            else:
                console.print("[yellow]It's a tie![/yellow]")
                game_var['total_draws'] += 1
                self.gameplay.super_over(game_var)

            game_var['lifetime_balls_faced'] += player_match_balls_faced
            game_var['lifetime_balls_bowled'] += player_match_balls_bowled
            game_var['lifetime_runs_conceded'] += player_match_runs_conceded

            current_match_data = self.manager.get_current_match_data(game_var)
            self.db_handler.save_match_data(game_var, self.name, player_match_runs_conceded, player_match_balls_bowled, game_var['current_match_wickets'])
            self.handle_profile_display_and_save(game_var, current_match_data)

        else:
            console.print("[bold red]Error:[/bold red] toss_result [red]must be either 'bowling' or 'batting'.[/red]")

    def reset_game(self, game_var):
        game_var['over'] = 0
        game_var['player_runs_1stinn'] = 0
        game_var['ai_runs_1stinn'] = 0
        game_var['player_runs_2ndinn'] = 0
        game_var['ai_runs_2ndinn'] = 0
        game_var['toss_result'] = ""
        game_var['match_summary'] = {}
        game_var['balls_played_first_inn'] = 0
        game_var['balls_played_sec_inn'] = 0
        game_var['current_match_wickets'] = 0
        game_var['balls_played_by_player'] = 0
        game_var['player_balls_bowled_1stinn'] = 0
        game_var['player_balls_bowled_2ndinn'] = 0
        game_var['player_runs_conceded_1stinn'] = 0
        game_var['player_runs_conceded_2ndinn'] = 0
        game_var['match_result'] = ""

    def get_choice(self):
        while True:
            try:
                menu = Panel.fit("[bold cyan]\n-------------------AVG HANDCRICKET-------------------[/bold cyan]\n" 
                    "\n[green]Type[/green] [blue]tutorial[/blue][green] to view the tutorial.[/green]" 
                    "\n[green]Type[/green] [blue]play[/blue][green] to play a match.[/green]" 
                    "\n[green]Type[/green] [blue]quit[/blue][green] to quit.[/green]" 
                    "\n[green]Type[/green] [blue]profile[/blue][green] to view the player profile.[/green]" 
                    "\n[green]Type[/green] [blue]save[/blue][green] to save current game state.[/green]" 
                    "\n[green]Type[/green] [blue]load[/blue][green] to load previous game state.[/green]" 
                    "\n[green]Type[/green] [blue]leaderboard[/blue][green] to view global stats.[/green]" 
                    "\n[green]Type[/green] [blue]switch[/blue][green] to switch player account.[/green]" 
                    "\n[green]Type[/green] [blue]settings[/blue][green] to change commentary or difficulty.[/green]",
                    title="Main Menu", border_style="bright_blue"
                )
                console.print(menu)
                choice = input(": ").strip().lower()
                if choice in ("tutorial","play", "quit", "profile", "save", "load", "leaderboard", "switch", "settings"):
                    return choice
                console.print("[red]Invalid input[/red]")
            except KeyboardInterrupt:
                console.print("\n[red]Game was interrupted[/red] [dark_grey]type no to exit properly[/dark_grey]")
                continue

    def handle_profile_display_and_save(self, game_var, current_match_data=None):
        if game_var['total_matches_played'] > 0:
            game_var['average_runs_per_match'] = (
                game_var['lifetime_runs'] / game_var['total_matches_played']
            )

        lifetime_strike_rate = self.gameplay.calculate_strike_rate(
            game_var['lifetime_runs'],
            game_var['lifetime_balls_faced']
        )

        lifetime_economy_rate = self.gameplay.calculate_economy_rate(
            game_var['lifetime_runs_conceded'],
            game_var['lifetime_balls_bowled']
        )

        lifetime_run_rate = self.gameplay.calculate_run_rate(
            game_var['lifetime_runs'],
            game_var['lifetime_balls_faced']
        )

        game_var['player_profile'] = {
            "Lifetime runs": game_var['lifetime_runs'],
            "Lifetime wickets": game_var['lifetime_wickets'],
            "Total Matches Played": game_var['total_matches_played'],
            "Total Wins": game_var['total_wins'],
            "Total Losses": game_var['total_losses'],
            "Total Draws": game_var['total_draws'],
            "Average Runs per Match": game_var['average_runs_per_match'],
            "Lifetime Balls Faced": game_var['lifetime_balls_faced'],
            "Lifetime Balls Bowled": game_var['lifetime_balls_bowled'],
            "Lifetime Runs Conceded": game_var['lifetime_runs_conceded'],
            "Lifetime Run Rate": lifetime_run_rate,
            "Lifetime Strike Rate": lifetime_strike_rate,
            "Lifetime Economy Rate": lifetime_economy_rate,
            "Centuries": game_var['centuries'],
            "Half Centuries": game_var['half_centuries']
        }
        self.display.print_player_profile(game_var['player_profile'])
        pw = self.db_handler.get_password(self.name)
        self.db_handler.save_profile(game_var['player_profile'], self.name, pw)

        if current_match_data:
            self.manager.save_profile_and_match_data(game_var['player_profile'], current_match_data)

    def show_settings_menu(self, game_var):
        while True:
            settings = Panel.fit("[bold cyan]\n-------------------Settings Menu-------------------[/bold cyan]\n"
                "\n[green]Type[/green] [blue]1[/blue][green] to toggle commentary on/off.[/green]"
                "\n[green]Type[/green] [blue]2[/blue][green] to change difficulty (easy/medium/hard).[/green]"
                "\n[green]Type[/green] [blue]3[/blue][green] to change player name.[/green]"
                "\n[green]Type[/green] [blue]4[/blue][green] to change password.[/green]"
                "\n[green]Type[/green] [blue]5[/blue][green] to go back to main menu.[/green]",
                title="Settings Menu", border_style="bright_blue"
            )
            console.print(f"[cyan]Enter choice (1-5):[/cyan]")
            choice = input("").strip()

            if choice == "1":
                self.toggle_commentary(game_var)
            elif choice == "2":
                self.change_difficulty(game_var)
            elif choice == "3":
                self.change_player_name()
            elif choice == "4":
                self.change_password()
            elif choice == "5":
                break
            else:
                console.print("[red]Invalid choice.[/red] [dark_grey]Please select[/dark_grey] 1, 2, 3, 4, [dark_grey]or[/dark_grey] 5")

    def toggle_commentary(self, game_var):
        current = self.gameplay.commentator.enabled
        self.gameplay.commentator.enabled = not current
        game_var["commentary_enabled"] = self.gameplay.commentator.enabled
        status = "enabled" if self.gameplay.commentator.enabled else "disabled"
        console.print(f"[green]Commentary is now[/green] [cyan]{status}[/cyan][green].[/green]")

    def change_difficulty(self, game_var):
        while True:
            console.print("[green]Enter new difficulty[/green] [cyan](easy, medium, hard)[/cyan][green]:[/green]")
            difficulty = input("").strip().lower()
            if difficulty in ("easy", "medium", "hard"):
                game_var["difficulty"] = difficulty
                self.gameplay.difficulty = difficulty
                console.print(f"[green]Difficulty set to [cyan]{difficulty}[/cyan][green].[/green]")
                break
            else:
                console.print("[red]Invalid choice[/red] [dark_grey]Try again.[/dark_grey]")

    def change_player_name(self):
        global name
        db_handler = self.db_handler
        old_name = self.name

        
        console.print("[green]Enter new player name [/green][dark_grey](max 30 characters)[/dark_grey][green]:[/green]")
        new_name = input("").strip().title()
        if not new_name.replace(" ", "").isalpha():
            console.print("[red]Invalid name[/red] [dark_grey]Only letters and spaces allowed.[/dark_grey]")
            return
        if len(new_name) == 0 or len(new_name) > 30:
            console.print("[red]Name must be 1-30 characters.[/red]")
            return
        if new_name == old_name:
            console.print("[red]New name must not match the current name.[/red]")
            return
        if db_handler.player_exists(new_name):
            console.print("[red]This player name already exists. Please choose another.[/red]")
            return

        console.print(f"[green]To confirm changing name from[/green] [cyan]{old_name}[/cyan] [green]to[/green] [cyan]{new_name}[/cyan][green], please enter your current password:[/green]")
        old_pass = pwinput.pwinput("").strip()
        if old_pass != db_handler.get_password(old_name):
            console.print("[red]Incorrect password. Name not changed.[/red]")
            return

        db = db_handler.get_db()
        cursor = db.cursor()
        cursor.execute("UPDATE player_profile SET name = %s WHERE name = %s", (new_name, old_name))
        cursor.execute("UPDATE match_data SET player_name = %s WHERE player_name = %s", (new_name, old_name))
        db.commit()
        cursor.close()

        old_json_file = f"{self.manager.json_folder_path}/{old_name}_game_save.json"
        new_json_file = f"{self.manager.json_folder_path}/{new_name}_game_save.json"
    
        old_excel_file = f"{self.manager.excel_folder_path}/{old_name}_profile_excel.xlsx"
        new_excel_file = f"{self.manager.excel_folder_path}/{new_name}_profile_excel.xlsx"
    
        files_renamed = []

        try:
            if os.path.exists(old_json_file):
                os.rename(old_json_file, new_json_file)
                files_renamed.append("game save")
        except Exception as e:
            console.print(f"[bold red]Warning:[/bold red] [red]Could not rename JSON save file: {e}[/red]")
    
        try:
            if os.path.exists(old_excel_file):
                os.rename(old_excel_file, new_excel_file)
                files_renamed.append("Excel profile")
        except Exception as e:
            console.print(f"[bold red]Warning:[/bold red] [red]Could not rename Excel file: {e}[/red]")

        console.print(f"[green]Name changed from[/green] [blue]'{old_name}'[/blue] [green]to[/green] [blue]'{new_name}'[/blue] [green]successfully.[/green]")
        if files_renamed:
            console.print(f"[yellow]Renamed save files: {', '.join(files_renamed)}[/yellow]")

        self.manager.name = new_name
        self.name = new_name
        name = new_name

    def change_password(self):
        db_handler = self.db_handler
        name = self.name

        old_pass = pwinput.pwinput("Enter your current password: ").strip()
        if old_pass != db_handler.get_password(name):
            console.print("[red]Incorrect password. Password not changed.[/red]")
            return

        while True:
            new_pass = pwinput.pwinput("Enter new password: ").strip()
            if len(new_pass) < 3:
                console.print("[red]Password must be at least 3 characters long.[/red]")
                continue
            confirm = pwinput.pwinput("Confirm new password: ").strip()
            if new_pass != confirm:
                console.print("[red]Passwords do not match.[/red] [dark_grey]Try again.[/dark_grey]")
            else:
                db_handler.update_password(name, new_pass)
                console.print("[bold yellow]Password updated successfully.[/bold yellow]")
                break

    def handle_save_game_and_profile(self, game_var):
        self.manager.save_game_to_file(game_var)
        if game_var['total_matches_played'] > 0:
            game_var['average_runs_per_match'] = (
                game_var['lifetime_runs'] / game_var['total_matches_played']
            )
        profile = {
            "Player Name": self.name,
            "Lifetime runs": game_var['lifetime_runs'],
            "Lifetime wickets": game_var['lifetime_wickets'],
            "Total Matches Played": game_var['total_matches_played'],
            "Total Wins": game_var['total_wins'],
            "Total Losses": game_var['total_losses'],
            "Total Draws": game_var['total_draws'],
            "Average Runs per Match": game_var['average_runs_per_match'],
            "Lifetime Balls Faced": game_var['lifetime_balls_faced'],
            "Lifetime Balls Bowled": game_var['lifetime_balls_bowled'],
            "Lifetime Runs Conceded": game_var['lifetime_runs_conceded'],
            "Centuries": game_var['centuries'],
            "Half Centuries": game_var['half_centuries']
        }

        current_match_data = self.manager.get_current_match_data(game_var)
        self.manager.save_profile_and_match_data(profile, current_match_data)
class Commentator:
    def __init__(self, enabled=True):
        self.enabled = enabled

    def say(self, messages):
        if self.enabled:
            console.print("\n[bright blue]Commentator:[/bright blue]", random.choice(messages))

    def run_commentary(self, runs):
        if runs == 0:
            sounds.play_run()
            self.say(["That's a dot ball!", "No run scored.", "Good ball, no run."], style = "bold dark_grey italic")
        elif runs == 1:
            sounds.play_run()
            self.say(["Just a single.", "Quick single taken.", "Keeps the scoreboard ticking."], style = "bold dark_grey italic")
        elif runs == 2:
            sounds.play_run()
            self.say(["Two runs added! Great work they're putting it!", "Great running between the wickets."], style = "bold green italic")
        elif runs == 3:
            sounds.play_run()
            self.say(["Three runs! That's rare.", "Excellent placement and running."], style = "bold green italic")
        elif runs == 4:
            sounds.play_score()
            self.say(["That's a FOUR!", "Cracking shot through the gap!", "That races to the boundary!"], style = "bold cyan italic")
        elif runs == 5:
            sounds.play_score()
            self.say(["Five runs! Unbelievable!", "What a shot! That's five!", "They really wanted that five!"], style = "bold cyan italic")
        elif runs == 6:
            sounds.play_score()
            self.say(["SIX! What a hit!", "That's out of the park!", "Massive hit!"], style = "bold red italic")
        elif runs == 7:
            sounds.play_score()
            self.say(["Seven runs! great thinking here.", "They chose for a seven this time.", "They're looking far ahead"], style = "bold red italic")
        elif runs == 8:
            sounds.play_score()
            self.say(["Eight! Great decision.", "An Eight! He's confident.", "He just smashed it!"], style = "bold red italic")
        elif runs == 9:
            sounds.play_score()
            self.say(["Nine runs! they know what they're doing here.", "Makes me wanna sing a song right now.", "Thats it! They can go home happy now!"], style = "bold red italic")
        elif runs == 10:
            sounds.play_score()
            self.say(["A TEN!, What a shot!", "That's out of the planet!","Holy mother of GOD I'm SCREAMINGGGGG"], style = "bold red italic")

    def wicket_commentary(self, out_batsman, bowler = "AI"):
        sounds.play_wicket()
        self.say([
            f"{out_batsman} is OUT! Bowled by {bowler}!",
            f"{out_batsman} departs! What a delivery from {bowler}!",
            f"What a catch! {out_batsman} is gone, bowled by {bowler}!",
            f"{out_batsman} has been dismissed! {bowler} takes the wicket!",
            f"{out_batsman} is back to the pavilion! {bowler} strikes",
            f"{out_batsman} is out! {bowler} with a brilliant delivery!",
            f"{out_batsman} has been sent packing! {bowler} gets the wicket!",
            f"{out_batsman} is gone! {bowler} with a fantastic ball!",
            f"{out_batsman} is out! {bowler} takes the wicket with a superb delivery!",
            f"What a day he must be having right now! Superb bowling from {bowler}!",
            f"He must be feeling on top of the world right now! {bowler} gets the wicket!",
            f"He must've missed his morning grace! {out_batsman} is out!",
            f"{out_batsman} is out! {bowler} bowls a peach of a delivery!",
            f"{out_batsman} is out! {bowler} with a brilliant delivery!",
            f"Wicket! {bowler} gets rid of {out_batsman}!"
        ], style="bold red italic" )

    def toss_commentary(self, winner, decision):
        self.say([
            f"{winner} wins the toss and chooses to {decision}, Dunno what they're thinking.",
            f"The toss is won by {winner}, they opt to {decision}, Makes sense at this time of the day.",
            f"{winner} decides to {decision} after winning the toss, some quality decision making here."
        ], style = "cyan italic")

    def milestone_commentary(self, player_name, milestone):
        sounds.play_milestone()
        if self.enabled:
            if milestone == 50:
                self.say([
                    f"That's a brilliant fifty for {player_name}!",
                    f"{player_name} reaches his half-century with style!",
                    f"Fifty up for {player_name}! What an innings this is turning out to be.",
                    f"{player_name} brings up his fifty! A well-deserved milestone!",
                    f"Half-century for {player_name}! He is looking in great touch!",
                    f"{player_name} has reached his fifty! A fantastic effort!",
                    f"Fifty! {player_name} has played a superb innings so far!",
                    f"What a partnership this is turning out to be! {player_name} has reached his fifty!",
                    f"He lifts his bat up!!! What a moment for {player_name}!",
                ], style = "magenta italic")
            elif milestone == 100:
                self.say([
                    f"A magnificent century for {player_name}! Take a bow!",
                    f"And he lifts his bat up! Imagine how he must be feeling right now!",
                    f"This is the moment he has been waiting for! God Bless Him!",
                    f"{player_name} has shown the world yet again why he is a class player!",
                    f"{player_name} brings up his hundred! An outstanding display of batting!",
                    f"Hundred! {player_name} has reached the magical three-figure mark!",
                ], style = "bold magenta italic")
class LeaderboardManager:
    def __init__(self, db):
        self.db = db  

    def show_menu(self):
        while True:
            print("\n")
            lb = Panel.fit("[bold cyan]\n-------------------Leaderboard Menu-------------------[/bold cyan]\n"
                "\n[green]Type[/green] [blue]0[/blue][green] to show leaderboard info.[/green]"
                "\n[green]Type[/green] [blue]1[/blue][green] to view most runs scored.[/green]"
                "\n[green]Type[/green] [blue]2[/blue][green] to view most wickets taken.[/green]"
                "\n[green]Type[/green] [blue]3[/blue][green] to view most balls played.[/green]"
                "\n[green]Type[/green] [blue]4[/blue][green] to view most active players.[/green]"
                "\n[green]Type[/green] [blue]5[/blue][green] to view highest scores scored.[/green]"
                "\n[green]Type[/green] [blue]6[/blue][green] to view most win rate.[/green]"
                "\n[green]Type[/green] [blue]7[/blue][green] to view highest all-rounders scores.[/green]"
                "\n[green]Type[/green] [blue]8[/blue][green] to view most lifetime run rate.[/green]"
                "\n[green]Type[/green] [blue]9[/blue][green] to view most lifetime strike rate.[/green]"
                "\n[green]Type[/green] [blue]10[/blue][green] to view most lifetime economy rate.[/green]"
                "\n[green]Type[/green] [blue]11[/blue][green] to view most half centuries.[/green]"
                "\n[green]Type[/green] [blue]12[/blue][green] to view most centuries.[/green]"
                "\n[green]Type[/green] [blue]13[/blue][green] to return to the Main Menu.[/green]",
                title="Leaderboard Menu", border_style="bright_blue"
            )
            console.print(lb)
            console.print(f"[cyan]Enter choice (0-13):[/cyan]")
            choice = input("").strip()
            sounds.play_click()
            options = {
                "0": self.leaderboard_info,
                "1": self.show_runs,
                "2": self.show_wickets,
                "3": self.show_balls_played,
                "4": self.show_most_active,
                "5": self.show_high_scores,
                "6": self.show_win_rate,
                "7": self.show_all_rounders,
                "8": self.show_lifetime_run_rate,
                "9": self.show_lifetime_strike_rate, 
                "10": self.show_lifetime_economy_rate, 
                "11": self.show_half_centuries,
                "12": self.show_centuries
            }

            if choice in options:
                options[choice]()
            elif choice == "13":
                break
            else:
                print("Invalid choice! Try again.")

    def leaderboard_info(self):
        info = Panel.fit("[bold cyan]\n-------------------Leaderboard Info-------------------[/bold cyan]\n"
            "\n[green]The leaderboard showcases the top players based on various statistics from all matches played.[/green]"
            "\n[green]It includes categories such as most runs, wickets, balls played, highest scores, win rates, and more.[/green]"
            "\n[green]Players are ranked based on their performance in each category, providing a competitive overview of the best players.[/green]"
            "\n[green]Stats are updated after each match to reflect the latest performances.[/green]",
            title="Leaderboard Information", border_style="bright_blue"
        )
        print("\n")
        console.print(info)
        time.sleep(5)

    def show_runs(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name, SUM(runs) as total_runs
            FROM match_data
            GROUP BY player_name
            ORDER BY total_runs DESC
            LIMIT 10
        """)
        runresult = cursor.fetchall()
        cursor.close()

        runstable = Table(title="🏆 Top Run Scorers", show_header=True, header_style="bold yellow", box=box.ROUNDED, border_style="bright_blue")
        runstable.add_column("Rank", style="cyan", justify="center")
        runstable.add_column("Player", style="bold green")
        runstable.add_column("Runs", style="bold magenta", justify="right")

        for i, (player_name, total_runs) in enumerate(runresult, start=1):
            runstable.add_row(str(i), player_name, str(total_runs))
        print("\n")
        console.print(runstable)
        time.sleep(5)

    def show_wickets(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name, SUM(wickets) as total_wickets
            FROM match_data
            GROUP BY player_name
            ORDER BY total_wickets DESC
            LIMIT 10
        """)
        wicketresults = cursor.fetchall()
        cursor.close()

        wickettable = Table(
            title="🎯 Top Wicket Takers",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_red",
            box=box.ROUNDED
        )
        wickettable.add_column("Rank", style="cyan", justify="center")
        wickettable.add_column("Player", style="bold green")
        wickettable.add_column("Wickets", style="bold magenta", justify="right")

        for i, (player_name, total_wickets) in enumerate(wicketresults, start=1):
            wickettable.add_row(str(i), player_name, str(total_wickets))

        print("\n")
        console.print(wickettable)
        time.sleep(5)

    def show_balls_played(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name, SUM(balls_faced) as total_balls
            FROM match_data
            GROUP BY player_name
            ORDER BY total_balls DESC
            LIMIT 10
        """)
        ballsresults = cursor.fetchall()
        cursor.close()

        ballstable = Table(
            title="⚡ Top Balls Faced",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_blue",
            box=box.ROUNDED
        )
        ballstable.add_column("Rank", style="cyan", justify="center")
        ballstable.add_column("Player", style="bold green")
        ballstable.add_column("Balls Faced", style="bold magenta", justify="right")

        for i, (player_name, total_balls) in enumerate(ballsresults, start=1):
            ballstable.add_row(str(i), player_name, str(total_balls))

        print("\n")
        console.print(ballstable)
        time.sleep(5)

    def show_most_active(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name, COUNT(*) as matches_played
            FROM match_data
            GROUP BY player_name
            ORDER BY matches_played DESC
            LIMIT 10
        """)
        activeresults = cursor.fetchall()
        cursor.close()

        activetable = Table(
            title="🔥 Most Active Players",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_cyan",
            box=box.ROUNDED
        )
        activetable.add_column("Rank", style="cyan", justify="center")
        activetable.add_column("Player", style="bold green")
        activetable.add_column("Matches Played", style="bold magenta", justify="right")

        for i, (player_name, matches_played) in enumerate(activeresults, start=1):
            activetable.add_row(str(i), player_name, str(matches_played))

        print("\n")
        console.print(activetable)
        time.sleep(5)

    def show_high_scores(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name, MAX(runs) as highest_score
            FROM match_data
            GROUP BY player_name
            ORDER BY highest_score DESC
            LIMIT 10
        """)
        highscoreresults = cursor.fetchall()
        cursor.close()

        highscoretable = Table(
            title="🏏 Highest Individual Scores",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_green",
            box=box.ROUNDED
        )
        highscoretable.add_column("Rank", style="cyan", justify="center")
        highscoretable.add_column("Player", style="bold green")
        highscoretable.add_column("High Score", style="bold magenta", justify="right")

        for i, (player_name, highest_score) in enumerate(highscoreresults, start=1):
            highscoretable.add_row(str(i), player_name, str(highest_score))

        print("\n")
        console.print(highscoretable)
        time.sleep(5)

    def show_win_rate(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name,
                   SUM(CASE WHEN result = 'win' THEN 1 ELSE 0 END) AS wins,
                   COUNT(*) AS matches,
                   ROUND(SUM(CASE WHEN result = 'win' THEN 1 ELSE 0 END) / COUNT(*) * 100, 2) AS win_rate
            FROM match_data
            GROUP BY player_name
            HAVING matches > 0
            ORDER BY win_rate DESC
            LIMIT 10
        """)

        winrateresults = cursor.fetchall()
        cursor.close()

        winratetable = Table(
            title="🏆 Win Rate Leaderboard",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_magenta",
            box=box.ROUNDED
        )
        winratetable.add_column("Rank", style="cyan", justify="center")
        winratetable.add_column("Player", style="bold green")
        winratetable.add_column("Win Rate %", style="bold blue", justify="right")
        winratetable.add_column("Record (W/M)", style="bold magenta", justify="center")

        for i, (player_name, wins, matches, win_rate) in enumerate(winrateresults, start=1):
            winratetable.add_row(
                str(i),
                player_name,
                f"{win_rate}%",
                f"{wins}/{matches}"
            )

        print("\n")
        console.print(winratetable)
        time.sleep(5)

    def show_all_rounders(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT player_name,
                   SUM(runs) + SUM(wickets * 20) as all_rounder_score
            FROM match_data
            GROUP BY player_name
            ORDER BY all_rounder_score DESC
            LIMIT 10
        """)
        rounderresults = cursor.fetchall()
        cursor.close()

        roundertable = Table(
            title="💫 Top All-Rounders",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_green",
            box=box.ROUNDED
        )
        roundertable.add_column("Rank", style="cyan", justify="center")
        roundertable.add_column("Player", style="bold green")
        roundertable.add_column("All-Rounder Score", style="bold magenta", justify="right")

        for i, (player_name, score) in enumerate(rounderresults, start=1):
            roundertable.add_row(str(i), player_name, str(score))

        console.print(roundertable)

    def show_lifetime_run_rate(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT name,
                    lifetime_runs,
                    lifetime_balls_faced,
                    CASE 
                       WHEN lifetime_balls_faced > 0 
                       THEN ROUND((lifetime_runs * 6.0) / lifetime_balls_faced, 2)
                       ELSE 0.0
                    END as lifetime_run_rate
            FROM player_profile
            WHERE lifetime_balls_faced > 0
            ORDER BY lifetime_run_rate DESC
            LIMIT 10
        """)
        
        liferunresults = cursor.fetchall()
        cursor.close()

        liferuntable = Table(
            title="⚡ Lifetime Run Rate Leaderboard",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_blue",
            box=box.ROUNDED
        )
        liferuntable.add_column("Rank", style="cyan", justify="center")
        liferuntable.add_column("Player", style="bold green")
        liferuntable.add_column("Run Rate", style="bold blue", justify="right")
        liferuntable.add_column("Runs/Balls", style="bold magenta", justify="center")

        for i, (player_name, lifetime_runs, lifetime_balls_faced, lifetime_run_rate) in enumerate(liferunresults, start=1):
            liferuntable.add_row(
                str(i),
                player_name,
                str(lifetime_run_rate),
                f"{lifetime_runs}/{lifetime_balls_faced}"
            )

        print("\n")
        console.print(liferuntable)
        time.sleep(5)

    def show_lifetime_strike_rate(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT name,
                    lifetime_runs,
                    lifetime_balls_faced,
                    CASE
                       WHEN lifetime_balls_faced > 0
                       THEN ROUND((lifetime_runs * 100.0) / lifetime_balls_faced, 2)
                       ELSE 0.0
                   END as lifetime_strike_rate
            FROM player_profile
            WHERE lifetime_balls_faced > 0
            ORDER BY lifetime_strike_rate DESC
            LIMIT 10
        """)
        lifestrikeresults = cursor.fetchall()
        cursor.close()

        lifestriketable = Table(
            title="🚀 Lifetime Strike Rate Leaderboard",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_red",
            box=box.ROUNDED
        )
        lifestriketable.add_column("Rank", style="cyan", justify="center")
        lifestriketable.add_column("Player", style="bold green")
        lifestriketable.add_column("Strike Rate", style="bold blue", justify="right")
        lifestriketable.add_column("Runs/Balls", style="bold magenta", justify="center")

        for i, (player_name, lifetime_runs, lifetime_balls_faced, strike_rate) in enumerate(lifestrikeresults, start=1):
            lifestriketable.add_row(
                str(i),
                player_name,
                str(strike_rate),
                f"{lifetime_runs}/{lifetime_balls_faced}"
            )   

        print("\n")
        console.print(lifestriketable)
        time.sleep(5)

    def show_lifetime_economy_rate(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
                        SELECT name,
                        lifetime_runs_conceded,
                        lifetime_balls_bowled,
                        CASE
                        WHEN lifetime_balls_bowled > 0
                        THEN ROUND(lifetime_runs_conceded * 6.0 / lifetime_balls_bowled, 2)
                        ELSE 0.0
                    END as lifetime_economy_rate
                FROM player_profile
                WHERE lifetime_balls_bowled > 0
                ORDER BY lifetime_economy_rate ASC 
                LIMIT 10
            """)
        lifeecoresults = cursor.fetchall()
        cursor.close()

        lifeecotable = Table(
            title="🛡️ Lifetime Economy Rate Leaderboard",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_cyan",
            box=box.ROUNDED
        )
        lifeecotable.add_column("Rank", style="cyan", justify="center")
        lifeecotable.add_column("Player", style="bold green")
        lifeecotable.add_column("Economy Rate", style="bold blue", justify="right")
        lifeecotable.add_column("Runs Conceded/Balls Bowled", style="bold magenta", justify="center")

        for i, (player_name, lifetime_runs_conceded, lifetime_balls_bowled, economy_rate) in enumerate(lifeecoresults, start=1):
            lifeecotable.add_row(
                str(i),
                player_name,
                str(economy_rate),
                f"{lifetime_runs_conceded}/{lifetime_balls_bowled}"
            )
        
        print("\n")
        console.print(lifeecotable)
        time.sleep(5)

    def show_half_centuries(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT name, SUM(half_centuries) AS total_half_centuries
            FROM player_profile
            GROUP BY name
            ORDER BY total_half_centuries DESC
            LIMIT 10
        """)
        halfcentresults = cursor.fetchall()
        cursor.close()

        halfcenttable = Table(
            title="🏏 Half Centuries Leaderboard",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_cyan",
            box=box.ROUNDED
        )

        halfcenttable.add_column("Rank", style="cyan", justify="center")
        halfcenttable.add_column("Player", style="bold green")
        halfcenttable.add_column("Half Centuries", style="bold magenta", justify="right")

        for i, (player_name, total_half) in enumerate(halfcentresults, start=1):
            halfcenttable.add_row(
                str(i),
                player_name,
                str(total_half)
            )

        print("\n")
        console.print(halfcenttable)
        time.sleep(5)

    def show_centuries(self):
        lbcursor = self.db.get_db()
        cursor = lbcursor.cursor()
        cursor.execute("""
            SELECT name, SUM(centuries) AS total_centuries
            FROM player_profile
            GROUP BY name
            ORDER BY total_centuries DESC
            LIMIT 10
        """)
        centresults = cursor.fetchall()
        cursor.close()

        centtable = Table(
            title="🏆 Top 10 Centuries",
            show_header=True,
            header_style="bold yellow",
            border_style="bright_cyan",
            box=box.ROUNDED
        )

        centtable.add_column("Rank", style="cyan", justify="center")
        centtable.add_column("Player", style="bold green")
        centtable.add_column("Centuries", style="bold magenta", justify="right")

        for i, (player_name, total_centuries) in enumerate(centresults, start=1):
            centtable.add_row(
                str(i),
                player_name,
                str(total_centuries)
            )

        print("\n")
        console.print(centtable)
        time.sleep(5)
class SoundManager:
    def __init__(self):
        pygame.mixer.init()
        pygame.mixer.music.load("assets/bgmusic.wav")
        self.milestone = pygame.mixer.Sound("assets/cheer.wav")
        self.wicket = pygame.mixer.Sound("assets/out.wav")
        self.score = pygame.mixer.Sound("assets/score.wav")
        self.click = pygame.mixer.Sound("assets/click.wav")
        self.run = pygame.mixer.Sound("assets/run.wav")

    def play_bg(self):
        pygame.mixer.music.set_volume(0.2)
        pygame.mixer.music.play(-1)
    
    def stop_bg(self):
        pygame.mixer.music.stop()
    
    def play_click(self):
        self.click.set_volume(0.5)
        self.click.play()

    def play_milestone(self):
        self.milestone.set_volume(0.5)
        self.milestone.play()
    
    def play_wicket(self):
        self.wicket.set_volume(0.5)
        self.wicket.play()
    
    def play_score(self):
        self.score.set_volume(0.5)
        self.score.play()

    def play_run(self):
        self.run.set_volume(0.7)
        self.run.play()
def ask_commentary_setting():
    while True:
        try:
            choice = Prompt.ask(f"[green]Do you want commentary?[/green] [cyan](yes/no)[/cyan][green]:[/green] ").strip().lower()
            sounds.play_click()
            if choice == "yes":
                return True
            elif choice == "no":
                return False
            else:
                console.print(f"[red]Invalid input! Please enter yes or no.[/red]")
        except KeyboardInterrupt:
            console.print(f"[bold red]Game interrupted! try again[/bold red]")
def ask_difficulty_setting():
    while True:
        try:
            difficulty = Prompt.ask("\n[green]Choose difficulty[/green] [cyan](easy / medium / hard)[/cyan][green]:[/green] ").strip().lower()
            sounds.play_click()
            if difficulty in ("easy", "medium", "hard"):
                console.print(f"\n[green]Difficulty set to[/green] [bright blue]{difficulty}[/bright blue][green].[/green]")
                return difficulty
            else:
                console.print(f"\n[red]Invalid choice. Try again.[/red]")
        except KeyboardInterrupt:
            console.print(f"\n[bold red]Game interrupted! try again[/bold red]")
def login_menu(db_handler, name):
    while True:
        sounds.play_click()

        print("\n--- Login Menu ---")
        print("1. Start Game")
        print("2. Change Password")
        print("3. Logout")
        choice = input("Choose: ").strip()

        sounds.play_click()
        
        if choice == "1":
            return True
        elif choice == "2":
            old_pass = pwinput.pwinput("Enter current password: ").strip()
            if old_pass == db_handler.get_password(name):
                new_pass = pwinput.pwinput("Enter new password: ").strip()
                confirm = pwinput.pwinput("Confirm new password: ").strip()
                if new_pass == confirm:
                    db_handler.update_password(name, new_pass)
                    print("Password updated successfully.")
                else:
                    print("Passwords do not match.")
            else:
                print("Incorrect current password.")
        elif choice == "3":
            return False
        else:
            print("Invalid choice.")
def player_login(db_handler):
    global game_var
    while True:
        try:
            name = Prompt.ask(f"\n[yellow]Enter your name[/yellow] [cyan](max 30 characters)[/cyan][green]:[/green] ").strip().title()
            sounds.play_click()
            if not name.replace(" ", "").isalpha():
                console.print(f"\n[red]Invalid name. Only letters and spaces allowed.[/red]")
                continue
            if len(name) == 0:
                console.print(f"\n[red]Name cannot be empty.[/red]")
                continue
            if len(name) > 30:
                print(f"\n[red]Name too long![/red] [blue]({len(name)}/30 characters)[/blue]")
                continue

            if db_handler.player_exists(name):
                console.print(f"\n[bold yellow]Welcome back,[/bold yellow] [blue]{name}[/blue][bold yellow]![/bold yellow]")
                stored_password = db_handler.get_password(name)
                authenticated = False
                console.print(f"\n[bold green italic]You have 3 attempts to enter your password.[/bold green italic]")
                i=0

                for attempt in range(3):
                    i+=1
                    console.print(f"\n[green]Enter your password:[/green] ", end="")
                    entered_pass = pwinput.pwinput("").strip()
                    if entered_pass == stored_password:
                        authenticated = True
                        break
                    elif i < 3:
                        console.print(f"\n[red]Incorrect password. You have[/red] [blue]{3 - i}[/blue] [red]attempts left.[/red]")
                if not authenticated:
                    console.print("[bold red]Too many failed attempts. Exiting login.[/bold red]")
                    return None
                
                console.print(f"[magenta italic]Logging in as[/magenta italic] [blue]{name}[/blue[magenta italic]...[/magenta italic]")
                time.sleep(2)
                reset = Prompt.ask(f"\n[green italic]Do you want to reset your stats?[/green italic] [yellow](yes/no)[/yellow][green italic]:[/green italic] ").strip().lower()
                sounds.play_click()
                if reset == "yes":
                    console.print("[cyan italic]Resetting profile and match history...[/cyan italic]")
                    time.sleep(2)
                    profile = {
                        "Lifetime runs": 0,
                        "Lifetime wickets": 0,
                        "Total Matches Played": 0,
                        "Total Wins": 0,
                        "Total Losses": 0,
                        "Total Draws": 0,
                        "Average Runs per Match": 0,
                        "Lifetime Balls Faced": 0,
                        "Lifetime Balls Bowled": 0,
                        "Lifetime Runs Conceded": 0,
                        "Centuries": 0,
                        "Half Centuries": 0,
                    }
                    db_handler.save_profile(profile, name, stored_password)
                    
                    db = db_handler.get_db()
                    cursor = db.cursor()
                    cursor.execute("DELETE FROM match_data WHERE player_name = %s", (name,))
                    db.commit()
                    cursor.close()
                    
                    console.print("\n[green]Profile and match history reset.[/green]")
                return name
            else:
                console.print(f"[yellow]Profile for[/yellow] [blue]{name}[/blue] [yellow]does not exist. Creating a new profile...[/yellow]")
                time.sleep(2)
                while True:
                    console.print(f"\n[green]Set a password (min 3 characters):[/green] ", end="")
                    password = pwinput.pwinput("").strip()
                    if len(password) < 3:
                        console.print(f"[red]Password too short![/red] [blue]({len(password)}/3 characters)[/blue]")
                        continue
                    console.print(f"[green]Confirm password:[/green] ", end="")
                    confirm = pwinput.pwinput("").strip()
                    if password != confirm:
                        print("Passwords do not match. Try again.")
                    else:
                        break
                console.print(f"\n[cyan italic]Creating profile for[/cyan italic] [blue]{name}[/blue][cyan italic]...[/cyan italic]")
                time.sleep(3)
                console.print(f"[bold green]Profile created successfully! Welcome,[/bold green] [blue]{name}[/blue][bold green]![/bold green]")
                profile = {
                    "Lifetime runs": 0,
                    "Lifetime wickets": 0,
                    "Total Matches Played": 0,
                    "Total Wins": 0,
                    "Total Losses": 0,
                    "Total Draws": 0,
                    "Average Runs per Match": 0,
                    "Lifetime Balls Faced": 0,
                    "Lifetime Balls Bowled": 0,
                    "Lifetime Runs Conceded": 0,
                    "Centuries": 0,
                    "Half Centuries": 0,
                }
                db_handler.save_profile(profile, name, password)

            return name
        except KeyboardInterrupt:
            console.print(f"\n[bold red]Game interrupted! try again[/bold red]")
def main(game_manager):
    global game_var, manager, name, gameplay, commentator, display
    while True:
        choice = game_manager.get_choice()

        sounds.play_click()

        if choice == "tutorial":
            print("\nTutorial: \nChoose Odd or even:")
            print("  :You will be playing against an AI opponent")
            print("  :AI will choose the opposite of what you will choose \n  :If the sum of the numbers corresponds to what you choose then you'll get the opportunity to choose to bat or bowl")
            print("  :There will be two innings from bat to bowl or bowl to bat")
            print("  :Using the same numbers will drop a wicket \n  :Each number will add to your total runs")
            print("  :At the end, who has the highest amount of runs wins!")
            print("  :Note that in case both Player and AI choose 0, it will lead to the 1 or 2 rule-which means that only 1 or 2 is allowed - but they don't count as runs!")
            print("  :In case of tie, a super over will occur! super over means only 1 over with just 6 balls!")
            print("\nHave a great time playing!!!")

        elif choice == "play":
            game_var['total_matches_played'] += 1
            game_manager.reset_game(game_var)
            toss_result = gameplay.toss(game_var)
            over = gameplay.match_over(game_var)
            game_var['balls'] = game_var['over'] * balls_per_over
            gameplay.first_in(game_var)
            gameplay.second_in(game_var)
            game_manager.result(game_var)

        elif choice == "quit":
            print("Thank you for playing")
            pygame.mixer.music.stop()
            db_handler.close_db()
            break

        elif choice == "save":
            print("Saving game state...")
            time.sleep(3)
            game_manager.handle_save_game_and_profile(game_var)

        elif choice == "load":
            print("Loading save game state...")
            time.sleep(3)
            loaded_game = manager.load_game_from_file()
            if loaded_game:
                game_var.update(loaded_game)
                print("Game state loaded successfully!")
                print(f"\nWelcome back {name}!!")
                print(f"Total matches played: {game_var['total_matches_played']}")
                print(f"Lifetime runs: {game_var['lifetime_runs']}")
                game_var['balls'] = game_var['over'] * balls_per_over

        elif choice == "profile":
            print("Loading profile...")
            time.sleep(2)
            game_manager.handle_profile_display_and_save(game_var)

        elif choice == "leaderboard":
            print("Attempting to open leaderboard...")
            time.sleep(2)
            leaderboard_manager = LeaderboardManager(db_handler)
            leaderboard_manager.show_menu()

        elif choice == "switch":
            print("Attempting to switch...")
            time.sleep(2)
            new_name = player_login(db_handler)
            if not new_name:
                print("Login failed. Staying with current player.")
                continue
            
            name = new_name
            
            game_var.update({
                "lifetime_runs": 0,
                "lifetime_wickets": 0,
                "total_matches_played": 0,
                "total_wins": 0,
                "total_losses": 0,
                "total_draws": 0,
                "average_runs_per_match": 0,
                "lifetime_balls_faced": 0,   
                "lifetime_balls_bowled": 0,
                "lifetime_runs_conceded": 0,
                "centuries": 0,
                "half_centuries": 0,
            })

            manager = GameDataManager(folder_path_for_json, folder_path_for_excel, name)

            if manager.file_exists():
                print("A pre-existing save with this name was found.")
                while True:
                    try:
                        autosave_load_choice = input("Would you like to load the save file? (yes/no): ").strip().lower()
                        if autosave_load_choice == "yes":
                            print("Loading save game state...")
                            time.sleep(3)
                            loaded_game = manager.load_game_from_file()
                            if loaded_game:
                                game_var.update(loaded_game)
                                print(f"\nWelcome back {name}!!")
                                print(f"Total matches played: {game_var['total_matches_played']}")
                                print(f"Lifetime runs: {game_var['lifetime_runs']}")
                            break
                        elif autosave_load_choice == "no":
                            commentary_enabled = ask_commentary_setting()
                            difficulty = ask_difficulty_setting()
                            game_var["commentary_enabled"] = commentary_enabled
                            game_var["difficulty"] = difficulty
                            profile_data = db_handler.load_profile(name)
                            if profile_data:
                                print("Profile stats loaded from database.")
                                game_var.update(profile_data)
                            break
                        else:
                            print("Enter a valid choice!")
                    except KeyboardInterrupt:
                        print("Game was interrupted, try again")
            else:
                commentary_enabled = ask_commentary_setting()
                difficulty = ask_difficulty_setting()
                game_var["commentary_enabled"] = commentary_enabled
                game_var["difficulty"] = difficulty
                profile_data = db_handler.load_profile(name)
                if profile_data:
                    print("Profile stats loaded from database.")
                    game_var.update(profile_data)

            commentator = Commentator(game_var.get("commentary_enabled", True))
            gameplay = GamePlay(name, min_choice, max_choice, score_alignment, game_var.get('difficulty', 'medium'), commentator)
            display = DisplayManager()
            game_manager = GameManager(gameplay, display, db_handler, manager, name)
            gameplay.intro()

        elif choice == "settings":
            print("Attempting to open the settings menu...")
            time.sleep(2)
            game_manager.show_settings_menu(game_var)

        else:
            print("Invalid Input")
def animated_print(*args, **kwargs):
    sep = kwargs.get("sep", " ")
    end = kwargs.get("end", "\n")
    text = sep.join(str(a) for a in args)

    for ch in text:
        sys.stdout.write(ch)
        sys.stdout.flush()
        time.sleep(0.02)

    sys.stdout.write(end)
    sys.stdout.flush()
if __name__ == "__main__":
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table
    from rich.prompt import Prompt
    from rich import box
    import random
    import pwinput
    import os
    import builtins
    import time
    import sys
    import mysql.connector
    from mysql.connector import Error
    import json
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    os.environ["PYGAME_HIDE_SUPPORT_PROMPT"] = "1"
    import pygame

    console = Console()

    sounds = SoundManager()
    sounds.play_bg()
    
    base_dir = os.path.join(os.path.expanduser("~"), "Documents", "Hand_cricket_saves")
    folder_path_for_json = os.path.join(base_dir, "json_saves")
    folder_path_for_excel = os.path.join(base_dir, "excel_saves")

    os.makedirs(folder_path_for_json, exist_ok=True)
    os.makedirs(folder_path_for_excel, exist_ok=True)

    original_print = print
    builtins.print = animated_print
    
    difficulty = ""

    while True:
        try:
            console.print(Panel.fit("[bold yellow]Enter your MySQL root password to connect to the database:[/bold yellow]", border_style="bright_blue"))
            db_password = pwinput.pwinput("")
            break
        except KeyboardInterrupt:
            print("\nGame was interrupted, Try again")
            
    DB_CONFIG = {
        "host": "localhost",
        "user": "root",
        "password": db_password,
        "database": "handcricket"
    }

    _DB = None

    db_handler = HandCricketDB(DB_CONFIG)

    db_handler.create_database()

    db_handler.init_db()

    game_var = {
        'player_runs_1stinn': 0,
        'ai_runs_1stinn': 0,
        'player_runs_2ndinn': 0,
        'ai_runs_2ndinn': 0,
        'toss_result': "",
        'match_summary': {},
        'player_profile': {},
        'balls_played_first_inn': 0,
        'balls_played_sec_inn': 0,
        'total_matches_played': 0,
        'total_wins': 0,
        'total_losses': 0,
        'lifetime_wickets': 0,
        'average_runs_per_match': 0,
        'current_match_wickets': 0,
        'lifetime_runs': 0,
        'over': 0,
        'total_draws': 0,
        'balls': 0,
        'balls_played_by_player': 0,
        'difficulty': 'medium',
        'commentary_enabled': True,
        'lifetime_balls_faced': 0,
        'lifetime_balls_bowled': 0,
        'lifetime_runs_conceded': 0,
        'centuries': 0,
        'half_centuries': 0,
        'player_runs_conceded_1stinn': 0, 
        'player_runs_conceded_2ndinn': 0, 
        'player_balls_bowled_1stinn': 0,  
        'player_balls_bowled_2ndinn': 0,
        'match_result': '',
    }

    name = player_login(db_handler)
    if not name:
        print("No valid player name provided. Exiting game.")
        exit()

    min_choice = 0
    max_choice = 10
    balls_per_over = 6
    score_alignment = 80

    manager = GameDataManager(folder_path_for_json, folder_path_for_excel, name)
    
    if manager.file_exists():
        console.print(Panel.fit("[bold yellow]A pre-existing save with this name was found.[/bold yellow]", border_style="bright_blue"))
        while True:
            try:
                autosave_load_choice = Prompt.ask("[green]Would you like to load the save file?[/green] (yes/no): ").strip().lower()
                sounds.play_click()
                if autosave_load_choice == "yes":
                    console.print("[dark_grey]Loading save game state...[/dark_grey]")
                    time.sleep(3)
                    loaded_game = manager.load_game_from_file()
                    if loaded_game:
                        game_var.update(loaded_game)
                        console.print(f"\n[bold green]Welcome back {name}!![/bold green]")
                        console.print(f"[bold green]Total matches played:[/bold green] {game_var['total_matches_played']}")
                        console.print(f"[bold green]Lifetime runs:[/bold green] {game_var['lifetime_runs']}")
                        break
                elif autosave_load_choice == "no":
                    commentary_enabled = ask_commentary_setting()
                    difficulty = ask_difficulty_setting()
                    game_var["commentary_enabled"] = commentary_enabled
                    game_var["difficulty"] = difficulty
                    profile_data = db_handler.load_profile(name)
                    if profile_data:
                        print("Profile stats loaded from database.")
                        game_var.update(profile_data)
                    break
                else:
                    print("Enter a valid choice!")
            except KeyboardInterrupt:
                print("Game was interrupted, try again")
    else:
        commentary_enabled = ask_commentary_setting()
        difficulty = ask_difficulty_setting()
        game_var["commentary_enabled"] = commentary_enabled
        game_var["difficulty"] = difficulty
        profile_data = db_handler.load_profile(name)
        if profile_data:
            print("Profile stats loaded from database.")
            game_var.update(profile_data)
    
    commentator = Commentator(game_var.get("commentary_enabled", True))
    gameplay = GamePlay(name, min_choice, max_choice, score_alignment, game_var.get('difficulty', 'medium'), commentator)
    display = DisplayManager()
    game_manager = GameManager(gameplay, display, db_handler, manager, name)
    gameplay.intro()
    time.sleep(1)
    main(game_manager)